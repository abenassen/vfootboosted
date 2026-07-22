"""Compare our classic voto puro against an external Italian fantasy-vote provider
(fantacalcio.it xlsx export: one file per giornata, sheets Fantacalcio/Statistico/
Italia; column `Voto` is the base pagella, bonus/malus are separate columns).

This is a QUANTITATIVE diagnostic, not a calibration step. It measures, on the same
matched player-matchday set:
  * distribution spread (is the external vote 'too flat'?);
  * agreement between the two base votes;
  * how much each base vote tracks SofaScore rating (performance signal);
  * how much each base vote is driven by goals scored (the 'too correlated with
    bonus' hypothesis — the base vote should be largely independent of the goal,
    which is a separate bonus).

    python manage.py compare_external_votes --sheet Fantacalcio
"""

from __future__ import annotations

import glob
import json
import math
import re
from collections import Counter, defaultdict

import openpyxl
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError

from realdata.models import Match, Player, PlayerTeamStint
from realdata.services.identity import norm_name
from vfoot.services.classic_rating import build_reference, voto_puro_for_match

DEFAULT_DIR = (str(Path(settings.VFOOT_DATA_DIR) / "data_fantacalcio" / "2025-2026"))
DEFAULT_CACHE = str(Path(settings.VFOOT_DATA_DIR) / "historical-data" / "serie-a" / "sofascore" / "cache")
EXTERNAL_TEAMS = {
    "atalanta", "bologna", "cagliari", "como", "cremonese", "fiorentina", "genoa",
    "verona", "hellas verona", "inter", "juventus", "lazio", "lecce", "milan",
    "napoli", "parma", "pisa", "roma", "sassuolo", "torino", "udinese",
}


def _pearson(xs, ys):
    n = len(xs)
    if n < 2:
        return float("nan")
    mx, my = sum(xs) / n, sum(ys) / n
    cov = sum((a - mx) * (b - my) for a, b in zip(xs, ys))
    sx = math.sqrt(sum((a - mx) ** 2 for a in xs))
    sy = math.sqrt(sum((b - my) ** 2 for b in ys))
    return cov / (sx * sy) if sx and sy else float("nan")


def _parse_voto(v):
    if v is None:
        return None
    m = re.search(r"(\d+(?:[.,]\d+)?)", str(v))
    return float(m.group(1).replace(",", ".")) if m else None


def _club_key(name):
    fillers = {"ac", "as", "fc", "ssc", "us", "ss", "hellas", "calcio"}
    return " ".join(t for t in norm_name(name).split() if t not in fillers)


class Command(BaseCommand):
    help = "Compare classic voto puro vs an external provider's base vote."

    def add_arguments(self, parser):
        parser.add_argument("--competition-season", type=int, default=2)
        parser.add_argument("--dir", default=DEFAULT_DIR)
        parser.add_argument("--sheet", default="Fantacalcio",
                            help="Fantacalcio | Statistico | Italia")
        parser.add_argument("--cache-dir", default=DEFAULT_CACHE)

    # -- external parsing ------------------------------------------------

    def _parse_file(self, path, sheet):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            return []
        ws = wb[sheet]
        out, team = [], None
        for r in ws.iter_rows(values_only=True):
            c0 = r[0]
            if isinstance(c0, str) and norm_name(c0) in EXTERNAL_TEAMS:
                team = c0
                continue
            if not isinstance(c0, (int, float)) or c0 == "Cod.":
                continue
            out.append({"team": team, "ruolo": r[1], "nome": r[2],
                        "voto": _parse_voto(r[3]),
                        "gf": r[4] or 0, "ass": r[12] or 0})
        return out

    # -- our-side indices ------------------------------------------------

    def _our_team_index(self, cs_id):
        """external club-key -> our team name (best match)."""
        teams = (PlayerTeamStint.objects
                 .filter(team_season__competition_season_id=cs_id)
                 .values_list("team_season__team__name", flat=True).distinct())
        ours = {_club_key(t): t for t in set(teams)}
        return ours

    def _our_player_index(self, cs_id):
        """(our_team_name, surname_norm) -> [player_id]; surname = last token."""
        idx = defaultdict(list)
        rows = (PlayerTeamStint.objects
                .filter(team_season__competition_season_id=cs_id, end_date__isnull=True)
                .values_list("player_id", "team_season__team__name",
                             "player__full_name", "player__short_name"))
        for pid, team, full, short in rows:
            keys = set()
            for nm in (full, short):
                toks = norm_name(nm).split()
                if toks:
                    keys.add(toks[-1])
            for k in keys:
                idx[(team, k)].append(pid)
        return idx

    def _ratings(self, ext_id, cache_dir):
        try:
            d = json.load(open(f"{cache_dir}/api_v1_event_{ext_id}_lineups.json"))
        except (FileNotFoundError, ValueError):
            return {}
        out = {}
        for side in ("home", "away"):
            for pl in d.get(side, {}).get("players", []):
                pid = (pl.get("player") or {}).get("id")
                st = pl.get("statistics") or {}
                if pid is not None and st.get("rating"):
                    out[str(pid)] = st["rating"]
        return out

    # -- main ------------------------------------------------------------

    def handle(self, *args, **opts):
        cs_id = opts["competition_season"]
        files = sorted(glob.glob(f"{opts['dir']}/*.xlsx"))
        if not files:
            raise CommandError(f"No .xlsx in {opts['dir']}")

        ref = build_reference(cs_id)
        team_map = self._our_team_index(cs_id)
        pidx = self._our_player_index(cs_id)
        sofa_ext = dict(Player.objects.filter(external_source="sofascore")
                        .values_list("id", "external_id"))

        # our voto puro + sofa rating per (matchday, player_id)
        our_vote, rating_by = {}, {}
        for m in Match.objects.filter(competition_season_id=cs_id):
            for row in voto_puro_for_match(m, ref):
                if row["rated"]:
                    our_vote[(m.matchday, row["player_id"])] = row["voto_puro"]
            r = self._ratings(m.external_id, opts["cache_dir"]) if m.external_id else {}
            for pid, extid in sofa_ext.items():
                if extid in r:
                    rating_by[(m.matchday, pid)] = r[extid]

        # walk external files
        gd_re = re.compile(r"Giornata_(\d+)")
        ext_rows, unmatched_team, unmatched_name = [], Counter(), 0
        ext_all_voti = []
        for f in files:
            mm = gd_re.search(f)
            if not mm:
                continue
            gd = int(mm.group(1))
            for e in self._parse_file(f, opts["sheet"]):
                if e["voto"] is not None:
                    ext_all_voti.append(e["voto"])
                our_team = team_map.get(_club_key(e["team"] or ""))
                if not our_team:
                    unmatched_team[e["team"]] += 1
                    continue
                surn = norm_name(e["nome"]).split()[-1] if e["nome"] else ""
                cands = pidx.get((our_team, surn), [])
                if len(cands) != 1:
                    unmatched_name += 1
                    continue
                pid = cands[0]
                ext_rows.append({"gd": gd, "pid": pid, "voto": e["voto"],
                                 "gf": e["gf"], "ass": e["ass"],
                                 "our": our_vote.get((gd, pid)),
                                 "rating": rating_by.get((gd, pid))})

        self._report(opts["sheet"], ext_all_voti, ext_rows,
                     unmatched_team, unmatched_name)

    def _report(self, sheet, ext_all, rows, unmatched_team, unmatched_name):
        w = self.stdout.write
        w(f"=== external '{sheet}' vs our voto puro ===")
        # overall external distribution (all parsed, rated)
        n = len(ext_all)
        mean = sum(ext_all) / n
        std = math.sqrt(sum((x - mean) ** 2 for x in ext_all) / n)
        w(f"External base vote (all {n}): mean={mean:.2f} std={std:.2f} "
          f"min={min(ext_all)} max={max(ext_all)}")

        paired = [r for r in rows if r["our"] is not None and r["voto"] is not None]
        w(f"\nMatched player-matchdays: {len(rows)} "
          f"(both rated: {len(paired)}); unmatched names: {unmatched_name}; "
          f"unmatched teams: {dict(unmatched_team)}")
        if not paired:
            return
        ext = [r["voto"] for r in paired]
        our = [r["our"] for r in paired]

        def stats(v):
            mu = sum(v) / len(v)
            sd = math.sqrt(sum((x - mu) ** 2 for x in v) / len(v))
            return mu, sd
        em, es = stats(ext)
        om, os_ = stats(our)
        w(f"\nOn the {len(paired)} both-rated pairs:")
        w(f"  external: mean={em:.2f} std={es:.2f}")
        w(f"  ours    : mean={om:.2f} std={os_:.2f}   (higher std = more spread)")
        w(f"  corr(external, ours) = {_pearson(ext, our):.3f}")

        # vs SofaScore rating (performance signal) on the subset with rating
        wr = [(r["voto"], r["our"], r["rating"]) for r in paired
              if r["rating"] is not None]
        if wr:
            ev = [a for a, _, _ in wr]
            ov = [b for _, b, _ in wr]
            rv = [c for _, _, c in wr]
            w(f"\nvs SofaScore rating (n={len(wr)}):")
            w(f"  corr(external, rating) = {_pearson(ev, rv):.3f}")
            w(f"  corr(ours,     rating) = {_pearson(ov, rv):.3f}")

        # decisive test: strip the goal confound — among NON-SCORERS, which base
        # vote tracks the pro rating better? (pure performance, no goal signal)
        wr0 = [(r["voto"], r["our"], r["rating"]) for r in paired
               if r["rating"] is not None and r["gf"] == 0]
        if wr0:
            ev = [a for a, _, _ in wr0]
            ov = [b for _, b, _ in wr0]
            rv = [c for _, _, c in wr0]
            w(f"\nvs rating, NON-SCORERS only (n={len(wr0)}) — goal confound removed:")
            w(f"  corr(external, rating) = {_pearson(ev, rv):.3f}")
            w(f"  corr(ours,     rating) = {_pearson(ov, rv):.3f}")

        # dependence on goals scored (is the base vote a goal proxy?)
        gf = [min(r["gf"], 2) for r in paired]
        w(f"\nGoal dependence  corr(vote, goals):")
        w(f"  external = {_pearson(ext, gf):.3f}   ours = {_pearson(our, gf):.3f}")
        w("  mean base vote by goals scored (0 / 1 / 2+):")
        for g in (0, 1, 2):
            eg = [r["voto"] for r in paired if min(r["gf"], 2) == g]
            og = [r["our"] for r in paired if min(r["gf"], 2) == g]
            if eg:
                w(f"    {g}{'+ ' if g==2 else '  '}goals (n={len(eg):>4}): "
                  f"external={sum(eg)/len(eg):.2f}  ours={sum(og)/len(og):.2f}")

        # side-by-side histograms
        w("\nHistogram (external | ours):")
        he, ho = Counter(round(x*2)/2 for x in ext), Counter(round(x*2)/2 for x in our)
        keys = sorted(set(he) | set(ho))
        pe, po = max(he.values()), max(ho.values())
        for k in keys:
            be = "█" * round(18 * he.get(k, 0) / pe)
            bo = "█" * round(18 * ho.get(k, 0) / po)
            w(f"  {k:4.1f} | {be:<18} {he.get(k,0):>4}  |  {bo:<18} {ho.get(k,0):>4}")
