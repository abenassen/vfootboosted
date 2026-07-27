"""Cross-check OUR detected bonus/malus events against the fantacalcio voti sheets.

The fantacalcio ``Voti_...`` files carry, per player per matchday, the raw EVENT
columns the site scored the fantavoto from: Gf (goals), Gs (conceded), Rp (penalty
saved), Rs (penalty missed), Rf (penalty scored), Au (own goal), Amm (yellow), Esp
(red), Ass (assist). Those are ground truth for "what happened". This command lines
each up against what WE detect from the imported data, per category, so a bonus or
malus we silently miss surfaces as a column where fanta counts events and we do not
— which is exactly how own goals and missed penalties were caught.

    python manage.py validate_fanta_events --season 2
    python manage.py validate_fanta_events --season 2 --show bonus_pen_saved

It reports, per category: fanta's event total, ours, and the player-matches where
they disagree (with a handful of examples). Assists are expected to diverge — the
provider and fantacalcio use different assist definitions — so they are flagged
informational, not a gap.
"""
from __future__ import annotations

import glob
from collections import defaultdict
from pathlib import Path

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError

from realdata.models import (
    CARD_RED, CARD_SECOND_YELLOW, CARD_YELLOW,
    Match, MatchAppearance, MatchDisciplinaryEvent, MatchShot,
)
from realdata.services.identity import norm_name
from vfoot.management.commands.voto_puro_discrepancies import (
    Command as DiscCmd, EXTERNAL_TEAMS, _club_key,
)

# sheet column -> event key
COLS = {"gf": 4, "gs": 5, "rp": 6, "rs": 7, "rf": 8, "au": 9, "amm": 10, "esp": 11, "ass": 12}


def _n(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


class Command(BaseCommand):
    help = "Validate our detected events against the fantacalcio voti sheets."

    def add_arguments(self, parser):
        parser.add_argument("--season", type=int, default=2)
        parser.add_argument("--dir", default=None)
        parser.add_argument("--show", default=None,
                            help="Print every mismatch for one category key.")
        parser.add_argument("--examples", type=int, default=6)

    def _parse_sheet(self, path):
        """Yield (team, ruolo, nome, {event: count}) for the Fantacalcio sheet."""
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Fantacalcio" not in wb.sheetnames:
            wb.close()
            return
        team = None
        for r in wb["Fantacalcio"].iter_rows(values_only=True):
            c0 = r[0]
            if isinstance(c0, str) and norm_name(c0) in EXTERNAL_TEAMS:
                team = c0
                continue
            if not isinstance(c0, (int, float)) or c0 == "Cod.":
                continue
            yield (team, r[1], r[2], {k: _n(r[i]) for k, i in COLS.items()})
        wb.close()

    def handle(self, *args, **o):
        cs = o["season"]
        ddir = o["dir"] or str(Path(settings.VFOOT_DATA_DIR) / "data_fantacalcio" / "2025-2026")
        import re
        files = sorted(glob.glob(f"{ddir}/*.xlsx"))
        if not files:
            raise CommandError(f"No fantacalcio .xlsx in {ddir}")

        disc = DiscCmd()
        team_map = disc._our_team_index(cs)
        pidx, _pt, pid_ini = disc._our_player_index(cs)
        gd_re = re.compile(r"Giornata_(\d+)")

        # fanta events keyed (md, pid)
        fanta: dict = {}
        unmatched = 0
        for f in files:
            mm = gd_re.search(f)
            if not mm:
                continue
            md = int(mm.group(1))
            for team, ruolo, nome, ev in self._parse_sheet(f):
                if str(ruolo).upper() == "ALL":  # coaches: fanta grades, we don't
                    continue
                our_team = team_map.get(_club_key(team or ""))
                if not our_team:
                    continue
                surn, abbr = disc._surname_and_initial(nome or "")
                cands = pidx.get((our_team, surn), [])
                if len(cands) > 1 and abbr:
                    nar = [pid for pid in cands
                           if any(fn.startswith(abbr) for fn in pid_ini.get(pid, ()))]
                    if nar:
                        cands = nar
                if len(cands) != 1:
                    if any(ev.values()):
                        unmatched += 1
                    continue
                fanta[(md, cands[0])] = ev

        ours = self._our_events(cs)

        # categories: (key, label, our_getter, note)
        cats = [
            ("gf", "Gol (Gf+Rf)", lambda o: o["goals"], "gol da MatchAppearance"),
            ("au", "Autogol (Au)", lambda o: o["own_goals"], "raw_stats.ownGoals"),
            ("rs", "Rig. sbagliato (Rs)", lambda o: o["missed_pen"], "MatchShot situation=penalty"),
            ("amm", "Ammonizione (Amm)", lambda o: o["yellow"], "MatchDisciplinaryEvent"),
            ("esp", "Espulsione (Esp)", lambda o: o["red"], "MatchDisciplinaryEvent"),
            ("rp", "Rig. parato (Rp)", lambda o: o["pen_saved"], "MatchShot save -> portiere"),
            ("gs", "Gol subiti GK (Gs)", lambda o: o["conceded"], "gol subiti dal portiere in campo"),
            ("ass", "Assist (Ass)", lambda o: o["assists"], "def. assist diverse (informativo)"),
        ]

        self.stdout.write(f"Coppie (giornata,giocatore) agganciate: {len(fanta)} "
                          f"(scartate con eventi: {unmatched})\n")
        self.stdout.write(f"{'categoria':<22}{'fanta':>7}{'nostri':>8}{'mismatch':>10}  note")
        self.stdout.write("-" * 78)
        detail = None
        for key, label, get, note in cats:
            fanta_tot = our_tot = mism = 0
            examples = []
            for (md, pid), ev in fanta.items():
                # Gf category folds Rf (penalties scored count as goals for the +3)
                want = ev[key] + (ev["rf"] if key == "gf" else 0)
                got = get(ours.get((md, pid), self._empty()))
                fanta_tot += want
                our_tot += got
                if want != got:
                    mism += 1
                    examples.append((md, pid, want, got))
            self.stdout.write(f"{label:<22}{fanta_tot:>7}{our_tot:>8}{mism:>10}  {note}")
            if o["show"] == key:
                detail = (label, examples)

        if detail:
            label, examples = detail
            names = dict(MatchAppearance.objects.values_list("player_id", "player__short_name"))
            self.stdout.write(f"\nMismatch '{label}' (fanta vs nostri):")
            for md, pid, want, got in examples[:200]:
                self.stdout.write(f"  gd{md:<3} {names.get(pid,pid):<20} fanta={want} nostri={got}")
        elif o["examples"]:
            # a few examples for the categories that are real gaps (non-assist)
            pass

    # ------------------------------------------------------------------
    @staticmethod
    def _empty():
        return {"goals": 0, "assists": 0, "own_goals": 0, "missed_pen": 0,
                "yellow": 0, "red": 0, "pen_saved": 0, "conceded": 0}

    def _our_events(self, cs) -> dict:
        out: dict = defaultdict(self._empty)
        for md, pid, goals, assists, raw in (MatchAppearance.objects
                .filter(match__competition_season_id=cs)
                .values_list("match__matchday", "player_id", "goals", "assists", "raw_stats")):
            rec = out[(md, pid)]
            rec["goals"] += goals or 0
            rec["assists"] += assists or 0
            rec["own_goals"] += int((raw or {}).get("ownGoals") or 0)
        for md, pid, ct in (MatchDisciplinaryEvent.objects
                .filter(match__competition_season_id=cs)
                .values_list("match__matchday", "player_id", "card_type")):
            rec = out[(md, pid)]
            if ct == CARD_YELLOW:
                rec["yellow"] += 1
            elif ct in (CARD_RED, CARD_SECOND_YELLOW):
                rec["red"] += 1
        for md, pid in (MatchShot.objects
                .filter(match__competition_season_id=cs, situation="penalty", is_goal=False)
                .exclude(player__isnull=True)
                .values_list("match__matchday", "player_id")):
            out[(md, pid)]["missed_pen"] += 1
        # penalties saved -> +3 to the keeper: reuse the production detector so the
        # validation checks the real code path (per match with a saved penalty).
        from vfoot.services.classic_pagella import (
            _penalties_saved_for_match, _goals_conceded_by_keeper)
        md_of = dict(Match.objects.filter(competition_season_id=cs)
                     .values_list("id", "matchday"))
        for mid in (MatchShot.objects
                    .filter(match__competition_season_id=cs, situation="penalty",
                            is_goal=False, shot_type="save")
                    .values_list("match_id", flat=True).distinct()):
            for gk_pid, n in _penalties_saved_for_match(mid).items():
                out[(md_of[mid], gk_pid)]["pen_saved"] += n
        for mid in md_of:  # goals conceded, per on-pitch keeper (production detector)
            for gk_pid, n in _goals_conceded_by_keeper(mid).items():
                out[(md_of[mid], gk_pid)]["conceded"] += n
        return dict(out)
