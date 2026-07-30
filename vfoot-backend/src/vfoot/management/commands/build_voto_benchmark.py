"""Benchmark globale del voto puro contro i due voti di fantacalcio.it.

    python manage.py build_voto_benchmark
    python manage.py build_voto_benchmark --matchdays 1-5 --out /tmp/bench

Genera una cartella di pagine HTML statiche e autoconsistenti (nessuna
dipendenza esterna, si aprono con un doppio click):

    index.html        il quadro d'insieme — accordo, distribuzioni, matrice
                      voto-a-voto, per ruolo, per giornata, le divergenze
                      piu' grosse;
    giornata-NN.html  una pagina per giornata, impaginata come il foglio
                      fantacalcio da cui vengono i dati (blocchi per squadra,
                      stesso ordine): Redazione | Statistico | NOSTRO | Δ.
                      Cliccando una riga si apre la nostra spiegazione del voto.

Perche' esiste
--------------
``voto_puro_discrepancies`` risponde "quanto siamo lontani" in aggregato,
``build_voto_tuner`` porta in Excel una trentina di casi scelti. Mancava il
terzo modo di guardare gli stessi numeri: TUTTI i voti, sfogliabili nel formato
del foglio da cui provengono, dove ogni singolo voto puo' essere interrogato
("perche' 6.5?") senza aprire un notebook.

Due scelte di metodo
--------------------
* ``6*`` nel foglio fantacalcio NON e' un sei: e' il senza voto (voto
  d'ufficio). Qui e' trattato come tale — mostrato "s.v." ed ESCLUSO dalle
  statistiche di accordo. Contarlo come un 6 gonfia l'accordo, perche' i due
  sistemi "concordano" su un numero che nessuno dei due ha misurato; e' anche
  la ragione per cui i numeri di questa pagina sono un filo peggiori di quelli
  stampati da ``voto_puro_discrepancies``, che il ``*`` lo perde.
* il confronto e' sul voto BASE (senza bonus/malus): e' l'unica parte che i tre
  sistemi calcolano davvero, e sommare i bonus — identici per tutti, sono fatti
  — non misurerebbe altro che il rumore residuo.

Il match giocatore-a-giocatore riusa il matcher di ``voto_puro_discrepancies``
(cognome + iniziale del nome, indicizzato sulle squadre in cui il giocatore ha
effettivamente giocato), cosi' c'e' una sola implementazione da mantenere.
"""
from __future__ import annotations

import glob
import html
import json
import math
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError

import vfoot.services.classic_rating as cr
from realdata.models import Match, MatchAppearance
from realdata.services.identity import norm_name
from vfoot.management.commands.voto_puro_discrepancies import Command as DiscCmd
from vfoot.services.classic_pagella import (
    get_reference, get_role_averages, pagella_for_match,
)

DEFAULT_DIR = str(Path(settings.VFOOT_DATA_DIR) / "data_fantacalcio" / "2025-2026")
DEFAULT_OUT = str(Path(settings.REPO_ROOT) / "voto_benchmark")
GD_RE = re.compile(r"Giornata_(\d+)")
EXTERNAL_TEAMS = {
    "atalanta", "bologna", "cagliari", "como", "cremonese", "fiorentina", "genoa",
    "verona", "hellas verona", "inter", "juventus", "lazio", "lecce", "milan",
    "napoli", "parma", "pisa", "roma", "sassuolo", "torino", "udinese",
}
# Colonne del foglio: Cod, Ruolo, Nome, Voto, Gf, Gs, Rp, Rs, Rf, Au, Amm, Esp, Ass
C_COD, C_RUOLO, C_NOME, C_VOTO, C_GF, C_AMM, C_ESP, C_ASS = 0, 1, 2, 3, 4, 10, 11, 12

ROLE_ORDER = {"POR": 0, "DIF": 1, "CEN": 2, "ATT": 3}
ROLE_LABEL = {"POR": "portieri", "DIF": "difensori", "CEN": "centrocampisti",
              "ATT": "attaccanti"}
SHEET_ROLE = {"P": "POR", "D": "DIF", "C": "CEN", "A": "ATT"}
# Istogramma e matrice: i voti vivono su una griglia di mezzo punto, quindi i bin
# SONO i valori possibili. Le code sono schiacciate sugli estremi.
HIST_BINS = [x / 2 for x in range(7, 19)]        # 3.5 … 9.0
MATRIX_BINS = [x / 2 for x in range(8, 18)]      # 4.0 … 8.5


def _clamp_bin(v, bins):
    return min(max(round(v * 2) / 2, bins[0]), bins[-1])


def _pearson(xs, ys):
    n = len(xs)
    if n < 2:
        return float("nan")
    mx, my = sum(xs) / n, sum(ys) / n
    cov = sum((a - mx) * (b - my) for a, b in zip(xs, ys))
    sx = math.sqrt(sum((a - mx) ** 2 for a in xs))
    sy = math.sqrt(sum((b - my) ** 2 for b in ys))
    return cov / (sx * sy) if sx and sy else float("nan")


def _mean_sd(v):
    if not v:
        return float("nan"), float("nan")
    mu = sum(v) / len(v)
    return mu, math.sqrt(sum((x - mu) ** 2 for x in v) / len(v))


def _club_key(name):
    fillers = {"ac", "as", "fc", "ssc", "us", "ss", "hellas", "calcio"}
    return " ".join(t for t in norm_name(name).split() if t not in fillers)


def _parse_vote(v):
    """(voto, senza_voto). '6*' e' il voto d'ufficio di fantacalcio: un
    segnaposto, non una misura — torna (6.0, True) e va escluso dai confronti."""
    if v is None:
        return None, False
    s = str(v)
    star = "*" in s
    m = re.search(r"(\d+(?:[.,]\d+)?)", s)
    if not m:
        return None, star
    return float(m.group(1).replace(",", ".")), star


def esc(s):
    return html.escape(str(s))


def f1(v, dash="—"):
    return dash if v is None or (isinstance(v, float) and math.isnan(v)) else f"{v:.1f}"


def f2(v, dash="—"):
    return dash if v is None or (isinstance(v, float) and math.isnan(v)) else f"{v:.2f}"


def signed(v, dash="—"):
    return dash if v is None else f"{v:+.2f}"


class Command(BaseCommand):
    help = "Costruisce il benchmark HTML del voto puro vs fantacalcio (Redazione/Statistico)."

    def add_arguments(self, parser):
        parser.add_argument("--season", type=int, default=2,
                            help="competition_season_id (default 2 = Serie A 2025-26)")
        parser.add_argument("--dir", default=DEFAULT_DIR,
                            help="cartella dei fogli fantacalcio (uno per giornata)")
        parser.add_argument("--out", default=DEFAULT_OUT,
                            help="cartella di destinazione delle pagine HTML")
        parser.add_argument("--matchdays", default=None,
                            help="sottoinsieme di giornate, es. '1-5,12' (default: tutte)")

    # ------------------------------------------------------------------ input

    @staticmethod
    def _wanted(spec):
        if not spec:
            return None
        out = set()
        for part in spec.split(","):
            part = part.strip()
            if "-" in part:
                a, b = part.split("-", 1)
                out.update(range(int(a), int(b) + 1))
            elif part:
                out.add(int(part))
        return out

    def _parse_sheet_file(self, path):
        """Un file-giornata -> (righe del foglio Redazione, {cod: (voto, sv)} dello
        Statistico). Le due schede elencano gli stessi giocatori con lo stesso
        codice fantacalcio, quindi lo Statistico si aggancia sul CODICE: nessun
        secondo match per nome, nessuna possibilita' di disallineamento."""
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        def rows_of(sheet):
            if sheet not in wb.sheetnames:
                return []
            out, team = [], None
            for r in wb[sheet].iter_rows(values_only=True):
                c0 = r[C_COD]
                if isinstance(c0, str) and norm_name(c0) in EXTERNAL_TEAMS:
                    team = c0
                    continue
                if not isinstance(c0, (int, float)):
                    continue
                voto, sv = _parse_vote(r[C_VOTO])
                out.append({
                    "cod": int(c0), "team": team,
                    "ruolo": str(r[C_RUOLO] or "").upper().strip(),
                    "nome": str(r[C_NOME] or "").strip(),
                    "voto": voto, "sv": sv,
                    "gf": int(r[C_GF] or 0), "ass": int(r[C_ASS] or 0),
                    "amm": int(r[C_AMM] or 0), "esp": int(r[C_ESP] or 0),
                })
            return out

        fanta = rows_of("Fantacalcio")
        stat = {r["cod"]: (r["voto"], r["sv"]) for r in rows_of("Statistico")}
        wb.close()
        return fanta, stat

    def _our_side(self, cs_id, matchdays):
        """{(gd, pid): riga nostra} + {(gd, nome_squadra): incontro}.

        La riga nostra e' quella della pagella reale: voto puro, spiegazione,
        minuti, eventi e fantavoto — esattamente cio' che il sito mostrerebbe."""
        ref, avgs = get_reference(cs_id), get_role_averages(cs_id)
        # Il rating SofaScore: terzo giudice, indipendente dalle pagelle italiane e
        # costruito sugli eventi. Su una divergenza dice da che parte sta chi guarda
        # la partita invece del tabellino.
        rating = {(md, pid): (rs or {}).get('rating') for pid, md, rs in
                  MatchAppearance.objects.filter(match__competition_season_id=cs_id)
                  .values_list('player_id', 'match__matchday', 'raw_stats')
                  if (rs or {}).get('rating')}
        ours, fixtures = {}, {}
        qs = (Match.objects.filter(competition_season_id=cs_id)
              .select_related("home_team__team", "away_team__team").order_by("matchday"))
        for m in qs:
            if matchdays and m.matchday not in matchdays:
                continue
            names = {"home": m.home_team.team.name, "away": m.away_team.team.name}
            # A postponed fixture and the match that replaced it are TWO rows with
            # the same matchday and the same two teams (5 of them in 2025-26), and
            # the postponed one carries no score. Keyed by (matchday, team) it
            # overwrote the played match and the block header read "None-None".
            # Only a match with a result describes an encounter; see the season-long
            # sibling of this trap in the matchday counting.
            scored = m.home_goals is not None and m.away_goals is not None
            for side, opp in (("home", "away"), ("away", "home")):
                key = (m.matchday, names[side])
                # never let a scoreless row replace a played one; a scoreless row
                # still fills an empty slot, so a match in progress keeps a header
                if not scored and (fixtures.get(key) or {}).get("gf") is not None:
                    continue
                fixtures[key] = {
                    "opp": names[opp], "home": side == "home",
                    "gf": m.home_goals if side == "home" else m.away_goals,
                    "gs": m.away_goals if side == "home" else m.home_goals,
                }
            pag = pagella_for_match(m, ref, averages=avgs, full_explanation=True)
            for side in ("home", "away"):
                for group in ("starters", "bench"):
                    for ln in pag[side][group]:
                        ours[(m.matchday, ln["player_id"])] = {
                            **ln, "team": names[side], "starter": group == "starters",
                            "rating": rating.get((m.matchday, ln["player_id"])),
                        }
        return ours, fixtures

    # ------------------------------------------------------------------- join

    def _resolve(self, pidx, pid_first, our_team, nome):
        """Il giocatore del foglio -> il nostro player_id (o None). Stessa regola
        di voto_puro_discrepancies: cognome, e in caso di omonimia in squadra
        l'abbreviazione del nome ('Lu.' -> Luca) come prefisso."""
        surname, abbr = DiscCmd._surname_and_initial(nome)
        cands = pidx.get((our_team, surname), [])
        if len(cands) > 1 and abbr:
            narrowed = [pid for pid in cands
                        if any(fn.startswith(abbr) for fn in pid_first.get(pid, ()))]
            if narrowed:
                cands = narrowed
        return cands[0] if len(cands) == 1 else None

    def _build_matchday(self, gd, fanta_rows, stat_by_cod, ours, fixtures,
                        team_map, pidx, pid_first):
        """Il modello di UNA giornata: blocchi-squadra nell'ordine del foglio,
        ogni riga con i tre voti affiancati e la nostra spiegazione."""
        blocks, by_team, order = [], defaultdict(list), []
        used_pids = set()
        for e in fanta_rows:
            if e["ruolo"] == "ALL":       # allenatori: fantacalcio li vota, noi no
                continue
            our_team = team_map.get(_club_key(e["team"] or ""))
            pid = self._resolve(pidx, pid_first, our_team, e["nome"]) if our_team else None
            o = ours.get((gd, pid)) if pid else None
            if o is not None:
                used_pids.add(pid)
            s_voto, s_sv = stat_by_cod.get(e["cod"], (None, False))
            our_v = None if (o is None or o.get("sv")) else o.get("voto_puro")
            f_voto = None if e["sv"] else e["voto"]
            s_voto = None if s_sv else s_voto
            row = {
                "pid": pid, "cod": e["cod"], "nome": e["nome"], "ruolo": e["ruolo"],
                "role": (o or {}).get("role") or SHEET_ROLE.get(e["ruolo"], ""),
                "f": f_voto, "f_sv": e["sv"], "s": s_voto, "s_sv": s_sv,
                "our": our_v,
                "our_sv": bool(o and o.get("sv")),
                "sv_reason": (o or {}).get("sv_reason"),
                "matched": o is not None,
                "minutes": (o or {}).get("minutes"),
                "events": (o or {}).get("events") or {},
                "starter": (o or {}).get("starter"),
                "bonus": (o or {}).get("bonus"), "malus": (o or {}).get("malus"),
                "fantavoto": (o or {}).get("fantavoto"),
                "expl": (o or {}).get("explanation"),
                "expl_text": (o or {}).get("explanation_text", ""),
                "rating": (o or {}).get("rating"),
                "conceded": (o or {}).get("conceded"),
                "sheet_gf": e["gf"], "sheet_ass": e["ass"],
                "amm": e["amm"], "esp": e["esp"],
            }
            row["gd"] = gd
            row["df"] = None if (our_v is None or f_voto is None) else round(our_v - f_voto, 2)
            row["ds"] = None if (our_v is None or s_voto is None) else round(our_v - s_voto, 2)
            # discordanza sugli EVENTI: non e' il voto, e' la materia prima. Un gol
            # che loro hanno e noi no rende insensato ogni confronto su quella riga.
            row["ev_mismatch"] = bool(
                o is not None and (
                    (row["events"].get("goals", 0) or 0) != e["gf"]
                    or (row["events"].get("assists", 0) or 0) != e["ass"]))
            key = e["team"] or "?"
            if key not in by_team:
                order.append(key)
            by_team[key].append(row)

        for team in order:
            our_team = team_map.get(_club_key(team))
            blocks.append({"team": team, "our_team": our_team,
                           "fixture": fixtures.get((gd, our_team)) if our_team else None,
                           "rows": by_team[team]})
        # I nostri votati che il foglio non elenca (o che il matcher non aggancia):
        # e' la misura onesta della copertura, quindi si vede, non si nasconde.
        missing = [o for (g, pid), o in ours.items()
                   if g == gd and pid not in used_pids and not o.get("sv")]
        missing.sort(key=lambda o: (o["team"], ROLE_ORDER.get(o["role"], 9), o["name"]))
        return {"gd": gd, "blocks": blocks, "missing": missing}

    # -------------------------------------------------------------- statistica

    @staticmethod
    def _agg(rows, key):
        """Accordo fra il nostro voto e la colonna ``key`` ('f' o 's'), sulle sole
        righe in cui ENTRAMBI hanno espresso un voto vero (niente s.v.)."""
        pairs = [(r["our"], r[key]) for r in rows
                 if r["our"] is not None and r[key] is not None]
        if not pairs:
            return {"n": 0, "mae": None, "bias": None, "corr": None,
                    "w05": None, "w10": None, "rmse": None}
        d = [a - b for a, b in pairs]
        n = len(pairs)
        return {
            "n": n,
            "mae": sum(abs(x) for x in d) / n,
            "rmse": math.sqrt(sum(x * x for x in d) / n),
            "bias": sum(d) / n,
            "corr": _pearson([a for a, _ in pairs], [b for _, b in pairs]),
            "w05": 100 * sum(1 for x in d if abs(x) <= 0.5) / n,
            "w10": 100 * sum(1 for x in d if abs(x) <= 1.0) / n,
        }

    @staticmethod
    def _sv_agreement(rows, key):
        """Quanto il NOSTRO insieme di senza-voto coincide con il loro.

        E' l'unica domanda sensata sui s.v.: il '6*' del foglio non e' un voto, e
        confrontarlo con un nostro numero non misurerebbe niente. Confrontare
        invece CHI viene lasciato senza voto misura una regola contro un'altra.
        Solo righe agganciate: dove non abbiamo la presenza non abbiamo regola."""
        m = [r for r in rows if r["matched"]]
        both = ours = theirs = neither = 0
        holes = 0
        for r in m:
            o = r["our_sv"]
            t = r[f"{key}_sv"] or r[key] is None
            holes += int(o and r["sv_reason"] == "dati_mancanti")
            if o and t:
                both += 1
            elif o:
                ours += 1
            elif t:
                theirs += 1
            else:
                neither += 1
        n = len(m) or 1
        prec = both / max(both + ours, 1)
        rec = both / max(both + theirs, 1)
        return {"n": len(m), "both": both, "only_ours": ours, "only_theirs": theirs,
                "neither": neither, "holes": holes,
                "acc": 100 * (both + neither) / n,
                "f1": 100 * 2 * prec * rec / max(prec + rec, 1e-9)}

    @staticmethod
    def _sv_by_minutes(rows):
        """La regola dell'uno e dell'altro, letta dai dati: quota di senza voto per
        fascia di minuti giocati. E' qui che si vede DOVE le due regole divergono,
        che e' l'unica cosa su cui si possa intervenire."""
        bands = [(1, 5), (5, 10), (10, 12), (12, 14), (14, 16), (16, 18),
                 (18, 20), (20, 25), (25, 45), (45, 91)]
        out = []
        for lo, hi in bands:
            g = [r for r in rows if r["matched"] and r["minutes"] is not None
                 and lo <= r["minutes"] < hi]
            if not g:
                continue
            out.append({
                "lo": lo, "hi": hi, "n": len(g),
                "ours": 100 * sum(1 for r in g if r["our_sv"]) / len(g),
                "theirs": 100 * sum(1 for r in g
                                    if r["f_sv"] or r["f"] is None) / len(g),
            })
        return out

    @staticmethod
    def _coverage(rows):
        c = {"tot": len(rows), "matched": 0, "both": 0, "our_only": 0,
             "ext_only": 0, "neither": 0, "unmatched": 0, "ev_mismatch": 0}
        for r in rows:
            if not r["matched"]:
                c["unmatched"] += 1
                continue
            c["matched"] += 1
            c["ev_mismatch"] += int(r["ev_mismatch"])
            ours, ext = r["our"] is not None, r["f"] is not None
            c["both"] += int(ours and ext)
            c["our_only"] += int(ours and not ext)
            c["ext_only"] += int(ext and not ours)
            c["neither"] += int(not ours and not ext)
        return c

    # ------------------------------------------------------------------- main

    def handle(self, *a, **o):
        cs_id = o["season"]
        files = sorted(glob.glob(f"{o['dir']}/*.xlsx"))
        if not files:
            raise CommandError(f"Nessun .xlsx in {o['dir']}")
        wanted = self._wanted(o["matchdays"])
        outdir = Path(o["out"])
        outdir.mkdir(parents=True, exist_ok=True)

        self.stdout.write("Calcolo voto puro + spiegazione per ogni partita…")
        ours, fixtures = self._our_side(cs_id, wanted)
        team_map = DiscCmd()._our_team_index(cs_id)
        pidx, _pid_team, pid_first = DiscCmd()._our_player_index(cs_id)

        self.stdout.write("Lettura fogli fantacalcio…")
        days = []
        for f in files:
            m = GD_RE.search(f)
            if not m:
                continue
            gd = int(m.group(1))
            if wanted and gd not in wanted:
                continue
            fanta_rows, stat_by_cod = self._parse_sheet_file(f)
            days.append(self._build_matchday(gd, fanta_rows, stat_by_cod, ours,
                                             fixtures, team_map, pidx, pid_first))
        if not days:
            raise CommandError("Nessuna giornata da elaborare.")
        days.sort(key=lambda d: d["gd"])
        gds = [d["gd"] for d in days]

        for i, d in enumerate(days):
            rows = [r for b in d["blocks"] for r in b["rows"]]
            d["rows"] = rows
            d["stats"] = {"f": self._agg(rows, "f"), "s": self._agg(rows, "s"),
                          "cov": self._coverage(rows)}
            d["prev"] = gds[i - 1] if i else None
            d["next"] = gds[i + 1] if i + 1 < len(gds) else None
            (outdir / f"giornata-{d['gd']:02d}.html").write_text(
                self._render_matchday(d), encoding="utf-8")

        (outdir / "index.html").write_text(self._render_index(days), encoding="utf-8")
        allrows_ = [r for d in days for r in d["rows"]]
        page = self._render_divergences(allrows_)
        if page:
            (outdir / "divergenze.html").write_text(page, encoding="utf-8")

        allrows = [r for d in days for r in d["rows"]]
        sf, ss = self._agg(allrows, "f"), self._agg(allrows, "s")
        cov = self._coverage(allrows)
        self.stdout.write(self.style.SUCCESS(
            f"\n{len(days)} giornate · {cov['tot']} righe di foglio · "
            f"agganciate {cov['matched']} ({100*cov['matched']/max(cov['tot'],1):.1f}%)"))
        self.stdout.write(
            f"  vs Redazione : n={sf['n']} MAE={f2(sf['mae'])} bias={signed(sf['bias'])} "
            f"corr={f2(sf['corr'])} entro 0.5={f1(sf['w05'])}%")
        self.stdout.write(
            f"  vs Statistico: n={ss['n']} MAE={f2(ss['mae'])} bias={signed(ss['bias'])} "
            f"corr={f2(ss['corr'])} entro 0.5={f1(ss['w05'])}%")
        self.stdout.write(f"report -> {outdir / 'index.html'}")

    # ---------------------------------------------------------------- rendering

    @staticmethod
    def _page(title, body, data=None):
        # charset esplicito: la pagina si apre da file:// e senza questo il
        # browser tira a indovinare (e sbaglia su ogni accento e ogni σ).
        payload = ""
        if data:
            chans, players = data
            payload = (f"const FCHAN={json.dumps(chans, ensure_ascii=False)};\n"
                       f"const FROWS={json.dumps(players)};\n")
        return ('<!doctype html>\n<meta charset="utf-8">\n'
                '<meta name="viewport" content="width=device-width,initial-scale=1">\n'
                f"<title>{esc(title)}</title>\n<style>{_CSS}</style>\n"
                f'<div class="wrap">{body}</div>\n'
                f"<script>{payload}{_JS}</script>\n")

    @staticmethod
    def _tile(label, value, note=""):
        return (f'<div class="tile"><span class="tl">{esc(label)}</span>'
                f'<span class="tv">{value}</span>'
                f'<span class="tn">{esc(note)}</span></div>')

    # -- una giornata ---------------------------------------------------------

    @staticmethod
    def _divergences(rows, threshold=1.0):
        """Le righe in cui siamo fuori da CIO' CHE DICE IL PIU' VICINO dei due voti
        esterni, di almeno ``threshold``.

        Non lo scarto dalla Redazione: se lei dice 6.5 e lo Statistico 7.5 e noi
        diciamo 7.5, non stiamo divergendo da nessuno — stiamo scegliendo fra due
        letture che gia' divergono fra loro. Il caso da guardare e' quello in cui
        entrambi ci contraddicono, ed e' quello che questo scarto misura."""
        out = []
        for r in rows:
            ext = [x for x in (r["f"], r["s"]) if x is not None]
            if not ext or r["our"] is None or r["ev_mismatch"]:
                continue
            hi, lo = max(ext), min(ext)
            gap = (r["our"] - hi) if r["our"] > hi else (
                (r["our"] - lo) if r["our"] < lo else 0.0)
            if abs(gap) >= threshold:
                out.append({**r, "gap": round(gap, 2)})
        out.sort(key=lambda r: -abs(r["gap"]))
        return out

    def _divergence_section(self, rows, threshold=1.0):
        """La sezione da leggere caso per caso: quattro giudici affiancati, filtri
        per ruolo e direzione, e per ognuno la tabella completa delle feature."""
        div = self._divergences(rows, threshold)
        if not div:
            return "", ([], {})
        payload = self._feature_payload(div, key=lambda r: f'{r["gd"]}-{r["pid"]}')
        trs = []
        for r in div:
            ev = []
            e = r["events"]
            for n, sym, klass in ((e.get("goals", 0), "gol", "g"),
                                  (e.get("assists", 0), "assist", "a"),
                                  (e.get("own_goals", 0), "autogol", "o"),
                                  (e.get("yellow", 0), "amm", "y"),
                                  (e.get("red", 0), "esp", "rd")):
                if n:
                    ev.append(_chip(n, sym, klass))
            if r["role"] == "POR" and r.get("conceded"):
                ev.append(_chip(r["conceded"], "gol subiti", "rd"))
            up = r["gap"] > 0
            trs.append(
                f'<tr class="r" data-role="{r["role"]}" '
                f'data-dir="{"up" if up else "down"}" data-gap="{abs(r["gap"]):.1f}" '
                f'tabindex="0">'
                f'<td class="c-n"><a href="giornata-{r["gd"]:02d}.html#p{r["pid"]}" '
                f'onclick="event.stopPropagation()">{esc(r["nome"])}</a>'
                f'<span class="tg">{esc(r["role"])} · {r["gd"]}ª</span></td>'
                f'<td class="c-v">{f1(r["f"])}</td><td class="c-v">{f1(r["s"])}</td>'
                f'<td class="c-v c-our"><b>{f1(r["our"])}</b></td>'
                f'{_delta_cell(r["gap"])}'
                f'<td class="c-v dim">{f1(r["rating"])}</td>'
                f'<td class="c-m">{r["minutes"] if r["minutes"] is not None else ""}</td>'
                f'<td class="c-e">{"".join(ev)}</td></tr>'
                + self._detail_html(r, key=f'{r["gd"]}-{r["pid"]}', cols=8))
        n_up = sum(1 for r in div if r["gap"] > 0)
        body = (
            "<p class=\"sub2\">Le presenze in cui il nostro voto sta fuori da quello "
            "che dice <b>il piu' vicino</b> dei due voti esterni, di almeno mezzo "
            f"punto e mezzo. Se la Redazione dice 6.5 e lo Statistico 7.5, un nostro "
            "7.5 non e' una divergenza: stiamo scegliendo fra due letture che gia' "
            "divergono. Qui ci sono solo i casi in cui <b>ci contraddicono "
            "entrambi</b>. Righe con eventi discordi escluse. Il <b>rating "
            "SofaScore</b> e' il terzo giudice: e' su un'altra scala (media 6.8 "
            "contro 6.0), quindi si legge come direzione, non come voto — ma e' "
            "costruito sugli eventi, quindi dice da che parte sta chi guarda la "
            "partita invece del tabellino. <b>Clicca una riga</b> per la spiegazione "
            "e tutte le voci del modello.</p>"
            f'<div class="bar"><div class="chips" id="dvrole">'
            f'<button class="chip on" data-role="">tutti i ruoli</button>'
            + "".join(f'<button class="chip" data-role="{r}">{ROLE_LABEL[r]}</button>'
                      for r in ("POR", "DIF", "CEN", "ATT"))
            + '</div><div class="chips" id="dvdir">'
            '<button class="chip on" data-dir="">entrambe</button>'
            f'<button class="chip" data-dir="up">siamo piu\' alti ({n_up})</button>'
            f'<button class="chip" data-dir="down">siamo piu\' bassi '
            f'({len(div)-n_up})</button>'
            '</div><div class="chips" id="dvthr">'
            '<button class="chip on" data-thr="1">&ge; 1</button>'
            '<button class="chip" data-thr="1.5">&ge; 1.5</button>'
            '<button class="chip" data-thr="2">&ge; 2</button>'
            '</div><span class="count" id="dvcount"></span></div>'
            '<div class="tw"><table class="tbl grid dv"><thead><tr>'
            '<th>giocatore</th><th class="c-v">red.</th><th class="c-v">stat.</th>'
            '<th class="c-v c-our">nostro</th><th class="c-d">fuori di</th>'
            '<th class="c-v">sofa</th><th class="c-m">min</th><th>eventi</th>'
            f'</tr></thead><tbody>{"".join(trs)}</tbody></table></div>')
        return body, payload

    @staticmethod
    def _feature_payload(rows, key=None):
        """Il dettaglio per-feature come dati, non come HTML.

        Una tabella di 42 righe per ciascuno dei ~250 giocatori votati sarebbe un
        megabyte e mezzo di markup per pagina, quasi tutto mai aperto. Qui va un
        payload compatto — il vocabolario del canale una volta sola, poi tre numeri
        per feature per giocatore — e la tabella la costruisce il browser al click.

        Il canale e' identificato dalla LISTA di feature, non dal ruolo: due
        giocatori con lo stesso elenco condividono per definizione lo stesso
        vocabolario, e cosi' non c'e' nessuna ipotesi sui ruoli da tenere allineata.
        """
        key = key or (lambda r: str(r["pid"]))
        chan_id, chans, players = {}, [], {}
        for r in rows:
            terms = ((r.get("expl") or {}).get("all_terms")) or []
            if not terms or not r["pid"]:
                continue
            sig = tuple(t["key"] for t in terms)
            if sig not in chan_id:
                chan_id[sig] = len(chans)
                chans.append([[t["key"], t["label"], t["kind"], t["weight"],
                               t.get("family", ""),
                               t.get("z_one", 0.0) if t.get("event") else None]
                              for t in terms])
            # punti di voto per punto di indice: col peso (condiviso) e la z (per
            # giocatore) il browser ricostruisce indice e media di ruolo, che e' il
            # metro di paragone di ogni voce
            flat = []
            for t in terms:
                flat += [t["value"], t["z"], t["points"]]
            players[key(r)] = [chan_id[sig],
                               (r["expl"] or {}).get("per_unit", 0.0)] + flat
        return chans, players

    def _render_matchday(self, d):
        gd, sf, ss, cov = d["gd"], d["stats"]["f"], d["stats"]["s"], d["stats"]["cov"]
        nav = ['<a class="nv" href="index.html">↑ indice</a>']
        if d["prev"]:
            nav.append(f'<a class="nv" href="giornata-{d["prev"]:02d}.html">← {d["prev"]}ª</a>')
        if d["next"]:
            nav.append(f'<a class="nv" href="giornata-{d["next"]:02d}.html">{d["next"]}ª →</a>')

        tiles = "".join([
            self._tile("confronti", str(sf["n"]), "voti veri da entrambi"),
            self._tile("scarto medio · redazione", f2(sf["mae"]),
                       f"bias {signed(sf['bias'])}"),
            self._tile("scarto medio · statistico", f2(ss["mae"]),
                       f"bias {signed(ss['bias'])}"),
            self._tile("entro mezzo voto", f"{f1(sf['w05'])}%", "sulla redazione"),
        ])

        blocks = []
        for b in d["blocks"]:
            fx = b["fixture"]
            sub = ""
            if fx:
                # il nostro nome squadra da entrambi i lati: col nome del foglio per
                # il soggetto e il nostro per l'avversario, la stessa partita si
                # leggeva in due modi nei due blocchi ("Verona 2-3 Bologna" e
                # "Hellas Verona 2-3 Bologna")
                me = b["our_team"] or b["team"]
                home, away = (me, fx["opp"]) if fx["home"] else (fx["opp"], me)
                if fx["gf"] is None or fx["gs"] is None:
                    # partita senza risultato (in corso, o rinviata): niente
                    # punteggio inventato, solo chi contro chi
                    sub = f"{home} – {away}"
                else:
                    gh, ga = ((fx["gf"], fx["gs"]) if fx["home"]
                              else (fx["gs"], fx["gf"]))
                    sub = f"{home} {gh}-{ga} {away}"
            rows = []
            for r in b["rows"]:
                rows.append(self._row_html(r))
            blocks.append(
                f'<section class="tm">'
                f'<h2>{esc(b["team"])}<span class="fx">{esc(sub)}</span></h2>'
                f'<div class="tw"><table class="grid"><thead><tr>'
                f'<th class="c-r">R</th><th class="c-n">giocatore</th>'
                f'<th class="c-v">red.</th><th class="c-v">stat.</th>'
                f'<th class="c-v c-our">nostro</th><th class="c-d">Δ red.</th>'
                f'<th class="c-d">Δ stat.</th><th class="c-m">min</th>'
                f'<th class="c-e">eventi</th></tr></thead>'
                f'<tbody>{"".join(rows)}</tbody></table></div></section>')

        missing = ""
        if d["missing"]:
            items = ", ".join(f'{esc(o["name"])} <i>({esc(o["team"])}, '
                              f'{f1(o["voto_puro"])})</i>' for o in d["missing"])
            missing = (f'<details class="note"><summary>{len(d["missing"])} nostri voti '
                       f'non presenti nel foglio (o non agganciati per nome)</summary>'
                       f'<p>{items}</p></details>')

        body = (
            f'<nav class="nav">{"".join(nav)}</nav>'
            f'<h1>Giornata {gd}<span class="h1sub">voto puro vs fantacalcio</span></h1>'
            f'<p class="sub">Ogni riga e\' un giocatore del foglio fantacalcio di questa '
            f'giornata. <b>Clicca una riga</b> per leggere perche\' il nostro voto e\' '
            f'quello che e\'.</p>'
            f'<div class="tiles">{tiles}</div>'
            f'<div class="bar">'
            f'<input id="q" class="q" type="search" placeholder="cerca un giocatore…" '
            f'autocomplete="off">'
            f'<div class="chips" id="filters">'
            f'<button class="chip on" data-thr="0">tutti</button>'
            f'<button class="chip" data-thr="1">Δ ≥ 1</button>'
            f'<button class="chip" data-thr="1.5">Δ ≥ 1.5</button>'
            f'<button class="chip" data-thr="sv">solo s.v. discordi</button>'
            f'</div><span class="count" id="count"></span></div>'
            f'{"".join(blocks)}'
            f'{missing}'
            f'<p class="foot">{esc(_foot_matchday(cov))}</p>')
        return self._page(f"Giornata {gd} · benchmark voto puro", body,
                          data=self._feature_payload(d["rows"]))

    def _row_html(self, r):
        """Una riga giocatore + la riga-dettaglio (nascosta) con la spiegazione."""
        cls = ["r"]
        if r["ev_mismatch"]:
            cls.append("evx")
        data = (f'data-n="{esc(norm_name(r["nome"]))}" '
                f'data-df="{abs(r["df"]) if r["df"] is not None else -1}" '
                f'data-sv="{1 if (r["our"] is None) != (r["f"] is None) else 0}"')
        pid = f' id="p{r["pid"]}"' if r["pid"] else ""

        def vcell(v, sv, extra=""):
            if sv:
                return f'<td class="c-v sv{extra}">s.v.</td>'
            return f'<td class="c-v{extra}">{f1(v)}</td>'

        if r["our"] is not None:
            our_cell = f'<td class="c-v c-our"><b>{f1(r["our"])}</b></td>'
        else:
            why = {"impiego_insufficiente": "troppo poco impiego",
                   "dati_mancanti": "dati mancanti"}.get(r["sv_reason"], "non agganciato")
            our_cell = f'<td class="c-v c-our sv" title="{esc(why)}">s.v.</td>'

        ev = []
        e = r["events"]
        for n, sym, klass in ((e.get("goals", 0), "gol", "g"),
                              (e.get("assists", 0), "assist", "a"),
                              (e.get("own_goals", 0), "autogol", "o"),
                              (e.get("yellow", 0), "amm", "y"),
                              (e.get("red", 0), "esp", "rd")):
            if n:
                ev.append(_chip(n, sym, klass))
        if not r["matched"]:
            # nessun aggancio: mostriamo almeno gli eventi che dichiara il foglio
            for n, sym, klass in ((r["sheet_gf"], "gol", "g"), (r["sheet_ass"], "assist", "a"),
                                  (r["amm"], "amm", "y"), (r["esp"], "esp", "rd")):
                if n:
                    ev.append(_chip(n, sym, klass))

        role_badge = ""
        if r["role"] and SHEET_ROLE.get(r["ruolo"]) and r["role"] != SHEET_ROLE[r["ruolo"]]:
            # loro lo mettono in difesa, noi lo valutiamo da centrocampista: il voto
            # e' z-scorato su un altro ruolo, e questo va detto sulla riga.
            role_badge = f'<span class="rb" title="ruolo con cui lo valutiamo">{r["role"]}</span>'

        nomatch = ("" if r["matched"] else
                   '<span class="nm" title="nessun aggancio col nostro archivio">?</span>')
        main = (
            f'<tr class="{" ".join(cls)}"{pid} {data} tabindex="0">'
            f'<td class="c-r">{esc(r["ruolo"])}</td>'
            f'<td class="c-n">{esc(r["nome"])}{role_badge}{nomatch}</td>'
            f'{vcell(r["f"], r["f_sv"])}{vcell(r["s"], r["s_sv"])}{our_cell}'
            f'{_delta_cell(r["df"])}{_delta_cell(r["ds"])}'
            f'<td class="c-m">{r["minutes"] if r["minutes"] is not None else ""}</td>'
            f'<td class="c-e">{"".join(ev)}</td></tr>')
        return main + self._detail_html(r)

    def _detail_html(self, r, key=None, cols=9):
        """La spiegazione: la frase, poi il bilancio additivo che ci arriva."""
        if r["our"] is None and not r["matched"]:
            inner = ('<p class="lead">Giocatore non agganciato al nostro archivio: '
                     'nessun voto da confrontare.</p>')
        elif r["our"] is None:
            why = {"impiego_insufficiente":
                   f'Senza voto: impiego sotto la soglia '
                   f'({cr.MIN_MINUTES_RATED}′ o {cr.MIN_TOUCHES_RATED} palloni giocati).',
                   "dati_mancanti": "Senza voto: dati della partita mancanti."}.get(
                       r["sv_reason"], "Senza voto.")
            inner = f'<p class="lead">{esc(why)}</p>'
        else:
            x = r["expl"] or {}
            contribs = x.get("contributions") or []
            scale = max(0.5, max((abs(c["points"]) for c in contribs), default=0.5))
            led = [f'<div class="lr"><span class="ll">base</span>'
                   f'{_bar(0, scale)}<span class="lv">{f2(x.get("base", 6.0))}</span></div>']
            for c in contribs:
                # Una riga fusa non e' una feature: dirlo qui evita la domanda
                # legittima "questo numero a quale riga della tabella corrisponde?"
                fam = (f'<span class="mrg" title="somma delle {c["family_size"]} voci '
                       f'&quot;{esc(c["family"])}&quot; nella tabella completa">'
                       f'{c["family_size"]} voci</span>') if c.get("family") else ""
                led.append(f'<div class="lr"><span class="ll">{esc(c["label"])}'
                           f'{fam}</span>{_bar(c["points"], scale)}'
                           f'<span class="lv">{c["points"]:+.2f}</span></div>')
            if abs(x.get("other_points", 0.0)) >= 0.005:
                led.append(f'<div class="lr o"><span class="ll">altre '
                           f'{x.get("other_count", 0)} voci</span>'
                           f'{_bar(x["other_points"], scale)}'
                           f'<span class="lv">{x["other_points"]:+.2f}</span></div>')
            led.append(f'<div class="lr t"><span class="ll">totale</span>'
                       f'<span class="lb"></span>'
                       f'<span class="lv">{f2(x.get("subtotal"))} → '
                       f'<b>{f1(r["our"])}</b></span></div>')
            fv = ""
            if r["fantavoto"] is not None:
                # il dettaglio in parentesi solo quando c'e' davvero qualcosa da
                # sommare: "fantavoto 5.0 (5.0)" non dice nulla a nessuno.
                detail = ((f' ({f1(r["our"])}'
                           f'{" +" + f1(r["bonus"]) if r["bonus"] else ""}'
                           f'{" −" + f1(r["malus"]) if r["malus"] else ""})')
                          if (r["bonus"] or r["malus"]) else "")
                fv = f' · fantavoto <b>{f1(r["fantavoto"])}</b>{detail}'
            note = f' · {esc(x.get("note"))}' if x.get("note") else ""
            warn = ""
            if r["ev_mismatch"]:
                warn = (f'<p class="warn">Eventi discordi col foglio: loro '
                        f'{r["sheet_gf"]} gol / {r["sheet_ass"]} assist, noi '
                        f'{r["events"].get("goals", 0)} / {r["events"].get("assists", 0)}. '
                        f'Il confronto su questa riga vale poco.</p>')
            # Tutte le voci, col nome tecnico che hanno nel foglio dei pesi. La
            # tabella non e' nel markup: la costruisce il browser dal payload (vedi
            # _feature_payload), altrimenti sarebbero ~1.5 MB per pagina di righe
            # che nel 99% dei casi nessuno apre.
            n_terms = len((r["expl"] or {}).get("all_terms") or [])
            full = ""
            if n_terms and r["pid"]:
                full = (f'<details class="ft" data-pid="{key or r["pid"]}">'
                        f'<summary>&nbsp;tutte le {n_terms} voci del modello — '
                        f'valore, σ, peso, contributo</summary>'
                        f'<div class="ftbody"></div></details>')
            inner = (f'<p class="lead">{esc(r["expl_text"] or "—")}</p>'
                     f'<div class="ledger">{"".join(led)}</div>'
                     f'<p class="meta">{r["minutes"]}′'
                     f'{" da titolare" if r["starter"] else " subentrato"}{fv}{note}</p>'
                     f'{warn}{full}')
        return (f'<tr class="d" hidden><td colspan="{cols}">'
                f'<div class="why">{inner}</div></td></tr>')

    def _sv_section(self, rows):
        """Il confronto sui SENZA VOTO: non un numero contro un numero, ma una
        regola contro una regola."""
        af, as_ = self._sv_agreement(rows, "f"), self._sv_agreement(rows, "s")

        def matrix(a, title):
            def cell(v, cls=""):
                return f'<td class="num {cls}">{v}</td>'
            return (
                f'<figure class="mx"><figcaption>{esc(title)} — accordo '
                f'<b>{a["acc"]:.2f}%</b></figcaption>'
                f'<table class="tbl sv"><thead><tr><th></th>'
                f'<th class="num">loro s.v.</th><th class="num">loro a voto</th>'
                f'</tr></thead><tbody>'
                f'<tr><th>noi s.v.</th>{cell(a["both"], "ok")}'
                f'{cell(a["only_ours"], "ko")}</tr>'
                f'<tr><th>noi a voto</th>{cell(a["only_theirs"], "ko")}'
                f'{cell(a["neither"], "ok")}</tr>'
                f'</tbody></table>'
                f'<p class="mxnote">Su {a["n"]} presenze agganciate · concordanza '
                f'sui soli s.v. (F1) {a["f1"]:.1f}% · dei nostri s.v. '
                f'{a["holes"]} sono buchi di dati, non una decisione.</p></figure>')

        bands = self._sv_by_minutes(rows)
        brows = []
        for b in bands:
            brows.append(
                f'<tr><td>{b["lo"]}-{b["hi"]}′</td><td class="num">{b["n"]}</td>'
                f'<td class="num">{b["theirs"]:.0f}%</td>'
                f'<td class="bul"><span class="t" style="width:{b["theirs"]:.1f}%">'
                f'</span><span class="o" style="width:{b["ours"]:.1f}%"></span></td>'
                f'<td class="num">{b["ours"]:.0f}%</td></tr>')
        band_tbl = (
            '<div class="tw"><table class="tbl"><thead><tr><th>minuti giocati</th>'
            '<th class="num">n</th><th class="num">s.v. loro</th>'
            '<th>quota di senza voto</th><th class="num">s.v. nostri</th></tr></thead>'
            f'<tbody>{"".join(brows)}</tbody></table></div>')

        return (
            "<p class=\"sub2\">Il <code>6*</code> del foglio non e' un voto, quindi "
            "non entra in nessun confronto di questa pagina. La domanda giusta sui "
            "senza voto e' un'altra: <b>lasciamo fuori le stesse persone?</b> "
            "La nostra regola e' "
            f"&ge;{cr.ALWAYS_RATED_MINUTES}′, oppure &ge;{cr.MIN_MINUTES_RATED}′ con "
            f"almeno {cr.MIN_TOUCHES_RATED} palloni giocati, piu' l'obbligo di voto "
            "per chi ha fatto gol, assist, autogol, un rigore o ha preso un "
            "cartellino.</p>"
            f'<div class="mx-wrap">{matrix(af, "vs Redazione")}'
            f'{matrix(as_, "vs Statistico")}</div>'
            '<p class="sub2 mt">La regola dell\'uno e dell\'altro, letta dai dati — '
            'quota di senza voto per fascia: <span class="kt">barra blu = loro</span>, '
            '<span class="ko2">barra arancio = noi</span>. Sopra i 16′ nessuno dei '
            'due tace piu\', sotto i 12′ tacciono quasi tutti e due: quel che resta '
            'si gioca nella fascia dei subentrati.</p>'
            f'{band_tbl}')

    def _render_divergences(self, rows, threshold=1.0):
        """La pagina dei casi da guardare a uno a uno. Sta per conto suo e non
        nell'indice per una ragione pratica: 600 casi col loro dettaglio pesano
        1.7 MB, e l'indice e' la porta d'ingresso."""
        body, payload = self._divergence_section(rows, threshold)
        if not body:
            return None
        head = ('<nav class="nav"><a class="nv" href="index.html">↑ indice</a></nav>'
                '<h1>Casi di disaccordo grande'
                '<span class="h1sub">dove nessuno dei due ci segue</span></h1>')
        return self._page("Disaccordi · benchmark voto puro", head + body, data=payload)

    # -- indice ---------------------------------------------------------------

    def _render_index(self, days):
        rows = [r for d in days for r in d["rows"]]
        sf, ss = self._agg(rows, "f"), self._agg(rows, "s")
        cov = self._coverage(rows)

        # serie confrontabili: le righe in cui TUTTI E TRE hanno votato davvero
        tri = [r for r in rows if r["our"] is not None and r["f"] is not None
               and r["s"] is not None]
        series = [("red", "Redazione", [r["f"] for r in tri]),
                  ("stat", "Statistico", [r["s"] for r in tri]),
                  ("our", "Nostro", [r["our"] for r in tri])]

        tiles = "".join([
            self._tile("voti confrontati", str(sf["n"]),
                       f"su {cov['tot']} righe di foglio"),
            self._tile("scarto medio · redazione", f2(sf["mae"]),
                       f"corr {f2(sf['corr'])} · bias {signed(sf['bias'])}"),
            self._tile("scarto medio · statistico", f2(ss["mae"]),
                       f"corr {f2(ss['corr'])} · bias {signed(ss['bias'])}"),
            self._tile("entro mezzo voto", f"{f1(sf['w05'])}%",
                       f"entro un voto {f1(sf['w10'])}%"),
        ])

        # per ruolo
        rrows = []
        for role in sorted({r["role"] for r in rows if r["role"]},
                           key=lambda x: ROLE_ORDER.get(x, 9)):
            sub = [r for r in rows if r["role"] == role]
            a, b = self._agg(sub, "f"), self._agg(sub, "s")
            rrows.append(
                f'<tr><td>{esc(ROLE_LABEL.get(role, role))}</td>'
                f'<td class="num">{a["n"]}</td><td class="num">{f2(a["mae"])}</td>'
                f'<td class="num">{signed(a["bias"])}</td><td class="num">{f2(a["corr"])}</td>'
                f'<td class="num">{f2(b["mae"])}</td><td class="num">{signed(b["bias"])}</td>'
                f'<td class="num">{f2(b["corr"])}</td></tr>')
        per_role = (
            '<div class="tw"><table class="tbl"><thead><tr><th>ruolo</th><th class="num">n</th>'
            '<th class="num">MAE red.</th><th class="num">bias</th><th class="num">corr</th>'
            '<th class="num">MAE stat.</th><th class="num">bias</th><th class="num">corr</th>'
            f'</tr></thead><tbody>{"".join(rrows)}</tbody></table></div>')

        div = self._divergences(rows)
        n_up = sum(1 for r in div if r['gap'] > 0)
        by_role = {r: sum(1 for x in div if x['role'] == r)
                   for r in ('POR', 'DIF', 'CEN', 'ATT')}
        div_teaser = (
            f'<p class="sub2">Su {sf["n"]} confronti, in <b>{len(div)}</b> il nostro voto sta fuori da quello che dice il piu\' vicino dei due voti esterni, di almeno un punto — {n_up} verso l\'alto, {len(div) - n_up} verso il basso. Per ruolo: ' + ' · '.join(f'{ROLE_LABEL[r]} {n}' for r, n in by_role.items())
            + '.</p><p class="sub2"><a class="bigl" href="divergenze.html">Aprili uno per uno &rarr;</a> &mdash; quattro giudici affiancati (le due pagelle, noi e il rating SofaScore), filtri per ruolo e direzione, e per ogni caso la spiegazione e tutte le voci del modello.</p>')

        # per giornata
        maxmae = max((d["stats"]["f"]["mae"] or 0) for d in days) or 1
        grows = []
        for d in days:
            a, b = d["stats"]["f"], d["stats"]["s"]
            w = 100 * (a["mae"] or 0) / maxmae
            grows.append(
                f'<tr><td><a href="giornata-{d["gd"]:02d}.html">giornata {d["gd"]}</a></td>'
                f'<td class="num">{a["n"]}</td>'
                f'<td class="num">{f2(a["mae"])}</td>'
                f'<td class="mbar"><span style="width:{w:.1f}%"></span></td>'
                f'<td class="num">{f2(b["mae"])}</td>'
                f'<td class="num">{f1(a["w05"])}%</td></tr>')
        per_gd = (
            '<div class="tw"><table class="tbl gd"><thead><tr><th>giornata</th>'
            '<th class="num">n</th>'
            '<th class="num">MAE red.</th><th></th><th class="num">MAE stat.</th>'
            f'<th class="num">entro 0.5</th></tr></thead>'
            f'<tbody>{"".join(grows)}</tbody></table></div>')

        # le divergenze piu' grosse, misurate sullo STATISTICO (voto senza alone-gol)
        divergent = sorted([r for r in rows if r["ds"] is not None and not r["ev_mismatch"]],
                           key=lambda r: -abs(r["ds"]))
        top = divergent[:40]
        cut = abs(top[-1]["ds"]) if top else 0
        n_over = sum(1 for r in divergent if abs(r["ds"]) >= cut)
        trows = []
        for r in top:
            gd = r["gd"]
            trows.append(
                f'<tr><td><a href="giornata-{gd:02d}.html#p{r["pid"]}">{esc(r["nome"])}</a>'
                f'<span class="tg">{esc(r["role"])} · {gd}ª</span></td>'
                f'<td class="num">{f1(r["f"])}</td><td class="num">{f1(r["s"])}</td>'
                f'<td class="num strong">{f1(r["our"])}</td>'
                f'{_delta_cell(r["ds"])}'
                f'<td class="why-inline">{esc(r["expl_text"])}</td></tr>')
        top_tbl = (
            '<div class="tw"><table class="tbl top"><thead><tr><th>giocatore</th>'
            '<th class="num">red.</th><th class="num">stat.</th><th class="num">nostro</th>'
            '<th class="num">Δ stat.</th><th>perche\' (nostra lettura)</th></tr></thead>'
            f'<tbody>{"".join(trows)}</tbody></table></div>')

        body = (
            f'<h1>Voto puro · benchmark globale<span class="h1sub">'
            f'{len(days)} giornate, Serie A 2025-26</span></h1>'
            f'<p class="sub">Il nostro voto base contro i due voti di fantacalcio.it: '
            f'la <b>Redazione</b> (pagella umana) e lo <b>Statistico</b> (voto '
            f'algoritmico). Il confronto e\' solo sul voto <b>base</b>: bonus e malus '
            f'sono fatti identici per tutti e si cancellano.</p>'
            f'<div class="tiles">{tiles}</div>'
            f'<h3 class="eyebrow">Come si distribuiscono i tre voti</h3>'
            f'{_histogram(series)}'
            f'<h3 class="eyebrow">Voto a voto — dove finiscono i nostri</h3>'
            f'<p class="sub2">Riga = il nostro voto, colonna = il loro. La diagonale '
            f'e\' l\'accordo pieno; sopra la diagonale siamo piu\' generosi, sotto piu\' '
            f'severi.</p>'
            f'<div class="mx-wrap">{_matrix(tri, "f", "vs Redazione")}'
            f'{_matrix(tri, "s", "vs Statistico")}</div>'
            f'<h3 class="eyebrow">Casi di disaccordo grande</h3>'
            f'{div_teaser}'
            f'<h3 class="eyebrow">Chi resta senza voto</h3>'
            f'{self._sv_section(rows)}'
            f'<h3 class="eyebrow">Per ruolo</h3>{per_role}'
            f'<h3 class="eyebrow">Le divergenze piu\' grosse dallo Statistico</h3>'
            "<p class=\"sub2\">Lo Statistico toglie i gol dal voto base: uno scarto "
            "grande qui e' pura divergenza sulla <i>lettura</i> della prestazione, "
            "non l'alone del gol. Righe con eventi discordi escluse. Clicca il nome "
            f"per aprire la riga nella sua giornata. Qui le prime {len(top)} di "
            f"{n_over} righe con |Δ| ≥ {cut:.1f}. "
            "<b>Attenzione</b> alle spiegazioni che dicono &laquo;Espulso&raquo; o "
            "&laquo;Autogol&raquo;: li' il Δ e' gonfiato per costruzione, perche' lo "
            "Statistico mette tutta la penalita' nel voto base mentre noi ne teniamo "
            f"la parte grossa nel malus del fantavoto.</p>{top_tbl}"
            f'<h3 class="eyebrow">Giornata per giornata</h3>{per_gd}'
            f'<p class="foot">{esc(_foot_index(cov, sf, ss))}</p>')
        return self._page("Voto puro · benchmark globale", body)


# ------------------------------------------------------------------ frammenti

def _chip(n, label, klass):
    return f'<span class="ev {klass}">{f"{n} " if n > 1 else ""}{label}</span>'


def _delta_cell(d):
    if d is None:
        return '<td class="c-d"></td>'
    a = abs(d)
    if a < 0.25:                       # accordo pieno: nessun segno, nessun colore
        return '<td class="c-d z">0</td>'
    lvl = 1 if a < 1 else (2 if a < 1.5 else 3)
    return f'<td class="c-d {"p" if d > 0 else "n"}{lvl}">{d:+.1f}</td>'


def _bar(points, scale):
    """Barra divergente centrata sullo zero: a destra cio' che alza il voto."""
    w = min(50.0, 50.0 * abs(points) / scale)
    if abs(points) < 0.005:
        return '<span class="lb"></span>'
    if points > 0:
        return (f'<span class="lb"><i class="up" style="left:50%;width:{w:.1f}%">'
                f'</i></span>')
    return (f'<span class="lb"><i class="dn" style="left:{50 - w:.1f}%;'
            f'width:{w:.1f}%"></i></span>')


def _histogram(series):
    """Barre raggruppate: quante volte esce ogni voto, per ciascuno dei tre."""
    W, H = 880, 250
    L, R, T, B = 44, 8, 18, 46
    n_bins, n_ser = len(HIST_BINS), len(series)
    counts, legend = [], []
    for key, label, vals in series:
        c = {b: 0 for b in HIST_BINS}
        for v in vals:
            c[_clamp_bin(v, HIST_BINS)] += 1
        tot = max(len(vals), 1)
        counts.append((key, label, {b: 100 * c[b] / tot for b in HIST_BINS}, c))
        mu, sd = _mean_sd(vals)
        legend.append(f'<span class="lg"><i class="sw {key}"></i>{esc(label)}'
                      f'<b>media {f2(mu)}</b><em>σ {f2(sd)}</em></span>')
    top = max(max(p.values()) for _, _, p, _ in counts)
    top = math.ceil(top / 5) * 5

    def X(i):
        return L + (W - L - R) * i / n_bins

    def Y(p):
        return T + (H - T - B) * (1 - p / top)

    parts = []
    for g in range(0, top + 1, 5 if top <= 40 else 10):
        y = Y(g)
        parts.append(f'<line class="gl" x1="{L}" y1="{y:.1f}" x2="{W-R}" y2="{y:.1f}"/>')
        parts.append(f'<text class="ax" x="{L-8}" y="{y+3.5:.1f}" text-anchor="end">{g}%</text>')
    gw = (W - L - R) / n_bins
    bw = (gw - 6) / n_ser
    peak = max((p[b], si, b) for si, (_, _, p, _) in enumerate(counts) for b in HIST_BINS)
    for si, (key, label, pct, cnt) in enumerate(counts):
        for bi, b in enumerate(HIST_BINS):
            x = X(bi) + 3 + si * bw
            y, h = Y(pct[b]), Y(0) - Y(pct[b])
            if h < 0.4:
                continue
            edge = "≤" if bi == 0 else ("≥" if bi == n_bins - 1 else "")
            parts.append(
                f'<rect class="bh {key}" x="{x:.1f}" y="{y:.1f}" width="{bw-2:.1f}" '
                f'height="{h:.1f}" rx="2"><title>{esc(label)} · {edge}{b:g} · '
                f'{cnt[b]} voti ({pct[b]:.1f}%)</title></rect>')
    _, psi, pb = peak
    px = X(HIST_BINS.index(pb)) + 3 + psi * bw + (bw - 2) / 2
    parts.append(f'<text class="pk" x="{px:.1f}" y="{Y(peak[0])-6:.1f}" '
                 f'text-anchor="middle">{peak[0]:.0f}%</text>')
    for bi, b in enumerate(HIST_BINS):
        lab = ("≤" if bi == 0 else ("≥" if bi == n_bins - 1 else "")) + f"{b:g}"
        parts.append(f'<text class="ax" x="{X(bi)+gw/2:.1f}" y="{H-B+18}" '
                     f'text-anchor="middle">{lab}</text>')
    parts.append(f'<text class="axt" x="{(L+W-R)/2:.0f}" y="{H-8}" '
                 f'text-anchor="middle">voto base</text>')
    return (f'<figure class="chart"><svg viewBox="0 0 {W} {H}" role="img" '
            f'aria-label="distribuzione dei voti">{"".join(parts)}</svg>'
            f'<figcaption class="legend">{"".join(legend)}</figcaption></figure>')


def _matrix(rows, key, title):
    """Matrice di accordo: nostro voto (righe) x voto esterno (colonne)."""
    grid = defaultdict(int)
    for r in rows:
        grid[(_clamp_bin(r["our"], MATRIX_BINS), _clamp_bin(r[key], MATRIX_BINS))] += 1
    mx = max(grid.values()) if grid else 1
    tot = sum(grid.values()) or 1
    head = "".join(f'<th class="num">{b:g}</th>' for b in MATRIX_BINS)
    body = []
    for our in reversed(MATRIX_BINS):
        cells = []
        for ext in MATRIX_BINS:
            n = grid.get((our, ext), 0)
            if not n:
                cells.append('<td class="mc z"></td>')
                continue
            # rampa sequenziale a una sola tinta: piu' scuro = piu' casi
            step = min(6, max(1, math.ceil(6 * (n / mx) ** 0.45)))
            diag = " dg" if our == ext else ""
            cells.append(f'<td class="mc s{step}{diag}" title="nostro {our:g} · '
                         f'loro {ext:g} · {n} voti ({100*n/tot:.1f}%)">{n}</td>')
        body.append(f'<tr><th class="num rh">{our:g}</th>{"".join(cells)}</tr>')
    return (f'<figure class="mx"><figcaption>{esc(title)}</figcaption>'
            f'<div class="tw"><table class="mxt"><thead><tr><th class="rh"></th>{head}'
            f'</tr></thead>'
            f'<tbody>{"".join(body)}</tbody></table></div>'
            f'<div class="mxleg"><span>pochi</span>'
            f'{"".join(f"<i class=s{i}></i>" for i in range(1, 7))}'
            f'<span>molti</span></div></figure>')


def _foot_matchday(cov):
    return (f"{cov['tot']} righe di foglio · agganciate al nostro archivio "
            f"{cov['matched']} · voto da entrambi {cov['both']} · solo nostro "
            f"{cov['our_only']} · solo loro {cov['ext_only']} · s.v. per entrambi "
            f"{cov['neither']} · eventi discordi {cov['ev_mismatch']}. "
            f"'6*' nel foglio e' il voto d'ufficio: qui vale s.v. e non entra nei conti.")


def _foot_index(cov, sf, ss):
    return (
        f"Copertura: {cov['matched']} righe su {cov['tot']} agganciate al nostro "
        f"archivio ({100*cov['matched']/max(cov['tot'],1):.1f}%); confronti utili "
        f"{sf['n']} sulla Redazione e {ss['n']} sullo Statistico. Escluse le righe "
        f"in cui uno dei due dice s.v. (per noi meno di {cr.MIN_MINUTES_RATED}′ o "
        f"{cr.MIN_TOUCHES_RATED} palloni giocati; per loro il '6*'). "
        f"Il nostro voto base e' 6 + {cr.VOTE_SPREAD_K} · (min/(min+"
        f"{cr.SHRINKAGE_MINUTES})) · z(indice) nel ruolo, poi mitigazione al "
        f"risultato e correzioni per espulsione/autogol/rigore sbagliato, "
        f"arrotondato a mezzo voto. MAE = scarto medio assoluto; bias positivo = "
        f"votiamo piu' alto di loro.")


# ------------------------------------------------------------------------ CSS

_CSS = """
/* fondo neutro caldo come gli altri report del progetto; le tinte dei DATI sono
   la palette categoriale validata (blu/arancio/verdeacqua) — blu = redazione,
   verdeacqua = statistico, arancio = noi. Δ usa una scala divergente arancio
   (siamo piu' alti) / blu (siamo piu' bassi) con il grigio al centro. */
:root{--bg:#fbfaf8;--fg:#20211d;--mut:#78756c;--fai:#a29e94;--line:#e7e4dd;
--card:#fff;--hov:#f4f2ed;--red:#2a78d6;--stat:#1baf7a;--our:#eb6834;
--warn:#b4530f;--seq:#2a78d6}
@media(prefers-color-scheme:dark){:root{--bg:#12120f;--fg:#eceae4;--mut:#98948a;
--fai:#6f6c64;--line:#2a2823;--card:#1a1915;--hov:#232219;--red:#3987e5;
--stat:#199e70;--our:#d95926;--warn:#e08a4a;--seq:#3987e5}}
:root[data-theme=dark]{--bg:#12120f;--fg:#eceae4;--mut:#98948a;--fai:#6f6c64;
--line:#2a2823;--card:#1a1915;--hov:#232219;--red:#3987e5;--stat:#199e70;
--our:#d95926;--warn:#e08a4a;--seq:#3987e5}
:root[data-theme=light]{--bg:#fbfaf8;--fg:#20211d;--mut:#78756c;--fai:#a29e94;
--line:#e7e4dd;--card:#fff;--hov:#f4f2ed;--red:#2a78d6;--stat:#1baf7a;
--our:#eb6834;--warn:#b4530f;--seq:#2a78d6}
*{box-sizing:border-box}
body{margin:0;background:var(--bg);color:var(--fg);padding:28px 24px 60px;
font:15px/1.5 ui-sans-serif,-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;
font-variant-numeric:tabular-nums;-webkit-font-smoothing:antialiased}
.wrap{max-width:1180px;margin:0 auto}
a{color:inherit}
h1{font-size:24px;line-height:1.2;margin:0 0 6px;letter-spacing:-.3px}
.h1sub{color:var(--mut);font-weight:400;font-size:15px;letter-spacing:0;margin-left:10px}
.sub{color:var(--mut);margin:0 0 20px;font-size:13.5px;max-width:78ch}
.sub2{color:var(--mut);margin:-6px 0 14px;font-size:12.5px;max-width:82ch}
.eyebrow{text-transform:uppercase;letter-spacing:.9px;font-size:11px;font-weight:670;
color:var(--mut);margin:34px 0 12px}
.nav{display:flex;gap:14px;margin-bottom:14px;font-size:12.5px}
.nv{color:var(--mut);text-decoration:none;border:1px solid var(--line);
background:var(--card);padding:4px 10px;border-radius:7px}
.nv:hover{color:var(--fg);border-color:var(--fai)}
/* --- riquadri numerici --- */
.tiles{display:grid;grid-template-columns:repeat(auto-fit,minmax(190px,1fr));
gap:11px;margin:0 0 26px}
.tile{background:var(--card);border:1px solid var(--line);border-radius:11px;
padding:11px 14px 12px;display:flex;flex-direction:column;gap:1px}
.tl{font-size:11px;text-transform:uppercase;letter-spacing:.6px;color:var(--mut);
font-weight:640}
.tv{font-size:27px;font-weight:700;letter-spacing:-.6px;line-height:1.15}
.tn{font-size:11.5px;color:var(--mut)}
/* --- barra filtri --- */
.bar{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin:0 0 18px;
position:sticky;top:0;background:var(--bg);padding:9px 0;z-index:3;
border-bottom:1px solid var(--line)}
.q{font:inherit;font-size:13px;padding:6px 11px;border-radius:8px;
border:1px solid var(--line);background:var(--card);color:var(--fg);min-width:230px}
.q:focus{outline:2px solid var(--our);outline-offset:1px}
.chips{display:flex;gap:6px;flex-wrap:wrap}
.chip{font:inherit;font-size:12px;padding:5px 11px;border-radius:8px;cursor:pointer;
border:1px solid var(--line);background:var(--card);color:var(--mut)}
.chip:hover{color:var(--fg)}
.chip.on{border-color:var(--our);color:var(--our);font-weight:640;
background:color-mix(in srgb,var(--our) 9%,var(--card))}
.count{font-size:12px;color:var(--mut);margin-left:auto}
/* --- blocco squadra --- */
.tm{margin:0 0 20px}
.tm h2{font-size:14px;margin:0 0 5px;font-weight:680;letter-spacing:.1px;
display:flex;align-items:baseline;gap:9px}
.fx{font-weight:400;font-size:12px;color:var(--mut)}
table.grid{width:100%;border-collapse:collapse;font-size:13.5px;
background:var(--card);border:1px solid var(--line);border-radius:10px;
overflow:hidden;table-layout:fixed}
.grid th{text-align:left;color:var(--mut);font-weight:620;font-size:10.5px;
text-transform:uppercase;letter-spacing:.4px;padding:6px 8px;
border-bottom:1px solid var(--line);background:var(--card)}
.grid td{padding:5px 8px;border-bottom:1px solid var(--line)}
.grid tr:last-child td{border-bottom:none}
tr.r{cursor:pointer}
tr.r:hover td{background:var(--hov)}
tr.r:focus{outline:2px solid var(--our);outline-offset:-2px}
tr.r.open td{background:var(--hov)}
.c-r{width:26px;color:var(--mut);font-size:11.5px}
/* nome a larghezza fissa e "eventi" a prendersi il resto: cosi' i voti stanno
   ACCANTO al nome invece che in fondo a una distesa vuota, e la colonna nome non
   si strozza quando la tabella scorre dentro il suo riquadro su schermo stretto */
.c-n{width:190px;font-weight:560;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.c-v{width:62px;text-align:right}
.c-our{border-left:2px solid color-mix(in srgb,var(--our) 45%,transparent)}
tr.r .c-our b{border-bottom:1px dotted var(--fai)}
.c-d{width:66px;text-align:right;font-size:12.5px;color:var(--mut)}
.c-d.z{color:var(--fai)}
/* zebra: la riga-dettaglio si intercala a quella del giocatore, quindi la
   striatura prende una riga ogni due (4n+1), non ogni due figli */
.grid tbody tr.r:nth-child(4n+1) td{background:color-mix(in srgb,var(--fg) 4%,transparent)}
.grid tbody tr.r:nth-child(4n+1):hover td,.grid tbody tr.r.open td{background:var(--hov)}
.c-m{width:46px;text-align:right;color:var(--mut);font-size:12px}
.c-e{width:auto;min-width:180px}
.grid th{white-space:nowrap}
.sv{color:var(--fai);font-size:11.5px}
.p1{background:color-mix(in srgb,var(--our) 9%,transparent)}
.p2{background:color-mix(in srgb,var(--our) 19%,transparent);color:var(--fg)}
.p3{background:color-mix(in srgb,var(--our) 32%,transparent);color:var(--fg);font-weight:640}
.n1{background:color-mix(in srgb,var(--red) 9%,transparent)}
.n2{background:color-mix(in srgb,var(--red) 19%,transparent);color:var(--fg)}
.n3{background:color-mix(in srgb,var(--red) 32%,transparent);color:var(--fg);font-weight:640}
.rb{font-size:9.5px;color:var(--mut);border:1px solid var(--line);border-radius:4px;
padding:0 3px;margin-left:5px;vertical-align:1px;letter-spacing:.3px}
.nm{color:var(--fai);margin-left:5px;font-size:11px}
.ev{font-size:10px;padding:1px 5px;border-radius:20px;margin-right:3px;
border:1px solid var(--line);color:var(--mut);white-space:nowrap}
.ev.g{border-color:color-mix(in srgb,var(--our) 50%,var(--line));color:var(--our)}
.ev.rd,.ev.o{border-color:color-mix(in srgb,var(--warn) 50%,var(--line));color:var(--warn)}
tr.evx .c-n:after{content:"⚠";color:var(--warn);font-size:10px;margin-left:5px}
/* --- dettaglio (la spiegazione) --- */
tr.d td{background:var(--hov);padding:0;border-bottom:1px solid var(--line)}
.why{padding:11px 14px 13px}
.lead{margin:0 0 9px;font-size:13.5px;max-width:88ch}
.ledger{display:flex;flex-direction:column;gap:2px;max-width:640px}
.lr{display:grid;grid-template-columns:1fr 190px 74px;align-items:center;gap:10px;
font-size:12.5px}
.lr .ll{color:var(--mut);text-align:right;overflow:hidden;text-overflow:ellipsis;
white-space:nowrap}
.lr .lv{text-align:right;font-variant-numeric:tabular-nums}
.lr.o .ll,.lr.o .lv{color:var(--fai);font-style:italic}
.lr.t{border-top:1px solid var(--line);margin-top:4px;padding-top:4px}
.lr.t .ll{color:var(--fg);font-weight:640}
.lb{position:relative;display:block;height:9px;background:linear-gradient(
to right,transparent calc(50% - .5px),var(--line) calc(50% - .5px),
var(--line) calc(50% + .5px),transparent calc(50% + .5px))}
.lb i{position:absolute;top:0;height:9px;display:block}
.lb i.up{background:var(--our);border-radius:0 3px 3px 0}
.lb i.dn{background:var(--red);border-radius:3px 0 0 3px}
.meta{margin:9px 0 0;font-size:12px;color:var(--mut)}
.warn{margin:7px 0 0;font-size:12px;color:var(--warn)}
/* --- tutte le feature (aperta al click, costruita dal payload) --- */
details.ft{margin:11px 0 0}
details.ft>summary{cursor:pointer;font-size:12px;color:var(--mut);
display:inline-block;border:1px solid var(--line);border-radius:7px;
padding:4px 10px;background:var(--card);list-style:none}
details.ft>summary::-webkit-details-marker{display:none}
details.ft>summary:before{content:"▸ ";font-size:10px}
details.ft[open]>summary:before{content:"▾ "}
details.ft>summary:hover{color:var(--fg);border-color:var(--fai)}
details.ft[open]>summary{margin-bottom:9px}
table.ftt{font-size:11.5px;min-width:820px}
table.ftt td,table.ftt th{padding:3px 8px}
/* niente maiuscolo automatico: la sigma minuscola diventerebbe Σ, che vuol dire
   un'altra cosa (somma) proprio in una tabella piena di somme */
table.ftt th{text-transform:none;font-size:10.5px;letter-spacing:.2px}
table.ftt .k{font-family:ui-monospace,SFMono-Regular,Menlo,monospace;font-size:11px}
table.ftt .lb{color:var(--mut)}
table.ftt .ki{color:var(--fai);font-size:10px;letter-spacing:.4px}
table.ftt .n{text-align:right;width:78px}
table.ftt .dim{color:var(--fai)}
table.ftt .pt{font-weight:700}
table.ftt tr.up .pt{color:var(--our)}
table.ftt tr.dn .pt{color:var(--red)}
table.ftt tr.ze td{color:var(--fai)}
table.ftt tr.ze .k{color:var(--mut)}
table.ftt tfoot th{text-align:right;text-transform:none;font-size:11px;
letter-spacing:0;color:var(--fg);border-top:1.5px solid var(--line);border-bottom:none}
table.ftt tfoot td{border-top:1.5px solid var(--line);border-bottom:none}
.ftnote{color:var(--mut);font-size:11px;margin:8px 0 0;max-width:100ch;line-height:1.6}
/* --- sezione disaccordi: quattro giudici affiancati --- */
table.dv{min-width:760px}
table.dv td:first-child{width:230px;font-weight:560}
table.dv td.dim{color:var(--mut);font-size:12px}
table.dv a{text-decoration:none;border-bottom:1px solid var(--line)}
table.dv a:hover{border-color:var(--our)}
table.dv .tg{white-space:nowrap}
#dvrole,#dvdir{gap:6px}
#dvdir{margin-left:14px}
/* marca delle famiglie fuse: la riga del riepilogo e le righe che la compongono
   portano lo stesso segno, cosi' il numero si ritrova */
.mrg{font-size:9.5px;letter-spacing:.3px;color:var(--mut);border:1px solid var(--line);
border-radius:4px;padding:0 4px;margin-left:6px;vertical-align:1px;white-space:nowrap;
background:color-mix(in srgb,var(--fg) 4%,transparent);cursor:help}
.lr .ll .mrg{margin-left:5px}
/* --- tabelle dell'indice --- */
.tbl{width:100%;border-collapse:collapse;font-size:13px;background:var(--card);
border:1px solid var(--line);border-radius:10px;overflow:hidden}
.tbl th{text-align:left;color:var(--mut);font-weight:620;font-size:10.5px;
text-transform:uppercase;letter-spacing:.4px;padding:7px 9px;
border-bottom:1px solid var(--line)}
.tbl td{padding:6px 9px;border-bottom:1px solid var(--line)}
.tbl tr:last-child td{border-bottom:none}
.tbl .num{text-align:right;width:78px}
.tbl .strong{font-weight:700;color:var(--our)}
.tbl a{font-weight:600;text-decoration:none;border-bottom:1px solid var(--line)}
.tbl a:hover{border-color:var(--our)}
.tg{color:var(--mut);font-size:11px;margin-left:7px;font-weight:400;white-space:nowrap}
.top td:first-child{width:210px}
.why-inline{color:var(--mut);font-size:12px}
.gd td:first-child{width:130px}
.mbar{width:130px}
.mbar span{display:block;height:7px;border-radius:0 3px 3px 0;background:var(--our);
opacity:.75}
/* --- grafico --- */
.chart{margin:0 0 8px;background:var(--card);border:1px solid var(--line);
border-radius:11px;padding:12px 14px 10px}
.chart svg{width:100%;height:auto;display:block;overflow:visible}
.gl{stroke:var(--line);stroke-width:1}
.ax{fill:var(--mut);font-size:10.5px}
.axt{fill:var(--mut);font-size:10.5px;letter-spacing:.4px;text-transform:uppercase}
.pk{fill:var(--mut);font-size:10.5px;font-weight:640}
.bh.red{fill:var(--red)}.bh.stat{fill:var(--stat)}.bh.our{fill:var(--our)}
.bh{transition:opacity .1s}.bh:hover{opacity:.75}
.legend{display:flex;gap:20px;flex-wrap:wrap;margin-top:9px;font-size:11.5px;
color:var(--mut)}
.lg{display:flex;align-items:center;gap:6px}
.lg b{color:var(--fg);font-weight:640}
.lg em{font-style:normal}
.sw{width:10px;height:10px;border-radius:3px;display:inline-block}
.sw.red{background:var(--red)}.sw.stat{background:var(--stat)}.sw.our{background:var(--our)}
/* --- matrice --- */
.mx-wrap{display:flex;gap:18px;flex-wrap:wrap}
.mx{margin:0;flex:1;min-width:330px;background:var(--card);border:1px solid var(--line);
border-radius:11px;padding:11px 13px 9px}
.mx figcaption{font-size:12px;font-weight:640;margin-bottom:7px}
.mxt{border-collapse:separate;border-spacing:2px;width:100%;font-size:10.5px}
.mxt th{color:var(--mut);font-weight:600;font-size:10px;padding:1px 2px}
.mxt .rh{width:24px;text-align:right}
.mc{text-align:center;padding:3px 2px;border-radius:3px;color:var(--mut);
background:color-mix(in srgb,var(--seq) 4%,transparent)}
.mc.z{color:transparent}
.mc.s1{background:color-mix(in srgb,var(--seq) 9%,transparent)}
.mc.s2{background:color-mix(in srgb,var(--seq) 20%,transparent)}
.mc.s3{background:color-mix(in srgb,var(--seq) 33%,transparent)}
.mc.s4{background:color-mix(in srgb,var(--seq) 48%,transparent);color:var(--bg)}
.mc.s5{background:color-mix(in srgb,var(--seq) 68%,transparent);color:var(--bg)}
.mc.s6{background:var(--seq);color:var(--bg);font-weight:640}
.mc.dg{box-shadow:inset 0 0 0 1px var(--fai)}
.mxleg{display:flex;align-items:center;gap:3px;margin-top:7px;font-size:10px;
color:var(--mut)}
.mxleg i{width:14px;height:9px;border-radius:2px;display:inline-block}
.mxleg i.s1{background:color-mix(in srgb,var(--seq) 9%,transparent)}
.mxleg i.s2{background:color-mix(in srgb,var(--seq) 20%,transparent)}
.mxleg i.s3{background:color-mix(in srgb,var(--seq) 33%,transparent)}
.mxleg i.s4{background:color-mix(in srgb,var(--seq) 48%,transparent)}
.mxleg i.s5{background:color-mix(in srgb,var(--seq) 68%,transparent)}
.mxleg i.s6{background:var(--seq)}
/* --- senza voto: matrice 2x2 e barre appaiate --- */
.tbl.sv{border:none}
.tbl.sv th{text-transform:none;font-size:11.5px;letter-spacing:0;color:var(--fg)}
.tbl.sv td.ok{color:var(--mut)}
.tbl.sv td.ko{font-weight:700;color:var(--our)}
.mxnote{color:var(--mut);font-size:11px;margin:8px 0 0;line-height:1.5}
.bul{width:190px}
.bul span{display:block;height:6px;border-radius:0 3px 3px 0;margin:2px 0}
.bul .t{background:var(--red);opacity:.55}
.bul .o{background:var(--our)}
.sub2.mt{margin-top:20px}
.kt{color:var(--red);font-weight:600}.ko2{color:var(--our);font-weight:600}
code{font:inherit;font-size:.95em;background:color-mix(in srgb,var(--fg) 7%,transparent);
padding:1px 4px;border-radius:4px}
/* --- code --- */
.note{margin:18px 0 0;font-size:12px;color:var(--mut)}
.note summary{cursor:pointer}
.note p{max-width:100ch;line-height:1.7}
.note i{font-style:normal;color:var(--fai)}
.foot{color:var(--mut);font-size:11.5px;margin-top:30px;max-width:96ch;line-height:1.65}
/* Le tabelle larghe scorrono DENTRO il loro riquadro: la pagina non deve mai
   scorrere in orizzontale. min-width sotto la quale le colonne diventano
   illeggibili — sopra, il contenitore e' piu' largo e non si nota. */
.tw{overflow-x:auto;-webkit-overflow-scrolling:touch}
.tw table.grid{min-width:640px}
.tw table.tbl{min-width:560px}
.tw table.mxt{min-width:300px}
@media(max-width:720px){
body{padding:18px 12px 40px}
.lr{grid-template-columns:1fr 90px 62px}
.h1sub{display:block;margin-left:0}
.mx{min-width:280px}
.tiles{grid-template-columns:1fr 1fr}
}
"""

_JS = """
(function(){
 var rows=[].slice.call(document.querySelectorAll('tr.r'));
 function detail(tr){var n=tr.nextElementSibling;return (n&&n.classList.contains('d'))?n:null;}
 function toggle(tr,force){
   var d=detail(tr); if(!d) return;
   var open=(force===undefined)?d.hidden:force;
   d.hidden=!open; tr.classList.toggle('open',open);
 }
 rows.forEach(function(tr){
   tr.addEventListener('click',function(){toggle(tr);});
   tr.addEventListener('keydown',function(e){
     if(e.key==='Enter'||e.key===' '){e.preventDefault();toggle(tr);}
   });
 });
 var q=document.getElementById('q'), cnt=document.getElementById('count'),
     chips=[].slice.call(document.querySelectorAll('#filters .chip')), thr='0';
 // --- filtri della sezione "disaccordo grande" (solo sull'indice) ----------
 (function(){
   var rr=document.getElementById('dvrole'), dd=document.getElementById('dvdir');
   if(!rr||!dd) return;
   var trs=[].slice.call(document.querySelectorAll('table.dv tr.r')),
       cnt=document.getElementById('dvcount'), tt=document.getElementById('dvthr'),
       role='', dir='', thr=0;
   function apply(){
     var n=0;
     trs.forEach(function(tr){
       var ok=(!role||tr.dataset.role===role)&&(!dir||tr.dataset.dir===dir)
              &&parseFloat(tr.dataset.gap)>=thr;
       tr.hidden=!ok;
       var d=tr.nextElementSibling;
       if(d&&d.classList.contains('d')&&!ok){d.hidden=true;tr.classList.remove('open');}
       if(ok) n++;
     });
     if(cnt) cnt.textContent=n+' / '+trs.length+' casi';
   }
   function wire(box,key,set){
     [].slice.call(box.querySelectorAll('.chip')).forEach(function(c){
       c.addEventListener('click',function(){
         [].slice.call(box.querySelectorAll('.chip')).forEach(function(x){
           x.classList.remove('on');});
         c.classList.add('on'); set(c.dataset[key]||''); apply();
       });
     });
   }
   wire(rr,'role',function(v){role=v;}); wire(dd,'dir',function(v){dir=v;});
   if(tt) wire(tt,'thr',function(v){thr=parseFloat(v)||0;});
   apply();
 })();
 function norm(s){return (s||'').toLowerCase().normalize('NFD')
   .replace(/[\\u0300-\\u036f]/g,'').replace(/[^a-z0-9]+/g,' ').trim();}
 function apply(){
   var term=norm(q?q.value:''), shown=0;
   rows.forEach(function(tr){
     var ok=(!term||(tr.dataset.n||'').indexOf(term)>=0);
     if(ok&&thr==='sv') ok=(tr.dataset.sv==='1');
     else if(ok&&thr!=='0') ok=(parseFloat(tr.dataset.df)>=parseFloat(thr));
     tr.hidden=!ok; var d=detail(tr); if(d&&!ok){d.hidden=true;tr.classList.remove('open');}
     if(ok) shown++;
   });
   document.querySelectorAll('section.tm').forEach(function(s){
     s.hidden=!s.querySelector('tr.r:not([hidden])');
   });
   if(cnt) cnt.textContent=shown+' / '+rows.length+' giocatori';
 }
 if(q) q.addEventListener('input',apply);
 chips.forEach(function(c){c.addEventListener('click',function(){
   chips.forEach(function(x){x.classList.remove('on');});
   c.classList.add('on'); thr=c.dataset.thr; apply();
 });});
 // --- tabella di TUTTE le feature, costruita al primo click ---------------
 // Il payload porta il vocabolario del canale una volta sola (nome tecnico,
 // descrizione, tipo, peso) e tre numeri per feature per giocatore: valore, z e
 // punti di voto. Indice = peso x z; media di ruolo = indice - punti/per_unit.
 function ftTable(pid){
   var row=(typeof FROWS!=='undefined')?FROWS[pid]:null;
   if(!row) return '<p class="ftnote">dettaglio non disponibile.</p>';
   var keys=FCHAN[row[0]], per=row[1], out=[];
   for(var i=0;i<keys.length;i++){
     var v=row[2+i*3], z=row[3+i*3], p=row[4+i*3], w=keys[i][3];
     var idx=w*z, avg=per?idx-p/per:0;
     // per un evento contato, quanto vale UNA occorrenza: l'unica unita' leggibile
     // dove la sigma e' una frazione di evento (vedi last_man_tackle)
     var one=(keys[i][5]!=null&&per)?keys[i][5]*w*per:null;
     out.push({key:keys[i][0], lab:keys[i][1], kind:keys[i][2], w:w,
               fam:keys[i][4]||'', one:one, v:v, z:z, idx:idx, avg:avg, p:p});
   }
   out.sort(function(a,b){return Math.abs(b.p)-Math.abs(a.p);});
   var tot=out.reduce(function(s,r){return s+r.p;},0);
   // le famiglie fuse: nel riepilogo sopra sono UNA riga, qui sono N
   var fams={};
   out.forEach(function(r){ if(r.fam) fams[r.fam]=(fams[r.fam]||0)+r.p; });
   var body=out.map(function(r){
     var cls=r.p>0.005?'up':(r.p<-0.005?'dn':'ze');
     var fam=r.fam?' <span class="mrg" title="il riepilogo fonde questa famiglia in '
       +'una riga sola: totale '+(fams[r.fam]>=0?'+':'')+fams[r.fam].toFixed(2)
       +'">'+r.fam+'</span>':'';
     return '<tr class="'+cls+'"><td class="k">'+r.key+'</td>'
       +'<td class="lb">'+(r.lab||'')+fam+'</td><td class="ki">'+r.kind+'</td>'
       +'<td class="n">'+r.v.toFixed(2)+'</td>'
       +'<td class="n">'+(r.z>=0?'+':'')+r.z.toFixed(2)+'</td>'
       +'<td class="n">'+(r.w>=0?'+':'')+r.w.toFixed(4)+'</td>'
       +'<td class="n ev">'+(r.one==null?'—':(r.one>=0?'+':'')+r.one.toFixed(2))+'</td>'
       +'<td class="n">'+(r.idx>=0?'+':'')+r.idx.toFixed(3)+'</td>'
       +'<td class="n dim">'+(r.avg>=0?'+':'')+r.avg.toFixed(3)+'</td>'
       +'<td class="n pt">'+(r.p>=0?'+':'')+r.p.toFixed(2)+'</td></tr>';
   }).join('');
   return '<div class="tw"><table class="tbl ftt"><thead><tr>'
     +'<th>feature (nome nel foglio)</th><th>descrizione</th><th>tipo</th>'
     +'<th class="n">valore</th><th class="n">σ</th><th class="n">peso</th>'
     +'<th class="n" title="quanto vale UNA occorrenza, per gli eventi contati">'
     +'1 evento</th>'
     +'<th class="n">indice</th><th class="n">media ruolo</th>'
     +'<th class="n">punti voto</th></tr></thead><tbody>'+body
     +'</tbody><tfoot><tr><th colspan="9">somma delle voci = voto di merito − 6</th>'
     +'<td class="n pt">'+(tot>=0?'+':'')+tot.toFixed(2)+'</td></tr></tfoot>'
     +'</table></div>'
     +'<p class="ftnote">valore = quello che ha fatto (per-90 dove il tipo dice '
     +'PER90, con pavimento a 55′); σ = quanto sopra o sotto la media della '
     +'popolazione, standardizzato e compresso; <b>1 evento</b> = quanto vale UNA '
     +'occorrenza per gli eventi contati, ed è quello da leggere per un evento raro '
     +'(dove una σ è una frazione di occorrenza, quindi il peso “per σ” inganna); '
     +'indice = peso × σ; '
     +'<b>punti voto = (indice − media del suo ruolo) × '+per.toFixed(3)
     +'</b>, cioè quanto quella voce ha spostato il voto. Le voci con punti a 0 '
     +'sono quelle in cui è esattamente nella media del ruolo, non quelle che '
     +'non contano.'
     +(Object.keys(fams).length?' <b>Il riepilogo qui sopra fonde le famiglie '
       +'marcate</b> ('+Object.keys(fams).map(function(f){
           return f+' '+(fams[f]>=0?'+':'')+fams[f].toFixed(2);}).join(', ')
       +') in una riga sola, perché lette separate direbbero assurdità: '
       +'<code>xg_shots</code> è sottratta per costruzione, quindi «tante posizioni '
       +'di tiro conquistate» comparirebbe fra le note negative. Un numero del '
       +'riepilogo che non trova riscontro in nessuna riga qui è la somma delle '
       +'righe con la stessa marca.':'')
     +'</p>';
 }
 document.querySelectorAll('details.ft').forEach(function(dt){
   dt.addEventListener('toggle',function(){
     if(!dt.open || dt.dataset.done) return;
     dt.dataset.done='1';
     dt.querySelector('.ftbody').innerHTML=ftTable(dt.dataset.pid);
   });
 });
 apply();
 if(location.hash){
   var t=document.querySelector(location.hash);
   if(t&&t.classList.contains('r')){toggle(t,true);t.scrollIntoView({block:'center'});}
 }
})();
"""
