"""Robustness probe for the classic *voto puro*: surface the player-matchdays where
our heuristic base vote diverges most from external fantacalcio.it base votes,
role by role, together with the model's own justification (explanation_text + top
additive contributions) and an INDEPENDENT third opinion (SofaScore rating).

The vote is compared against BOTH fantacalcio sheets in one pass:
  * Fantacalcio  — the editorial pagella (human graded);
  * Statistico   — the algorithmic base vote (a cleaner, goal-stripped base).

The high-value output is the JSON dump (``--json PATH``): one record per matched
player-matchday carrying our vote, both external votes, the SofaScore rating,
minutes, goals/assists and the explanation. The report builder then flags the
cases where we disagree with BOTH the external vote AND the SofaScore rating —
the genuine outliers.

    python manage.py voto_puro_discrepancies --json /tmp/discrep.json

Pagelle (our vote + explanation + rating) are sheet-independent, so they are
computed ONCE and matched against every sheet.
"""

from __future__ import annotations

import glob
import json
import math
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError

from realdata.models import Match, MatchAppearance, PlayerTeamStint
from realdata.services.identity import norm_name
from vfoot.services.classic_pagella import (
    get_reference, get_role_averages, pagella_for_match,
)

DEFAULT_DIR = str(Path(settings.VFOOT_DATA_DIR) / "data_fantacalcio" / "2025-2026")
SHEETS = ["Fantacalcio", "Statistico"]
EXTERNAL_TEAMS = {
    "atalanta", "bologna", "cagliari", "como", "cremonese", "fiorentina", "genoa",
    "verona", "hellas verona", "inter", "juventus", "lazio", "lecce", "milan",
    "napoli", "parma", "pisa", "roma", "sassuolo", "torino", "udinese",
}
ROLE_ORDER = {"POR": 0, "DIF": 1, "CEN": 2, "ATT": 3}


def _parse_voto(v):
    if v is None:
        return None
    m = re.search(r"(\d+(?:[.,]\d+)?)", str(v))
    return float(m.group(1).replace(",", ".")) if m else None


def _club_key(name):
    fillers = {"ac", "as", "fc", "ssc", "us", "ss", "hellas", "calcio"}
    return " ".join(t for t in norm_name(name).split() if t not in fillers)


def _pearson(xs, ys):
    n = len(xs)
    if n < 2:
        return float("nan")
    mx, my = sum(xs) / n, sum(ys) / n
    cov = sum((a - mx) * (b - my) for a, b in zip(xs, ys))
    sx = math.sqrt(sum((a - mx) ** 2 for a in xs))
    sy = math.sqrt(sum((b - my) ** 2 for b in ys))
    return cov / (sx * sy) if sx and sy else float("nan")


class Command(BaseCommand):
    help = "Dump voto-puro vs fantacalcio (both sheets) + SofaScore discrepancies for the report builder."

    def add_arguments(self, parser):
        parser.add_argument("--competition-season", type=int, default=2)
        parser.add_argument("--dir", default=DEFAULT_DIR)
        parser.add_argument("--json", default=None, help="Write full paired rows here.")

    # -- external parsing -------------------------------------------------

    def _parse_file(self, path, sheet):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[sheet]
        out, team = [], None
        for r in ws.iter_rows(values_only=True):
            c0 = r[0]
            if isinstance(c0, str) and norm_name(c0) in EXTERNAL_TEAMS:
                team = c0
                continue
            if not isinstance(c0, (int, float)) or c0 == "Cod.":
                continue
            out.append({"team": team, "ruolo": r[1], "nome": r[2],
                        "voto": _parse_voto(r[3]), "gf": r[4] or 0, "ass": r[12] or 0})
        wb.close()
        return out

    def _our_team_index(self, cs_id):
        teams = (PlayerTeamStint.objects
                 .filter(team_season__competition_season_id=cs_id)
                 .values_list("team_season__team__name", flat=True).distinct())
        return {_club_key(t): t for t in set(teams)}

    # letters unicode NFKD leaves alone, but our DB and fantacalcio disagree on:
    # Turkish ı, Nordic ø/å, Icelandic ð/þ, Slavic ł/đ...
    _FOLD = {"ı": "i", "İ": "i", "ø": "o", "å": "a", "ð": "d", "þ": "t",
             "đ": "d", "ł": "l", "æ": "ae", "œ": "oe", "ß": "ss"}

    @classmethod
    def _afold(cls, s):
        """Aggressive fold to bare a-z (accents + the letters above + apostrophes),
        so 'Østigård'=='Ostigard', 'Yıldız'=='Yildiz', "N'Dicka"=='Ndicka'."""
        import unicodedata
        s = "".join(cls._FOLD.get(c, c) for c in (s or "").lower())
        s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
        return re.sub(r"[^a-z ]", " ", s)

    @classmethod
    def _surname_and_initial(cls, name):
        """(surname, first-name abbreviation) for a name in any format we see:
        'Nico Paz', 'N. Paz', 'Paz N.', 'Pellegrini Lu.', 'Ederson D.S.', 'De Ketelaere'.

        Fantacalcio puts the surname FIRST and appends the first name abbreviated
        ('Lu.', 'Se.') only to disambiguate a shared surname. A trailing run of
        abbreviations (a token ending '.' or a lone letter) is stripped to recover
        the surname; the leftmost abbreviation (folded, e.g. 'lu') then disambiguates
        by the candidate's first-name PREFIX."""
        toks = (name or "").split()
        stripped = []
        while len(toks) > 1 and (toks[-1].endswith(".") or len(toks[-1].rstrip(".")) == 1):
            stripped.insert(0, toks.pop())
        surname = cls._afold(toks[-1]).replace(" ", "") if toks else ""
        abbr = (cls._afold(stripped[0]).replace(" ", "") if stripped
                else (cls._afold(toks[0]).replace(" ", "")[:1] if len(toks) > 1 and toks[0] else ""))
        return surname, abbr

    def _our_player_index(self, cs_id):
        idx = defaultdict(list)
        pid_team = {}
        pid_first = defaultdict(set)  # pid -> folded first names, for prefix disambiguation
        # Index on the teams a player ACTUALLY appeared for (not the single open TM
        # stint): this follows mid-season transfers per match AND catches players who
        # played but are missing from the TM roster — both of which we only compare
        # because they have a voto puro (i.e. they appeared).
        rows = (MatchAppearance.objects
                .filter(match__competition_season_id=cs_id)
                .values_list("player_id", "team_season__team__name",
                             "player__full_name", "player__short_name")
                .distinct())
        for pid, team, full, short in rows:
            pid_team.setdefault(pid, team)
            keys = set()
            for nm in (full, short):
                surn, _ = self._surname_and_initial(nm)
                if surn:
                    keys.add(surn)
            ft = self._afold(full or "").split()  # our names are 'First Last'
            if ft:
                pid_first[pid].add(ft[0])
            for k in keys:
                idx[(team, k)].append(pid)
        return idx, pid_team, pid_first

    def _match_external(self, files, sheet, team_map, pidx, pid_first):
        """{(gd, pid): {voto, gf, ass}} for one sheet. Head coaches (ruolo ALL) are
        skipped: fantacalcio grades them, we do not rate them."""
        gd_re = re.compile(r"Giornata_(\d+)")
        out, unmatched = {}, 0
        for f in files:
            mm = gd_re.search(f)
            if not mm:
                continue
            gd = int(mm.group(1))
            for e in self._parse_file(f, sheet):
                if e["voto"] is None or str(e["ruolo"]).upper() == "ALL":
                    continue
                our_team = team_map.get(_club_key(e["team"] or ""))
                if not our_team:
                    continue
                surn, abbr = self._surname_and_initial(e["nome"])
                cands = pidx.get((our_team, surn), [])
                # shared surname on one team: the appended abbreviation is a prefix
                # of the right player's first name (Lu.->Luca, Lo.->Lorenzo).
                if len(cands) > 1 and abbr:
                    narrowed = [pid for pid in cands
                                if any(fn.startswith(abbr) for fn in pid_first.get(pid, ()))]
                    if narrowed:
                        cands = narrowed
                if len(cands) != 1:
                    unmatched += 1
                    continue
                out[(gd, cands[0])] = {"voto": e["voto"], "gf": e["gf"], "ass": e["ass"]}
        return out, unmatched

    # -- main -------------------------------------------------------------

    def handle(self, *args, **opts):
        cs_id = opts["competition_season"]
        files = sorted(glob.glob(f"{opts['dir']}/*.xlsx"))
        if not files:
            raise CommandError(f"No .xlsx in {opts['dir']}")

        rows = self.discrepancy_rows(cs_id, files)
        self._quick_corr(rows)
        if opts["json"]:
            Path(opts["json"]).write_text(json.dumps(rows, ensure_ascii=False))
            self.stdout.write(f"\nWrote {len(rows)} rows -> {opts['json']}")

    def discrepancy_rows(self, cs_id, files):
        """Unified our-vs-external rows for a season. Reusable by other commands
        (e.g. build_voto_tuner) so the external-sheet parsing lives in one place."""
        ref = get_reference(cs_id)
        averages = get_role_averages(cs_id)
        team_map = self._our_team_index(cs_id)
        pidx, pid_team, pid_ini = self._our_player_index(cs_id)

        # our side (sheet-independent): compute pagella ONCE
        self.stdout.write("Computing voto puro + explanation for every match…")
        ours, rating_by = {}, {}
        for m in Match.objects.filter(competition_season_id=cs_id):
            pag = pagella_for_match(m, ref, averages=averages)
            for side in ("home", "away"):
                for ln in pag[side]["starters"] + pag[side]["bench"]:
                    if ln.get("sv") or ln.get("voto_puro") is None:
                        continue
                    exp = ln.get("explanation") or {}
                    ours[(m.matchday, ln["player_id"])] = {
                        "our": ln["voto_puro"], "role": ln["role"],
                        "minutes": ln["minutes"], "goals": ln["events"]["goals"],
                        "assists": ln["events"]["assists"], "name": ln["name"],
                        "explanation_text": ln.get("explanation_text", ""),
                        "contributions": exp.get("contributions") or [],
                        "other_points": exp.get("other_points", 0.0),
                    }
        for pid, md, raw in (MatchAppearance.objects
                             .filter(match__competition_season_id=cs_id)
                             .values_list("player_id", "match__matchday", "raw_stats")):
            if raw and raw.get("rating"):
                rating_by[(md, pid)] = raw["rating"]

        # match both sheets
        ext = {}
        for sheet in SHEETS:
            ext[sheet], unm = self._match_external(files, sheet, team_map, pidx, pid_ini)
            self.stdout.write(f"  {sheet}: matched {len(ext[sheet])} "
                              f"(unmatched names {unm})")

        # unified rows
        rows = []
        for (gd, pid), o in ours.items():
            fanta = ext["Fantacalcio"].get((gd, pid))
            stat = ext["Statistico"].get((gd, pid))
            if fanta is None and stat is None:
                continue
            rows.append({
                "gd": gd, "pid": pid, "name": o["name"],
                "team": pid_team.get(pid, ""), "role": o["role"],
                "minutes": o["minutes"], "goals": o["goals"], "assists": o["assists"],
                "our": o["our"],
                "fanta": fanta["voto"] if fanta else None,
                "statistico": stat["voto"] if stat else None,
                "sofa": rating_by.get((gd, pid)),
                "explanation_text": o["explanation_text"],
                "contributions": o["contributions"],
                "other_points": o["other_points"],
            })
        return rows

    def _quick_corr(self, rows):
        w = self.stdout.write
        w(f"\nUnified rows (our + >=1 external): {len(rows)}")
        w("Per-role corr with SofaScore rating (non-scorers):")
        w(f"  {'role':<5}{'n':>6}  {'our~sofa':>9}{'fanta~sofa':>11}{'stat~sofa':>10}")
        by_role = defaultdict(list)
        for r in rows:
            by_role[r["role"]].append(r)
        for role in sorted(by_role, key=lambda x: ROLE_ORDER.get(x, 9)):
            g = [r for r in by_role[role]
                 if not r["goals"] and r["sofa"] is not None]
            def c(key):
                pairs = [(r["our"] if key == "our" else r[key], r["sofa"])
                         for r in g if r.get(key if key != "our" else "our") is not None]
                return _pearson([a for a, _ in pairs], [b for _, b in pairs])
            w(f"  {role:<5}{len(g):>6}  {c('our'):>9.3f}{c('fanta'):>11.3f}"
              f"{c('statistico'):>10.3f}")
