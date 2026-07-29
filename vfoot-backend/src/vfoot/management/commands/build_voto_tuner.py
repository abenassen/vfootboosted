"""Build the interactive voto-puro weight tuner spreadsheet.

    python manage.py build_voto_tuner                 # -> <repo>/voto_puro_tuner.xlsx
    python manage.py build_voto_tuner --season 2 --cases 24 --out /path/file.xlsx

An analyst opens the sheet and edits the WEIGHTS (Tuner!C): the voto puro of every
selected case recomputes live via spreadsheet formulas (μ/σ per role are rebuilt
from a stored covariance matrix, so no Python round-trip). It replaces a pile of
throwaway scripts — hence a real, versioned command rather than a temp file.

Design notes:
  * features = the deployed outfield vector (TOTAL + PER90 + defensive exposure)
    plus the shot-outcome detail (post/goal/save/miss/block, weight 0) so the
    SGA_Pali paradigm can be reconstructed by hand from the weights.
  * the √ compression toggle applies √ ONLY to PER90 features; TOTAL/exposure/shot
    detail stay linear (integer counts stay integer, xG/xGOT stay linear).
  * cases are chosen for DISCREPANCY with fantacalcio, and deliberately span
    scorers and heavy defeats — fantacalcio ties the base vote to the result, our
    voto puro does not, and that is exactly where they part ways.
"""
from __future__ import annotations

import glob
import math
from collections import defaultdict
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter as col
from openpyxl.worksheet.datavalidation import DataValidation

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError

import vfoot.services.classic_rating as cr
from realdata.models import Match, MatchAppearance, Player, Team
from vfoot.management.commands.voto_puro_discrepancies import Command as DiscCmd

OUT_ROLES = ["DIF", "CEN", "ATT"]
SHOTMAP = {"post": "shots_post", "goal": "shots_goal", "save": "shots_saved",
           "miss": "shots_off", "block": "shots_blocked"}
# Provider stats we ingest but do not weight — shown at weight 0 for inspection.
# (crosses_completed / dribbles_attempted are now weighted; possession_lost was
# dropped as 79% redundant with the errors_* it overlaps — see PER90_WEIGHTS.)
# big_chance_created IS weighted now (it is the passer's stat, see TOTAL_WEIGHTS) so
# it arrives from there; big_chance_missed stays dropped, already in the SGA.
UNUSED = ["tackles", "possession_lost"]
# Emblematic cases to always show, tagged by disagreement type:
#   "DISAC. 2x"    — we disagree with BOTH fantacalcio and SofaScore. Scorers who
#                    also wasted chances: the goal is our +3 bonus, not the base
#                    vote, and both benchmarks apply a scoring halo we don't.
#   "DISAC. fanta" — we disagree with fantacalcio but SofaScore AGREES with us:
#                    individual merit in a team defeat vs fanta's collective punishment.
FORCED = [
    (1072, 11, "DISAC. 2x"),     # Delprato (DIF) 1 gol, 2 big chance mancate
    (1101, 36, "DISAC. 2x"),     # Malen (ATT) 2 gol ma sprechi
    (768, 11, "DISAC. 2x"),      # Berardi (CEN) 2 gol
    (1188, 10, "DISAC. 2x"),     # Castro (ATT) 2 gol
    (934, 28, "DISAC. 2x"),      # Bowie (ATT) 1 gol
    (740, 29, "DISAC. 2x"),      # Simeone (ATT) 1 gol
    (1002, 25, "DISAC. 2x"),     # Loyola (CEN) 1 gol + rigore concesso
    (926, 37, "DISAC. fanta"),   # Edmundsson (DIF) SofaScore 9.1 conferma noi
    (907, 21, "DISAC. fanta"),   # Koopmeiners (CEN)
    (1160, 33, "DISAC. fanta"),  # Buongiorno (DIF)
]

TEAL = "1F5C53"; YEL = "FFF2CC"
CF_GREEN = ("C6EFCE", "006100"); CF_YEL = ("FFEB9C", "9C6500"); CF_RED = ("FFC7CE", "9C0006")


class Command(BaseCommand):
    help = "Build the interactive voto-puro weight tuner spreadsheet."

    def add_arguments(self, parser):
        parser.add_argument("--season", type=int, default=2)
        parser.add_argument("--cases", type=int, default=28,
                            help="Target number of case columns (emblematic + defender "
                                 "extremes + Statistico discrepancies + fillers).")
        parser.add_argument("--dir", default=None,
                            help="Folder with the external fantacalcio .xlsx sheets.")
        parser.add_argument("--out", default=None,
                            help="Output path (default <repo>/voto_puro_tuner.xlsx).")

    def handle(self, *args, **o):
        cs = o["season"]
        out = o["out"] or str(Path(__file__).resolve().parents[5] / "voto_puro_tuner.xlsx")
        ddir = o["dir"] or str(Path(settings.VFOOT_DATA_DIR) / "data_fantacalcio" / "2025-2026")
        files = sorted(glob.glob(f"{ddir}/*.xlsx"))
        if not files:
            raise CommandError(f"No fantacalcio .xlsx in {ddir}")

        # ---- feature set (exactly the deployed vector) ----
        # Since the standardisation, a case column holds the STANDARDISED value of
        # each feature, so the index is a plain SUMPRODUCT of weights and column:
        # no RAW/SQRT duality to reconstruct in the sheet, and a weight in column C
        # means the same thing as in the code — index points per 1σ.
        TOTAL = list(cr.TOTAL_WEIGHTS)
        PER90 = list(cr.PER90_WEIGHTS)
        FEATS = TOTAL + PER90 + [cr.EXPOSURE_KEY]
        nF = len(FEATS)
        is_p90 = {f: (f in PER90) for f in FEATS}
        curw = {**cr.TOTAL_WEIGHTS, **cr.PER90_WEIGHTS}
        SCALES = cr.feature_scales(gk=False)

        def w_of(f):
            return -cr.EXPOSURE_WEIGHT if f == cr.EXPOSURE_KEY else curw.get(f, 0.0)

        # ---- per-match data ----
        mids = list(Match.objects.filter(competition_season_id=cs).values_list("id", flat=True))
        md_of = {m.id: m.matchday for m in Match.objects.filter(competition_season_id=cs)}
        # _per_match_player_totals already merges the shot-outcome detail (post/goal/
        # save/miss/block), with own goals excluded from shots_goal — so no manual
        # re-injection here (it would re-count own goals as goals scored).
        totals = cr._per_match_player_totals(mids)
        minutes = cr._minutes_map(mids)
        expo = cr.defensive_exposure(mids, minutes)
        gd_on_map = cr.on_pitch_goal_difference(mids, minutes)

        roles = {p: r for p, r in cr.current_role_map().items() if r}
        teamname = {t.id: t.name for t in Team.objects.all()}
        # match result + player side (for margin / heavy-loss selection)
        mgoals = {m.id: (m.home_goals, m.away_goals)
                  for m in Match.objects.filter(competition_season_id=cs)}
        side_of = dict(MatchAppearance.objects.filter(match__competition_season_id=cs)
                       .values_list("id", "side")) if False else {}
        side_of = {(a["match_id"], a["player_id"]): a["side"]
                   for a in MatchAppearance.objects
                   .filter(match__competition_season_id=cs)
                   .values("match_id", "player_id", "side")}

        # Which features are counted EVENTS (integer-valued) rather than continuous
        # quantities. For those, "one sigma" can be a meaningless unit — an error
        # leading to a goal has sigma 0.10, i.e. a tenth of an event, so the column
        # D reading understates it by a factor of ten. The right question there is
        # what ONE occurrence is worth, which is what column E answers.
        is_event = {}
        z_one = {}
        for f in FEATS:
            if f == cr.EXPOSURE_KEY or f in cr.DERIVED_FEATURES or f in cr.MERGED_FEATURES:
                is_event[f] = False
                continue
            vals = [totals[k].get(f, 0.0) for k in totals]
            is_event[f] = all(abs(v - round(v)) < 1e-9 for v in vals)
            if is_event[f]:
                # a count of 1, at 90' the per-90 scale is exactly 1.0
                z_one[f] = cr._feature_z(f, 1.0, SCALES)

        def rawvec(tot, m, ex):
            """the feature values in provider units — for the readable sheets"""
            v = cr.raw_feature_values(tot, m, ex, gk=False)
            return np.array([v.get(f, 0.0) for f in FEATS])

        def zvec(tot, m, ex):
            """the STANDARDISED values the index is actually built from"""
            v = cr.raw_feature_values(tot, m, ex, gk=False)
            return np.array([cr._feature_z(f, v.get(f, 0.0), SCALES) for f in FEATS])

        # ---- population -> per-role mean/cov of the STANDARDISED vector ----
        pop, popraw = defaultdict(list), defaultdict(list)
        gp2mp = {}
        for (mid, pid), tot in totals.items():
            role = roles.get(pid)
            if role not in OUT_ROLES:
                continue
            m = minutes.get((mid, pid), 0)
            gp2mp[(md_of.get(mid), pid)] = (mid, pid)
            if m < cr.MIN_MINUTES_REFERENCE or not cr.is_rated(m, tot):
                continue
            ex = expo.get((mid, pid), 0.0)
            pop[role].append(zvec(tot, m, ex))
            popraw[role].append(rawvec(tot, m, ex))
        STATS = {}
        for role in OUT_ROLES:
            Z = np.array(pop[role]); R = np.array(popraw[role])
            # ddof=0: build_reference computes the POPULATION std, and a sheet
            # that used the sample one would disagree with the deployed vote by a
            # rounding step on cases sitting exactly on a 0.25 boundary.
            STATS[role] = {"mean_z": Z.mean(0), "cov_z": np.cov(Z.T, ddof=0),
                           "mean_raw": R.mean(0), "sd_raw": R.std(0),
                           "n": len(Z)}
        # With POOLED_ROLE_SPREAD the outfield roles share ONE spread: the std of
        # the residuals pooled across them. Projected on a weight vector that is
        # w'Cw with C the sample-weighted average of the WITHIN-role covariances —
        # each role's residuals are already centred on its own mean, so pooling
        # them is exactly this average. Stored as an extra block so the sheet's
        # live sigma keeps matching the deployed one whatever the weights.
        _n = sum(STATS[r]["n"] for r in OUT_ROLES)
        COV_POOLED = sum(STATS[r]["n"] * STATS[r]["cov_z"] for r in OUT_ROLES) / _n

        # ---- discrepancy rows (reuses the external-sheet parsing) ----
        rows = DiscCmd().discrepancy_rows(cs, files)
        name = {p.id: (p.short_name or p.full_name) for p in Player.objects.all()}

        def margin(gd, pid):
            mp = gp2mp.get((gd, pid))
            if not mp:
                return 0
            mid, _ = mp
            hg, ag = mgoals.get(mid, (0, 0))
            return (hg - ag) if side_of.get((mid, pid)) == "home" else (ag - hg)

        cand = []
        for r in rows:
            if (r["minutes"] or 0) < 60 or r["role"] not in OUT_ROLES:
                continue
            # need our vote and at least one external benchmark to compare against
            if (r["gd"], r["pid"]) not in gp2mp or r["our"] is None:
                continue
            if r["fanta"] is None and r["statistico"] is None:
                continue
            r = {**r,
                 "absd": abs(r["our"] - r["fanta"]) if r["fanta"] is not None else 0.0,
                 "absd_stat": (abs(r["our"] - r["statistico"])
                               if r["statistico"] is not None else None),
                 "margin": margin(r["gd"], r["pid"])}
            cand.append(r)

        # ---- case selection (dedup by player) ----
        sel, seen = [], set()

        def take(r, tipo):
            if r["pid"] in seen:
                return False
            sel.append({**r, "tipo": tipo}); seen.add(r["pid"]); return True

        by_key = {(r["gd"], r["pid"]): r for r in cand}
        for pid, gd, tag in FORCED:                              # emblematic cases
            r = by_key.get((gd, pid))
            if r:
                take(r, tag)
        # THE EXTREMES OF THE DEFENDER SCALE. A defender reaching 9 (or dropping to
        # 3) is where the model is least constrained by the reference and most
        # likely to be caught out, so those cases belong in the sheet BY RULE rather
        # than by remembering to add them: they are re-picked automatically on every
        # rebuild, including after a reweighting that creates new ones. Both ends —
        # the exposure term can push a defender down as hard as a goal lifts him up.
        defenders = [x for x in cand if x["role"] == "DIF"]
        for r in sorted(defenders, key=lambda x: -x["our"])[:2]:
            take(r, "ESTREMO alto")
        for r in sorted(defenders, key=lambda x: x["our"])[:2]:
            take(r, "ESTREMO basso")
        # THE STRANGE-VOTE HUNT: biggest gaps vs the Statistico base vote. Statistico
        # is the goal-stripped algorithmic grade, so a large gap here is a pure
        # performance-read disagreement — exactly the votes to sanity-check.
        n = 0
        for r in sorted([x for x in cand if x["absd_stat"] is not None],
                        key=lambda x: -x["absd_stat"]):
            if take(r, "DISAC. stat"):
                n += 1
            if n >= 10:
                break
        # scorers with a base-vote discrepancy
        n = 0
        for r in sorted([x for x in cand if x["goals"]], key=lambda x: -x["absd"]):
            if take(r, "GOL"):
                n += 1
            if n >= 5:
                break
        # heavy defeat, we are high, fanta is low: the result-correlation gap
        n = 0
        for r in sorted([x for x in cand if x["margin"] <= -3 and x["our"] >= 6.5
                         and x["fanta"] <= 5.5], key=lambda x: -x["our"]):
            if take(r, "KO netto"):
                n += 1
            if n >= 4:
                break
        # top discrepancy per role
        for role in OUT_ROLES:
            n = 0
            for r in sorted([x for x in cand if x["role"] == role], key=lambda x: -x["absd"]):
                if take(r, "OUTLIER"):
                    n += 1
                if n >= 2:
                    break
        # a couple of good-agreement anchors per role (sanity)
        for role in OUT_ROLES:
            n = 0
            for r in sorted([x for x in cand if x["role"] == role and x["absd"] <= 0.3
                             and 5.6 <= x["our"] <= 7.6], key=lambda x: x["absd"]):
                if take(r, "buono"):
                    n += 1
                if n >= 1:
                    break
        sel = sel[:o["cases"]]

        casedata = []
        for r in sel:
            gd, pid = r["gd"], r["pid"]
            mid, _ = gp2mp[(gd, pid)]; tot = totals[(mid, pid)]; m = minutes[(mid, pid)]
            role = roles[pid]; ex = expo.get((mid, pid), 0.0)
            M = Match.objects.get(id=mid)
            res = f"gd{gd}: {teamname.get(M.home_team_id,'?')} {M.home_goals}-{M.away_goals} {teamname.get(M.away_team_id,'?')}"
            if r["goals"]:
                res += f" · {r['goals']}⚽"
            casedata.append({"name": name.get(pid, str(pid)), "tipo": r["tipo"], "role": role,
                             "min": m, "match": res, "goals": r["goals"],
                             "raw": rawvec(tot, m, ex), "z": zvec(tot, m, ex),
                             "fanta": r["fanta"], "stat": r["statistico"],
                             "sofa": r["sofa"], "our": r["our"],
                             "expl": r.get("explanation_text", ""),
                             # weight-independent inputs to the post-adjustment layer
                             "gd_on": gd_on_map.get((mid, pid), 0),
                             "red_adj": round(cr.red_card_adjustments(mid).get(pid, 0.0)
                                              + cr.own_goal_adjustments(mid).get(pid, 0.0)
                                              + cr.penalty_missed_adjustments(mid).get(pid, 0.0), 3)})
        casedata.sort(key=lambda c: (OUT_ROLES.index(c["role"]), c["tipo"], c["name"]))

        # ---- the discrepancy ledger: top gaps vs Statistico, with our text ----
        expl_rows, seen_e = [], set()
        for r in sorted([x for x in cand if x["absd_stat"] is not None],
                        key=lambda x: -x["absd_stat"]):
            if r["pid"] in seen_e:
                continue
            seen_e.add(r["pid"])
            mid, _ = gp2mp[(r["gd"], r["pid"])]
            M = Match.objects.get(id=mid)
            expl_rows.append({
                "name": name.get(r["pid"], str(r["pid"])), "role": r["role"], "gd": r["gd"],
                "match": (f"{teamname.get(M.home_team_id,'?')} {M.home_goals}-{M.away_goals} "
                          f"{teamname.get(M.away_team_id,'?')}"),
                "goals": r["goals"], "assists": r["assists"],
                "our": r["our"], "stat": r["statistico"], "fanta": r["fanta"],
                "delta": round(r["our"] - r["statistico"], 1),
                "expl": r.get("explanation_text", "") or "(nessuna voce sopra soglia)"})
            if len(expl_rows) >= 30:
                break

        self._build_xlsx(out, FEATS, nF, is_p90, w_of, STATS, casedata, expl_rows,
                         cov_pooled=COV_POOLED, is_event=is_event, z_one=z_one)
        self.stdout.write(self.style.SUCCESS(
            f"scritto {out} | casi: {len(casedata)} "
            f"(gol: {sum(1 for c in casedata if c['goals'])}, "
            f"KO netto: {sum(1 for c in casedata if c['tipo']=='KO netto')})"))

    # ------------------------------------------------------------------
    def _build_xlsx(self, out, FEATS, nF, is_p90, w_of, STATS, casedata, expl_rows=(),
                    cov_pooled=None, is_event=None, z_one=None):
        is_event = is_event or {}
        z_one = z_one or {}
        wb = openpyxl.Workbook()
        fillh = PatternFill("solid", fgColor=TEAL); yel = PatternFill("solid", fgColor=YEL)

        # ---- ref: means as COLUMNS + covariance blocks ----
        ref = wb.create_sheet("ref")
        meancol = {}
        for ci, role in enumerate(OUT_ROLES):
            c = 2 + ci; meancol[role] = col(c)
            ref.cell(1, c, f"{role}_z")
            for i, v in enumerate(STATS[role]["mean_z"]):
                ref.cell(2 + i, c, float(v))
        covpos = {}; rr = nF + 4
        for role in list(OUT_ROLES) + ["_POOLED"]:
            C = STATS[role]["cov_z"] if role in STATS else cov_pooled
            covpos[role] = rr
            for i in range(nF):
                for j in range(nF):
                    ref.cell(rr, 2 + j, float(C[i, j]))
                rr += 1
            rr += 1

        def meanrng(role):
            return f"ref!${meancol[role]}$2:${meancol[role]}${1+nF}"

        def covrng(role):
            p = covpos[role]
            return f"ref!$B${p}:${col(1+nF)}${p+nF-1}"

        # ---- calc: w_outer + mu/sigma ----
        calc = wb.create_sheet("calc")
        for i in range(nF):
            for j in range(nF):
                calc.cell(2 + i, 2 + j, f"=Tuner!$C${7+i}*Tuner!$C${7+j}")
        wouter = f"calc!$B$2:${col(1+nF)}${1+nF}"; wvec = f"Tuner!$C$7:$C${6+nF}"
        sigma_src = "_POOLED" if cr.POOLED_ROLE_SPREAD else None
        murow = {}; rr = nF + 4
        for role in OUT_ROLES:
            calc.cell(rr, 1, role)
            calc.cell(rr, 2, f"=SUMPRODUCT({wvec},{meanrng(role)})")
            # sigma dalla covarianza CONDIVISA (POOLED_ROLE_SPREAD): il centro
            # resta per ruolo, la dispersione e' unica fra i ruoli di movimento.
            calc.cell(rr, 3,
                      f"=SQRT(SUMPRODUCT({covrng(sigma_src or role)},{wouter}))")
            murow[role] = rr; rr += 1

        # ---- cases: feature vectors + highlight + per-role mean/var/σ ----
        cs = wb.create_sheet("cases")

        def zfill(v, mean, var):
            if var <= 0:
                return None
            z = (v - mean) / (var ** 0.5)
            if z >= 2: return PatternFill("solid", fgColor="ED7D31")
            if z >= 1: return PatternFill("solid", fgColor="FFD966")
            if z <= -2: return PatternFill("solid", fgColor="2E75B6")
            if z <= -1: return PatternFill("solid", fgColor="9DC3E6")
            return None
        for ci, c in enumerate(casedata):
            S = STATS[c["role"]]
            for i in range(nF):
                rc = cs.cell(2 + i, 2 + ci, float(c["z"][i])); rc.number_format = "0.00"
                f = zfill(c["z"][i], S["mean_z"][i], S["cov_z"][i, i])
                if f: rc.fill = f
                qc = cs.cell(2 + nF + 2 + i, 2 + ci, float(c["raw"][i]))
                qc.number_format = "0.00"
                f2 = zfill(c["raw"][i], S["mean_raw"][i], S["sd_raw"][i] ** 2)
                if f2: qc.fill = f2

        def craw(ci): return f"cases!${col(2+ci)}$2:${col(2+ci)}${1+nF}"
        cs["A1"] = ("VALORI STANDARDIZZATI (entrano nell'indice: indice = "
                    "SOMMAPRODOTTO(pesi, questa colonna))")
        cs.cell(nF + 3, 1, "valori GREZZI (per-90 dove serve) — solo per leggerli")
        for i, f in enumerate(FEATS):
            cs.cell(2 + i, 1, f); cs.cell(nF + 4 + i, 1, f)
        for ci, c in enumerate(casedata):
            cs.cell(1, 2 + ci, c["name"]).font = Font(bold=True)
            cs.cell(nF + 3, 2 + ci, c["name"]).font = Font(bold=True)
        base = 2 + len(casedata) + 1
        for j, role in enumerate(OUT_ROLES):
            mc = base + 3 * j; vc = mc + 1; sc = mc + 2
            for hr in (1, nF + 3):
                cs.cell(hr, mc, f"{role} media").font = Font(bold=True, color=TEAL)
                cs.cell(hr, vc, f"{role} var").font = Font(bold=True, color=TEAL)
                cs.cell(hr, sc, f"{role} σ").font = Font(bold=True, color=TEAL)
            for i in range(nF):
                vz = STATS[role]["cov_z"][i, i]; sdr = STATS[role]["sd_raw"][i]
                cs.cell(2 + i, mc, round(float(STATS[role]["mean_z"][i]), 3))
                cs.cell(2 + i, vc, round(float(vz), 4))
                cs.cell(2 + i, sc, round(float(vz) ** 0.5, 3))
                cs.cell(nF + 4 + i, mc, round(float(STATS[role]["mean_raw"][i]), 3))
                cs.cell(nF + 4 + i, vc, round(float(sdr) ** 2, 4))
                cs.cell(nF + 4 + i, sc, round(float(sdr), 3))
        # TERZO BLOCCO: il contributo vero, valore x peso. Vive di formule, quindi
        # segue i pesi mentre li editi. Serve perche' il blocco standardizzato in
        # cima invita a confrontare i NUMERI fra righe, mentre quello che arriva al
        # voto e' il prodotto: un 1.36 su un peso 0.017 vale meno di uno 0.3 su 0.16.
        r0 = 2 * nF + 5
        cs.cell(r0 - 1, 1, "CONTRIBUTO all'indice (valore × peso, live)").font = Font(
            bold=True, color=TEAL)
        for i, f in enumerate(FEATS):
            cs.cell(r0 + i, 1, f)
        for ci, c in enumerate(casedata):
            cs.cell(r0 - 1, 2 + ci, c["name"]).font = Font(bold=True)
            for i in range(nF):
                cc = cs.cell(r0 + i, 2 + ci,
                             f"={col(2+ci)}{2+i}*Tuner!$C${7+i}")
                cc.number_format = "0.000"
            cs.cell(r0 + nF, 2 + ci,
                    f"=SUM({col(2+ci)}{r0}:{col(2+ci)}{r0+nF-1})").font = Font(bold=True)
        cs.cell(r0 + nF, 1, "TOTALE = indice").font = Font(bold=True)

        cs.cell(2 * nF + 6, 1, "Evidenziazione (vs media del RUOLO del caso): ambra=+1σ "
                "arancio=+2σ (alto); azzurro=-1σ blu=-2σ (basso). Colonne a destra: "
                "media/var/σ per ruolo.").font = Font(italic=True, size=9)
        cs.column_dimensions["A"].width = 22
        for ci in range(len(casedata)):
            cs.column_dimensions[col(2 + ci)].width = 11
        for j in range(9):
            cs.column_dimensions[col(base + j)].width = 10

        # ---- Tuner ----
        tun = wb.create_sheet("Tuner"); wb.move_sheet("Tuner", -(len(wb.sheetnames) - 1))
        tun["A1"] = "VOTO PURO — tuner dei pesi"; tun["A1"].font = Font(bold=True, size=14)
        tun["A2"] = ("Edita i PESI (col C, celle gialle) e K mitigazione (B5). Interruttore "
                     "RAW/SQRT in B4. Il VOTO FINALE (riga 24) si colora: verde=accordo, "
                     "giallo=borderline, rosso=outlier.")
        tun["A3"] = ("Pipeline: voto base = 6+0.8·(min/(min+25))·(indice−media)/σ in [3,10]; "
                     "poi + mitigazione risultato (solo divergenze, cap ±1) + red/autogol; "
                     "poi clamp [3,10] e arrotondamento 0.5. gd_on e red/autogol sono FISSI "
                     "(non dipendono dai pesi).")
        tun["A4"] = ("OGNI PESO = contributo all'indice di 1σ di quella feature. Le colonne "
                     f"dei casi contengono i valori GIA' STANDARDIZZATI (standardizza → "
                     f"comprimi con {cr.COMPRESS_K:g}·log(1+u/{cr.COMPRESS_K:g}) → "
                     "ristandardizza), quindi indice = SOMMAPRODOTTO(pesi, colonna). "
                     "Per far contare A la meta' di B, dai ad A meta' peso: adesso e' vero.")
        tun["A4"].font = Font(italic=True, size=9)
        tun["A5"] = "K mitigazione:"; tun["B5"] = cr.RESULT_MITIGATION_K
        tun["C5"] = "base sc/vitt:"; tun["D5"] = cr.RESULT_MITIGATION_BASE
        for cell in ("B5", "D5"):
            tun[cell].font = Font(bold=True); tun[cell].fill = yel
            tun[cell].number_format = "0.00"
        murow_dif = murow["DIF"]
        tun["A6"] = "feature"; tun["B6"] = "tipo"; tun["C6"] = "PESO"
        # Column D answers the question the sheet used to hide: a case column holds
        # STANDARDISED values, so a 1.36 next to a 0.00 looks big — but what reaches
        # the vote is value x weight, and the weights span a factor of 20. Without
        # this column an analyst reads the value and infers an importance that is
        # not there (it happened: an xA of +0.9 sigma read as decisive when it was
        # worth 0.02 of a vote).
        tun["D6"] = "1σ in VOTI"
        tun["E6"] = "1 EVENTO in VOTI"
        for cc in ("A6", "B6", "C6", "D6", "E6"):
            tun[cc].font = Font(bold=True, color="FFFFFF"); tun[cc].fill = fillh
        for i, f in enumerate(FEATS):
            tun.cell(7 + i, 1, f)
            tun.cell(7 + i, 2, "PER90" if is_p90[f] else ("EXPOS" if f == "_exposure" else "TOT"))
            # 4 decimals, not 3: the weights are now "index points per 1σ" and live
            # around 0.01-0.15, where a third decimal is a 1% truncation and shows up
            # as a rounding step on cases sitting near a 0.25 vote boundary.
            wc = tun.cell(7 + i, 3, round(w_of(f), 4)); wc.fill = yel
            wc.number_format = "0.0000"
            # |peso| / sigma dell'indice x K x shrinkage a 90' = punti di voto che
            # vale UNA sigma di quella feature. Live: segue i pesi che editi.
            dc = tun.cell(7 + i, 4,
                          f"=ABS(C{7+i})/calc!$C${murow_dif}*{cr.VOTE_SPREAD_K}"
                          f"*90/(90+{cr.SHRINKAGE_MINUTES})")
            dc.number_format = "0.00"
            # SIGNED, unlike D: an error that costs half a vote should read -0.50,
            # not 0.50. Blank for continuous features, where "one occurrence" of an
            # xA or of conceded danger means nothing.
            if is_event.get(f):
                ec = tun.cell(7 + i, 5,
                              f"={z_one[f]:.6f}*C{7+i}/calc!$C${murow_dif}"
                              f"*{cr.VOTE_SPREAD_K}*90/(90+{cr.SHRINKAGE_MINUTES})")
                ec.number_format = "+0.00;-0.00"
            else:
                tun.cell(7 + i, 5, "—")
        # c0 = 6: le etichette di riga dei casi stanno in F, non in E, perche' la
        # colonna E ora appartiene alla tabella dei pesi ("1 EVENTO in VOTI").
        KM = "Tuner!$B$5"; BB = "Tuner!$D$5"; c0 = 6
        rowlab = [(7, "giocatore"), (8, "TIPO"), (9, "ruolo"), (10, "partita (gd, risultato, gol)"),
                  (11, "minuti"), (12, "fanta"), (13, "statistico"), (14, "sofascore"),
                  (15, "nostro(attuale)"), (16, "indice"), (17, "media INDICE ruolo"),
                  (18, "sigma INDICE ruolo"), (19, "gd_on (in campo)"), (20, "red/autogol/rigore (fisso)"),
                  (21, "voto base"), (22, "mitigazione"), (24, "VOTO FINALE (live)")]
        for rr2, lab in rowlab:
            tun.cell(rr2, c0, lab).font = Font(bold=True, size=9)
        for ci, c in enumerate(casedata):
            cc = c0 + 1 + ci; L = col(cc); mr = murow[c["role"]]
            tun.cell(7, cc, c["name"]).font = Font(bold=True)
            _tipo_fill = {"buono": "C9E7DF", "DISAC. 2x": "F8CBAD",
                          "DISAC. fanta": "FFE699"}.get(c["tipo"], "F7ECDD")
            tun.cell(8, cc, c["tipo"]).fill = PatternFill("solid", fgColor=_tipo_fill)
            tun.cell(9, cc, c["role"]); tun.cell(10, cc, c["match"]); tun.cell(11, cc, c["min"])
            tun.cell(12, cc, c["fanta"]); tun.cell(13, cc, c["stat"] if c["stat"] is not None else "-")
            tun.cell(14, cc, c["sofa"] if c["sofa"] is not None else "-"); tun.cell(15, cc, c["our"])
            tun.cell(16, cc, f'=SUMPRODUCT({wvec},{craw(ci)})')
            tun.cell(17, cc, f'=calc!$B${mr}')
            tun.cell(18, cc, f'=calc!$C${mr}')
            tun.cell(19, cc, c["gd_on"])
            tun.cell(20, cc, c["red_adj"])
            # voto base (clamp [3,10], pre-arrotondamento)
            tun.cell(21, cc, f'=MAX(3,MIN(10,6+0.8*({L}11/({L}11+25))*(({L}16-{L}17)/{L}18)))')
            # mitigazione: solo divergenze, gravità = base + K·|gd_on|, cap ±1.
            #   sconfitta (gd_on<0): voto alto scende di (voto-6)·(base+K·|gd_on|)
            #   vittoria  (gd_on>0): voto basso sale di (6-voto)·(base+K·gd_on)
            tun.cell(22, cc,
                     f'=MAX(-1,MIN(1,IF({L}19<0,-MAX(0,{L}21-6)*({BB}+{KM}*(-{L}19)),'
                     f'IF({L}19>0,MAX(0,6-{L}21)*({BB}+{KM}*{L}19),0))))')
            # voto finale = clamp(base + mitigazione + red_adj), arrotondato a 0.5
            tun.cell(24, cc, f'=ROUND(MAX(3,MIN(10,{L}21+{L}22+{L}20))*2)/2')
            tun.cell(24, cc).font = Font(bold=True, size=12)
            for rr in (16, 17, 18, 20, 21, 22):
                tun.cell(rr, cc).number_format = "0.000"
            tun.cell(24, cc).number_format = "0.0"
            for rr in (12, 13, 14, 15):
                if isinstance(tun.cell(rr, cc).value, (int, float)):
                    tun.cell(rr, cc).number_format = "0.0"
            tun.column_dimensions[col(cc)].width = 13
        # conditional colour on the FINAL voto row vs fanta/statistico
        first = col(c0 + 1); last = col(c0 + len(casedata)); rng = f"{first}24:{last}24"
        D = f"MIN(ABS({first}24-{first}12),IF(ISNUMBER({first}13),ABS({first}24-{first}13),99))"
        for cond, (bg, fg) in ((f"{D}<=0.75", CF_GREEN), (f"AND({D}>0.75,{D}<=1.5)", CF_YEL), (f"{D}>1.5", CF_RED)):
            tun.conditional_formatting.add(rng, FormulaRule(
                formula=[cond], fill=PatternFill("solid", fgColor=bg),
                font=Font(bold=True, size=12, color=fg)))
        tun["F26"] = ("NOTE: 'media/sigma INDICE ruolo' sono media e dev.std dell'INDICE (somma pesata) "
                      "tra i giocatori del ruolo → cambiano coi pesi. gd_on e red_adj sono FISSI. "
                      "'nostro(attuale)' (riga 15) è il deployato e dovrebbe ≈ VOTO FINALE (riga 24).")
        tun["F27"] = ("TIPO: 'DISAC. 2x'=disaccordo con fanta E SofaScore (marcatori "
                      "che sprecano: il gol è bonus +3, non voto base; loro fanno l'alone-gol, "
                      "noi no); 'DISAC. fanta'=disaccordo con fanta ma SofaScore ci dà ragione "
                      "(merito individuale in sconfitta vs punizione collettiva); "
                      "GOL=marcatore, 'KO netto'=sconfitta ≥3 gol, OUTLIER, buono=accordo; "
                      "'ESTREMO alto/basso'=i voti difensori più alti e più bassi della "
                      "stagione, ripescati a ogni rebuild (lì il modello è meno vincolato).")
        tun["F28"] = ("Mitigazione: solo divergenze (voto>6 in sconfitta → giù; voto<6 in vittoria → su), "
                      "gravità = base + K·|gd_on|, cap ±1. K in B5, 'base sc/vitt' (contributo "
                      "discreto sconfitta/vittoria, oltre i gol) in D5. SGA_Pali: xgOT−xg + palo.")
        tun["F29"] = (
            f"EXPOSURE (_exposure, peso {-cr.EXPOSURE_WEIGHT:.2f}): pericolo SUBITO addebitato "
            f"a chi era in quella zona. Per ogni tiro avversario (rigori esclusi, solo minuti "
            f"in campo) l'addebito è λ·esito + (1−λ)·xGOT con λ={cr.EXPOSURE_LAMBDA:.2f}; "
            f"esito = 1 gol, {cr.EXPOSURE_POST_OUTCOME:.3f} legno (= shots_post/shots_goal dei "
            f"nostri pesi d'attacco), 0 altrimenti — quindi tiri fuori/ribattuti costano zero. "
            f"L'addebito è diviso fra i giocatori di MOVIMENTO in campo in quel minuto in "
            f"proporzione alla presenza (heatmap) nella zona specchiata + "
            f"{cr.EXPOSURE_KERNEL:.2f}× le adiacenti; le quote fanno 100%, il portiere è "
            f"escluso dalla divisione (il suo canale risponde già delle parate). Il valore in "
            f"colonna è FISSO (non dipende dai pesi), il PESO no: editalo in C.")
        for a in ("F26", "F27", "F28", "F29"):
            tun[a].font = Font(italic=True, size=9)
        tun.column_dimensions["A"].width = 21; tun.column_dimensions["C"].width = 8
        tun.column_dimensions["D"].width = 11; tun.column_dimensions["E"].width = 15
        tun.column_dimensions["F"].width = 19

        # ---- medie (readable per-feature role means) ----
        med = wb.create_sheet("medie")
        med["A1"] = "Medie per-FEATURE per ruolo (FISSE)."; med["A1"].font = Font(bold=True)
        for j, h in enumerate(["feature", "DIF grezza", "CEN grezza", "ATT grezza",
                               "DIF σ grezza", "CEN σ grezza", "ATT σ grezza"]):
            hc = med.cell(3, 1 + j, h); hc.font = Font(bold=True, color="FFFFFF"); hc.fill = fillh
        for i, f in enumerate(FEATS):
            med.cell(4 + i, 1, f)
            for jr, role in enumerate(OUT_ROLES):
                med.cell(4 + i, 2 + jr, round(float(STATS[role]["mean_raw"][i]), 3))
                med.cell(4 + i, 5 + jr, round(float(STATS[role]["sd_raw"][i]), 3))
        med.column_dimensions["A"].width = 21

        # ---- spiegazioni: the strange-vote ledger with our generated text ----
        if expl_rows:
            sp = wb.create_sheet("spiegazioni")
            wb.move_sheet("spiegazioni", -(len(wb.sheetnames) - 2))  # right after Tuner
            sp["A1"] = ("Voti più lontani dallo STATISTICO (voto fanta senza bonus/malus), "
                        "con la spiegazione testuale che genera il nostro sistema.")
            sp["A1"].font = Font(bold=True, size=12)
            sp["A2"] = ("Lo Statistico toglie i gol dal voto, quindi un Δ grande = pura "
                        "divergenza sulla LETTURA della prestazione. 'nostro' è il voto puro "
                        "deployato (mitigazione + drop-prestazione da espulsione/autogol già "
                        "dentro). ATTENZIONE alle righe con 'Espulso.'/'Autogol.' nella "
                        "spiegazione: lì il Δ è GONFIATO, perché lo Statistico mette tutta la "
                        "penalità nel voto base mentre noi ci aggiungiamo il malus -1/-2 a "
                        "livello di FANTAVOTO (non qui). Leggi la spiegazione e giudica.")
            sp["A2"].font = Font(italic=True, size=9)
            heads = ["giocatore", "ruolo", "gd", "partita", "gol", "assist",
                     "nostro", "statistico", "fanta", "Δ ns−stat", "spiegazione"]
            for j, h in enumerate(heads):
                hc = sp.cell(4, 1 + j, h)
                hc.font = Font(bold=True, color="FFFFFF"); hc.fill = fillh
            for i, e in enumerate(expl_rows):
                r = 5 + i
                sp.cell(r, 1, e["name"]).font = Font(bold=True)
                sp.cell(r, 2, e["role"]); sp.cell(r, 3, e["gd"]); sp.cell(r, 4, e["match"])
                sp.cell(r, 5, e["goals"] or ""); sp.cell(r, 6, e["assists"] or "")
                sp.cell(r, 7, e["our"]).number_format = "0.0"
                sp.cell(r, 8, e["stat"]).number_format = "0.0"
                sp.cell(r, 9, e["fanta"] if e["fanta"] is not None else "-")
                dc = sp.cell(r, 10, e["delta"]); dc.number_format = "+0.0;-0.0"
                dc.font = Font(bold=True)
                ad = abs(e["delta"])
                fg = (CF_RED if ad >= 2 else (CF_YEL if ad >= 1.5 else CF_GREEN))
                dc.fill = PatternFill("solid", fgColor=fg[0]); dc.font = Font(bold=True, color=fg[1])
                tc = sp.cell(r, 11, e["expl"]); tc.alignment = Alignment(wrap_text=True, vertical="top")
            widths = {"A": 18, "B": 6, "C": 5, "D": 26, "E": 5, "F": 6,
                      "G": 8, "H": 9, "I": 7, "J": 10, "K": 88}
            for c, w in widths.items():
                sp.column_dimensions[c].width = w
            sp.freeze_panes = "A5"

        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        wb.save(out)
