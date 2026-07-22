"""Fit the FREE feature weights of the classic voto-puro heuristic to an external
target vote (SofaScore ``rating`` or a fantacalcio.it sheet), as a DIAGNOSTIC.

Our voto puro is, by construction, LINEAR in its feature weights:

    index  = Σ_k  w_k · compress(feature_k)          # TOTAL feats as totals,
                                                      # PER90 feats per-90+floored
    vote   = 6 + spread_k · w · (index − mean_role) / std_role

so "find the weights that best reproduce the target vote" is a plain RIDGE
regression of the SAME compressed feature basis (plus per-role intercepts) onto
the target. The cross-validated correlation of that fit is the CEILING our data
can reach for that target; the fitted coefficients are the provider's IMPLIED
feature weighting. Comparing that ceiling to our hand-tuned model's correlation
(printed as the baseline) answers two questions:

  1. is the residual gap because we MIS-WEIGHT features we already have, or because
     the signal simply isn't in our data (ceiling ≈ our model → near the limit)?
  2. which features does the target actually reward, and do our hand weights agree
     in sign / ranking?

This does NOT modify the heuristic. It only measures.

    python manage.py classic_fit_weights --target rating
    python manage.py classic_fit_weights --target statistico
    python manage.py classic_fit_weights --target fantacalcio --per-role
"""

from __future__ import annotations

import glob
import json
import math
import re
from collections import Counter, defaultdict

import numpy as np
import openpyxl
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError

from realdata.models import Match, Player, PlayerTeamStint
from realdata.services.identity import norm_name
from vfoot.services.classic_rating import (
    EXTRAP_FLOOR_MINUTES, MIN_MINUTES_REFERENCE, PER90_WEIGHTS, TOTAL_WEIGHTS,
    WEIGHTS, _compress, _minutes_map, _per_match_player_totals, build_reference,
    is_rated, voto_puro_for_match,
)

FEATURES = list(TOTAL_WEIGHTS) + list(PER90_WEIGHTS)  # stable column order
ROLES = [Player.ROLE_DEF, Player.ROLE_MID, Player.ROLE_FWD]  # outfield only

DEFAULT_DIR = str(Path(settings.VFOOT_DATA_DIR) / "data_fantacalcio" / "2025-2026")
DEFAULT_CACHE = str(Path(settings.VFOOT_DATA_DIR) / "historical-data" / "serie-a" / "sofascore" / "cache")
SHEET_FOR_TARGET = {"statistico": "Statistico", "fantacalcio": "Fantacalcio",
                    "italia": "Italia"}
EXTERNAL_TEAMS = {
    "atalanta", "bologna", "cagliari", "como", "cremonese", "fiorentina", "genoa",
    "verona", "hellas verona", "inter", "juventus", "lazio", "lecce", "milan",
    "napoli", "parma", "pisa", "roma", "sassuolo", "torino", "udinese",
}


def _feature_row(totals: dict, minutes: int) -> list:
    """The per-feature compressed basis our index sums (BEFORE weighting)."""
    scale = 90.0 / max(minutes, EXTRAP_FLOOR_MINUTES)
    row = []
    for k in TOTAL_WEIGHTS:
        row.append(_compress(totals.get(k, 0.0)))
    for k in PER90_WEIGHTS:
        row.append(_compress(totals.get(k, 0.0) * scale))
    return row


def _pearson(a, b):
    a, b = np.asarray(a, float), np.asarray(b, float)
    if a.size < 2 or a.std() == 0 or b.std() == 0:
        return float("nan")
    return float(np.corrcoef(a, b)[0, 1])


def _club_key(name):
    fillers = {"ac", "as", "fc", "ssc", "us", "ss", "hellas", "calcio"}
    return " ".join(t for t in norm_name(name).split() if t not in fillers)


def _ridge_cv(X, y, role_oh, alphas, folds, rng):
    """K-fold ridge; intercept columns (role one-hots) are NOT penalised.
    Returns (best_alpha, pooled out-of-fold predictions, refit coefs on full data)."""
    n, p = X.shape
    Xa = np.hstack([X, role_oh])            # features + role intercepts
    pen = np.concatenate([np.ones(p), np.zeros(role_oh.shape[1])])
    idx = rng.permutation(n)
    fold_id = np.array_split(idx, folds)

    def fit(tr, alpha):
        Xt, yt = Xa[tr], y[tr]
        A = Xt.T @ Xt + alpha * np.diag(pen)
        return np.linalg.solve(A, Xt.T @ yt)

    best, best_corr, best_pred = alphas[0], -2.0, None
    for alpha in alphas:
        pred = np.empty(n)
        for f in range(folds):
            te = fold_id[f]
            tr = np.concatenate([fold_id[g] for g in range(folds) if g != f])
            pred[te] = Xa[te] @ fit(tr, alpha)
        c = _pearson(pred, y)
        if c > best_corr:
            best, best_corr, best_pred = alpha, c, pred
    coefs = fit(np.arange(n), best)
    return best, best_pred, coefs


class Command(BaseCommand):
    help = "Fit voto-puro feature weights to an external target vote (diagnostic)."

    def add_arguments(self, parser):
        parser.add_argument("--competition-season", type=int, default=2)
        parser.add_argument("--target", default="rating",
                            help="rating | statistico | fantacalcio | italia")
        parser.add_argument("--dir", default=DEFAULT_DIR)
        parser.add_argument("--cache-dir", default=DEFAULT_CACHE)
        parser.add_argument("--per-role", action="store_true",
                            help="fit a separate regression per role and pool")
        parser.add_argument("--with-bonus", action="store_true",
                            help="add goals+assists (the excluded bonus events) as "
                                 "extra fit features, to measure their effect")
        parser.add_argument("--folds", type=int, default=5)
        parser.add_argument("--min-minutes", type=int, default=MIN_MINUTES_REFERENCE)
        parser.add_argument("--seed", type=int, default=0)

    # -- target loaders --------------------------------------------------

    def _ratings(self, ext_id, cache_dir):
        try:
            d = json.load(open(f"{cache_dir}/api_v1_event_{ext_id}_lineups.json"))
        except (FileNotFoundError, ValueError):
            return {}
        out = {}
        for side in ("home", "away"):
            for pl in d.get(side, {}).get("players", []):
                pid = (pl.get("player") or {}).get("id")
                st = pl.get("statistics") or {}
                if pid is not None and st.get("rating"):
                    out[str(pid)] = float(st["rating"])
        return out

    def _bonus_map(self, cs_id, cache_dir):
        """(matchday, our_player_id) -> (goals, assists) from cache lineups stats.

        These are the events we DELIBERATELY exclude from voto puro (they're the
        +3/+1 bonus layer). Added as extra fit features only to measure how much of
        the residual gap to a goal-inflated base vote is explained by the goal/assist
        signal we hold back."""
        sofa_ext = dict(Player.objects.filter(external_source="sofascore")
                        .values_list("id", "external_id"))
        ext_to_pid = {v: k for k, v in sofa_ext.items()}
        out = {}
        for m in Match.objects.filter(competition_season_id=cs_id):
            if not m.external_id:
                continue
            try:
                d = json.load(open(f"{cache_dir}/api_v1_event_"
                                   f"{m.external_id}_lineups.json"))
            except (FileNotFoundError, ValueError):
                continue
            for side in ("home", "away"):
                for pl in d.get(side, {}).get("players", []):
                    pid = ext_to_pid.get(str((pl.get("player") or {}).get("id")))
                    if pid is None:
                        continue
                    st = pl.get("statistics") or {}
                    out[(m.matchday, pid)] = (float(st.get("goals", 0) or 0),
                                              float(st.get("goalAssist", 0) or 0))
        return out

    def _rating_target(self, cs_id, cache_dir):
        """(matchday, our_player_id) -> SofaScore rating."""
        sofa_ext = dict(Player.objects.filter(external_source="sofascore")
                        .values_list("id", "external_id"))
        ext_to_pid = {v: k for k, v in sofa_ext.items()}
        out = {}
        for m in Match.objects.filter(competition_season_id=cs_id):
            if not m.external_id:
                continue
            for extid, val in self._ratings(m.external_id, cache_dir).items():
                pid = ext_to_pid.get(extid)
                if pid is not None:
                    out[(m.matchday, pid)] = val
        return out

    def _external_target(self, cs_id, dir_, sheet):
        """(matchday, our_player_id) -> external base vote, matched team+surname."""
        teams = (PlayerTeamStint.objects
                 .filter(team_season__competition_season_id=cs_id)
                 .values_list("team_season__team__name", flat=True).distinct())
        team_map = {_club_key(t): t for t in set(teams)}
        pidx = defaultdict(list)
        rows = (PlayerTeamStint.objects
                .filter(team_season__competition_season_id=cs_id, end_date__isnull=True)
                .values_list("player_id", "team_season__team__name",
                             "player__full_name", "player__short_name"))
        for pid, team, fnm, snm in rows:
            surnames = set()  # dedupe: full & short often share a surname token
            for nm in (fnm, snm):
                toks = norm_name(nm or "").split()
                if toks:
                    surnames.add(toks[-1])
            for s in surnames:
                pidx[(team, s)].append(pid)

        gd_re = re.compile(r"Giornata_(\d+)")
        out, unm_team, unm_name = {}, Counter(), 0
        for f in sorted(glob.glob(f"{dir_}/*.xlsx")):
            mm = gd_re.search(f)
            if not mm:
                continue
            gd = int(mm.group(1))
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            if sheet not in wb.sheetnames:
                continue
            team = None
            for r in wb[sheet].iter_rows(values_only=True):
                c0 = r[0]
                if isinstance(c0, str) and norm_name(c0) in EXTERNAL_TEAMS:
                    team = c0
                    continue
                if not isinstance(c0, (int, float)) or c0 == "Cod.":
                    continue
                voto = _parse_voto(r[3])
                if voto is None:
                    continue
                our_team = team_map.get(_club_key(team or ""))
                if not our_team:
                    unm_team[team] += 1
                    continue
                surn = norm_name(r[2]).split()[-1] if r[2] else ""
                cands = pidx.get((our_team, surn), [])
                if len(cands) != 1:
                    unm_name += 1
                    continue
                out[(gd, cands[0])] = voto
        self._unm = (dict(unm_team), unm_name)
        return out

    # -- main ------------------------------------------------------------

    def handle(self, *args, **opts):
        cs_id = opts["competition_season"]
        target = opts["target"].lower()
        w = self.stdout.write

        if target == "rating":
            tgt = self._rating_target(cs_id, opts["cache_dir"])
            tname = "SofaScore rating"
        elif target in SHEET_FOR_TARGET:
            tgt = self._external_target(cs_id, opts["dir"], SHEET_FOR_TARGET[target])
            tname = f"fantacalcio.it '{SHEET_FOR_TARGET[target]}'"
        else:
            raise CommandError(f"unknown --target {target!r}")
        if not tgt:
            raise CommandError("no target votes loaded")

        # Build the sample: rated outfield player-matchdays that HAVE a target vote.
        match_ids = list(Match.objects.filter(competition_season_id=cs_id)
                         .values_list("id", flat=True))
        md_by_match = dict(Match.objects.filter(competition_season_id=cs_id)
                           .values_list("id", "matchday"))
        totals = _per_match_player_totals(match_ids)
        minutes = _minutes_map(match_ids)
        roles = dict(Player.objects.exclude(classic_role="")
                     .values_list("id", "classic_role"))

        feat_names = list(FEATURES)
        bonus = {}
        if opts["with_bonus"]:
            bonus = self._bonus_map(cs_id, opts["cache_dir"])
            feat_names += ["goals", "assists"]
        nfeat = len(feat_names)

        # our current model's prediction on the same sample = the baseline
        ref = build_reference(cs_id)
        our_pred = {}
        for m in Match.objects.filter(competition_season_id=cs_id):
            for row in voto_puro_for_match(m, ref):
                if row["rated"]:
                    our_pred[(m.matchday, row["player_id"])] = row["voto_puro"]

        Xrows, yrows, rrows, ourrows = [], [], [], []
        for (mid, pid), feats in totals.items():
            role = roles.get(pid)
            if role not in ROLES:
                continue
            mins = minutes.get((mid, pid), 0)
            if mins < opts["min_minutes"] or not is_rated(mins, feats):
                continue
            key = (md_by_match[mid], pid)
            if key not in tgt or key not in our_pred:
                continue
            row = _feature_row(feats, mins)
            if opts["with_bonus"]:
                g, a = bonus.get(key, (0.0, 0.0))
                row = row + [g, a]
            Xrows.append(row)
            yrows.append(tgt[key])
            rrows.append(role)
            ourrows.append(our_pred[key])

        n = len(Xrows)
        if n < 200:
            raise CommandError(f"only {n} matched samples — too few")
        X = np.asarray(Xrows, float)
        y = np.asarray(yrows, float)
        our = np.asarray(ourrows, float)
        role_arr = np.asarray(rrows)

        # standardize feature columns (coefs become 'vote pts per std of feature')
        mu, sd = X.mean(0), X.std(0)
        sd[sd == 0] = 1.0
        Xs = (X - mu) / sd
        role_oh = np.column_stack([(role_arr == r).astype(float) for r in ROLES])
        rng = np.random.default_rng(opts["seed"])
        alphas = [0.3, 1.0, 3.0, 10.0, 30.0, 100.0]

        w(f"=== fit voto-puro weights → {tname} (cs={cs_id}) ===")
        w(f"samples: {n}  (per role: " +
          ", ".join(f"{r}={int((role_arr==r).sum())}" for r in ROLES) + ")")
        if target != "rating":
            ut, un = getattr(self, "_unm", ({}, 0))
            w(f"unmatched names: {un}; unmatched teams: {ut}")

        base_corr = _pearson(our, y)
        w(f"\nBASELINE  corr(our current model, {target}) = {base_corr:.3f}")

        if opts["per_role"]:
            preds = np.empty(n)
            for r in ROLES:
                m = role_arr == r
                a, pr, _ = _ridge_cv(Xs[m], y[m],
                                     np.ones((int(m.sum()), 1)),  # single intercept
                                     alphas, opts["folds"], rng)
                preds[m] = pr
                w(f"  [{r}] n={int(m.sum())} alpha={a:g} "
                  f"CV corr={_pearson(pr, y[m]):.3f}")
            w(f"\nFITTED (per-role)  pooled CV corr = {_pearson(preds, y):.3f}")
            # refit per-role coefs on full data for the weight table
            coef_by_role = {}
            for r in ROLES:
                m = role_arr == r
                _, _, c = _ridge_cv(Xs[m], y[m], np.ones((int(m.sum()), 1)),
                                    alphas, opts["folds"], rng)
                coef_by_role[r] = c[:nfeat]
            self._weight_table(w, feat_names,
                               np.mean(list(coef_by_role.values()), axis=0))
        else:
            alpha, pred, coefs = _ridge_cv(Xs, y, role_oh, alphas,
                                           opts["folds"], rng)
            cv_corr = _pearson(pred, y)
            fitted = Xs @ coefs[:nfeat] + role_oh @ coefs[nfeat:]
            ins = _pearson(fitted, y)
            w(f"FITTED (global, alpha={alpha:g})  CV corr = {cv_corr:.3f}"
              f"   in-sample = {ins:.3f}")
            gain = cv_corr - base_corr
            w(f"\nGAIN vs our model: {gain:+.3f}  "
              f"({'signal is there but mis-weighted' if gain > 0.04 else 'near the ceiling of our data — weights already good'})")
            self._weight_table(w, feat_names, coefs[:nfeat],
                               role_int=dict(zip(ROLES, coefs[nfeat:])))

    def _weight_table(self, w, feat_names, coefs, role_int=None):
        if role_int:
            w("\nrole intercepts: " +
              ", ".join(f"{r}={v:.2f}" for r, v in role_int.items()))
        w("\nfitted standardized coef (vote pts / std)  vs  our hand weight:")
        w(f"  {'feature':<28} {'fitted':>8} {'ours':>7}  {'bucket':<6} note")
        order = sorted(range(len(feat_names)), key=lambda i: -abs(coefs[i]))
        for i in order:
            k = feat_names[i]
            ours = WEIGHTS.get(k)  # None for bonus extras (goals/assists)
            bucket = ("TOTAL" if k in TOTAL_WEIGHTS
                      else "BONUS" if ours is None else "per90")
            ours_s = f"{ours:>7.2f}" if ours is not None else f"{'—':>7}"
            note = ""
            if ours is not None and (coefs[i] > 0) != (ours > 0) and abs(coefs[i]) > 0.02:
                note = "SIGN DISAGREES"
            elif abs(coefs[i]) < 0.02:
                note = "~0 (target ignores it)"
            w(f"  {k:<28} {coefs[i]:>8.3f} {ours_s}  {bucket:<6} {note}")


def _parse_voto(v):
    if v is None:
        return None
    m = re.search(r"(\d+(?:[.,]\d+)?)", str(v))
    return float(m.group(1).replace(",", ".")) if m else None
