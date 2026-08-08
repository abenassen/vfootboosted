from __future__ import annotations

import csv
import io
import logging
from datetime import timedelta
from random import Random

from django.contrib.auth.models import User
from django.db import transaction
from django.db.models import Count, Q
from django.http import Http404
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from rest_framework import status
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from realdata.models import (
    CompetitionSeason,
    Match,
    MatchAppearance,
    Player,
    PlayerTeamStint,
)
from vfoot.api.league_serializers import (
    AddRosterPlayerSerializer,
    AuctionAssignSerializer,
    BulkAssignRosterSerializer,
    CompetitionStageBuildSerializer,
    CompetitionStageCreateSerializer,
    CompetitionStageUpdateSerializer,
    CompetitionStageRuleCreateSerializer,
    CompetitionScheduleSerializer,
    CompetitionSchedulePreviewSerializer,
    CompetitionPrizeCreateSerializer,
    CompetitionUpdateSerializer,
    CompetitionTemplateSerializer,
    CompetitionWizardPreviewSerializer,
    CompetitionWizardSerializer,
    CreateAuctionSerializer,
    CreateLeagueSerializer,
    ImportRosterCSVSerializer,
    JoinLeagueSerializer,
    MarketToggleSerializer,
    MatchdayConcludeSerializer,
    NominateSerializer,
    PlaceBidSerializer,
    QualificationRuleCreateSerializer,
    RemoveRosterPlayerSerializer,
    UpdateMemberRoleSerializer,
    UpdateMyTeamSerializer,
)
from vfoot.models import (
    AuctionBid,
    AuctionEvent,
    AuctionNomination,
    AuctionSession,
    CompetitionQualificationRule,
    CompetitionStage,
    CompetitionStageParticipant,
    CompetitionStageRule,
    CompetitionTeam,
    CompetitionPrize,
    FantasyCompetition,
    FantasyFixture,
    FantasyFixtureDetail,
    FantasyLeague,
    FantasyMatchday,
    LeagueDecision,
    FantasyRosterSlot,
    FantasyTeam,
    LeagueMembership,
    LeaguePlayerRole,
    OfficeOverride,
    SavedLineupSnapshot,
)
from vfoot.services.auction_engine import (
    ROLES as AUCTION_ROLES, check_purchase, league_role_map, player_role, team_budgets,
)
from vfoot.services.auction_realtime import broadcast_auction
import os as _os
from functools import lru_cache as _lru_cache

from django.conf import settings as _settings

from vfoot.services.name_search import matches as name_matches
from vfoot.services.player_profiles import player_profiles
from vfoot.services.vector_zone_scoring import load_calibration
from vfoot.services.fantasy_simulation import (
    bulk_assign_players_to_teams,
    generate_knockout_fixtures,
    generate_round_robin_fixtures,
)
from vfoot.services import competition_calendar, competition_plan
from vfoot.services.competition_prizes import describe_condition, prize_winner_team_ids
from vfoot.services.competition_stages import (
    MAX_LEGS,
    build_default_stage_graph,
    competition_round_rows,
    recompute_round_layout,
    resolve_pending_stages,
    resolve_stage,
    stage_has_results,
)
from vfoot.services.competition_wizard import WizardError, create_competition, qualified_team_count
from vfoot.services.league_competitions import main_competition
from vfoot.services.formation_rules import CLASSIC_CONSTRAINTS, validate_classic_lineup
from vfoot.services.classic_pagella import (
    elapsed_minutes, get_reference, match_in_progress, pagella_for_match,
)
from vfoot.services.league_decisions import (
    accept_all_proposals, attention_count, cast_vote, market_blocked_reason,
    open_role_decisions, resolve as resolve_decision, unavailable_players,
    undecided_player_ids,
)
from vfoot.services.listone import snapshot_league_listone
from vfoot.services.listone import eligible_player_ids
from vfoot.services.player_ratings import (
    latest_market_values, player_values, previous_season_with_data,
)
from vfoot.services.match_resolver import matchday_fixtures_by_team
from vfoot.services import (
    currency, honours, knockout, lineup_deadline, lineup_repair, matchday_state,
)

log = logging.getLogger(__name__)

# Frozen listone role (POR/DIF/CEN/ATT) -> frontend lineup taxonomy (GK/DEF/MID/ATT).
_CLASSIC_ROLE_TO_LINEUP = {"POR": "GK", "DIF": "DEF", "CEN": "MID", "ATT": "ATT"}


def _membership_or_404(league: FantasyLeague, user_id: int) -> LeagueMembership:
    m = LeagueMembership.objects.filter(league=league, user_id=user_id).first()
    if not m:
        raise Http404("Not a member of this league")
    return m


def _ensure_admin(league: FantasyLeague, user_id: int) -> LeagueMembership:
    m = _membership_or_404(league, user_id)
    if m.role != LeagueMembership.ROLE_ADMIN:
        raise Http404("Admin privileges required")
    return m


class LeagueListCreateView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        memberships = (LeagueMembership.objects.filter(user=request.user)
                       .select_related("league", "team",
                                       "league__reference_season__competition",
                                       "league__reference_season__season"))
        data = []
        for m in memberships:
            season = m.league.reference_season
            data.append(
                {
                    "league_id": m.league_id,
                    "name": m.league.name,
                    "role": m.role,
                    "invite_code": m.league.invite_code,
                    "market_open": m.league.market_open,
                    "team_name": m.team.name if hasattr(m, "team") else None,
                    "team_crest": m.team.crest if hasattr(m, "team") else "",
                    "reference_season": (
                        {
                            "id": season.id,
                            "name": str(season),
                            "competition": season.competition.name,
                            "season": season.season.code,
                        }
                        if season else None
                    ),
                }
            )
        return Response(data)

    @transaction.atomic
    def post(self, request):
        s = CreateLeagueSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        data = s.validated_data

        reference_season = get_object_or_404(
            CompetitionSeason, id=data["reference_season_id"])
        league = FantasyLeague.objects.create(
            name=data["name"], owner=request.user, reference_season=reference_season,
            mode=data.get("mode", FantasyLeague.MODE_AURA),
            initial_budget=data.get("initial_budget", 1000),
            slots_gk=data.get("slots_gk", 3), slots_def=data.get("slots_def", 8),
            slots_mid=data.get("slots_mid", 8), slots_fwd=data.get("slots_fwd", 6))
        membership = LeagueMembership.objects.create(
            league=league,
            user=request.user,
            role=LeagueMembership.ROLE_ADMIN,
        )
        team = FantasyTeam.objects.create(league=league, manager=membership, name=data["team_name"])

        # Draw the listone straight away: a classic league IS its frozen roles, and
        # the questions the inference cannot answer are the admin's first piece of
        # work, not a surprise on the morning of the auction. Deferring it to the
        # market opening meant a freshly created league reported "nessuna decisione
        # in sospeso" while a dozen were in fact waiting to be raised.
        decisions = 0
        if league.mode == FantasyLeague.MODE_CLASSIC and league.reference_season_id:
            decisions = snapshot_league_listone(league).get("decisions_opened", 0)

        return Response(
            {
                "league_id": league.id,
                "name": league.name,
                "invite_code": league.invite_code,
                "invite_link": f"/join/{league.invite_code}",
                "team_id": team.id,
                "decisions_opened": decisions,
            },
            status=status.HTTP_201_CREATED,
        )


class LeagueJoinView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        s = JoinLeagueSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        data = s.validated_data

        league = get_object_or_404(FantasyLeague, invite_code=data["invite_code"])

        if LeagueMembership.objects.filter(league=league, user=request.user).exists():
            return Response({"detail": "Sei già iscritto a questa lega."}, status=status.HTTP_200_OK)

        membership = LeagueMembership.objects.create(
            league=league,
            user=request.user,
            role=LeagueMembership.ROLE_MANAGER,
        )
        team = FantasyTeam.objects.create(league=league, manager=membership, name=data["team_name"])

        return Response(
            {
                "league_id": league.id,
                "team_id": team.id,
                "name": league.name,
                "role": membership.role,
            },
            status=status.HTTP_201_CREATED,
        )


class LeagueDetailView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _membership_or_404(league, request.user.id)

        members = LeagueMembership.objects.filter(league=league).select_related("user")
        teams = FantasyTeam.objects.filter(league=league).select_related("manager__user")
        records = _league_wide_records(league)

        season = league.reference_season
        return Response(
            {
                "league_id": league.id,
                "name": league.name,
                "mode": league.mode,
                "market_open": league.market_open,
                "max_substitutions": league.max_substitutions,
                "defense_bonus_enabled": league.defense_bonus_enabled,
                "defense_bonus_mode": league.defense_bonus_mode,
                "keeper_clean_sheet_enabled": league.keeper_clean_sheet_enabled,
                "home_advantage_bonus": league.home_advantage_bonus,
                "enforce_lineup_deadline": league.enforce_lineup_deadline,
                "lineup_lock_mode": league.lineup_lock_mode,
                "initial_budget": league.initial_budget,
                "roster_slots": {"POR": league.slots_gk, "DIF": league.slots_def,
                                 "CEN": league.slots_mid, "ATT": league.slots_fwd},
                "roster_size": league.roster_size(),
                # Once an auction has been created the economy is frozen (editing it
                # would rewrite what was paid); the settings page uses this to lock.
                "auction_started": league.auction_sessions.exclude(
                    status=AuctionSession.STATUS_DRAFT).exists(),
                "invite_code": league.invite_code,
                "invite_link": f"/join/{league.invite_code}",
                "reference_season": (
                    {
                        "id": season.id,
                        "name": str(season),
                        "competition": season.competition.name,
                        "season": season.season.code,
                    }
                    if season
                    else None
                ),
                "members": [
                    {
                        "membership_id": m.id,
                        "user_id": m.user_id,
                        "username": m.user.username,
                        "role": m.role,
                    }
                    for m in members
                ],
                "teams": [
                    {
                        "team_id": t.id,
                        "name": t.name,
                        "crest": t.crest,
                        "manager_user_id": t.manager.user_id,
                        "manager_username": t.manager.user.username,
                        # Record aggregated across ALL competitions, not one chosen
                        # implicitly: a league has no single table, so points and
                        # rank would be a lie. Wins/draws/losses and goals for/against
                        # are the format-agnostic summary that always makes sense.
                        "record": records.get(t.id, {"played": 0, "wins": 0, "draws": 0,
                                                     "losses": 0, "goals_for": 0,
                                                     "goals_against": 0}),
                    }
                    for t in teams
                ],
            }
        )


_COMPETITION_END_DETAIL = {
    FantasyCompetition.FORMAT_LEAGUE: "campionato concluso",
    FantasyCompetition.FORMAT_CUP: "coppa assegnata",
    FantasyCompetition.FORMAT_GROUPS_KNOCKOUT: "gironi e finali conclusi",
}

# Above which a real transfer is NEWS to a fantasy league. Deliberately its own
# number, and well above league_decisions.RELEVANCE_MIN_VALUE_EUR (€5M, "worth an
# admin ruling on his role"): leagues are drawn up in August with the real market
# still running, so signings keep landing for weeks, and a floor that admitted
# every squad filler would turn the home page into a transfer ticker and bury the
# league's own life underneath it. Only the genuine coups.
NEWS_MIN_VALUE_EUR = 10_000_000


def _eur_short(value: int) -> str:
    """12_500_000 -> "12,5 M€". Italian decimal comma, and no decimals when round."""
    m = value / 1_000_000
    return (f"{m:.0f} M€" if abs(m - round(m)) < 0.05
            else f"{m:.1f} M€".replace(".", ","))


# More seed rows than this written in the same MINUTE is a listone being drawn, not
# a market. The real market tops a league up one player at a time, hours apart; a
# seeding writes hundreds at once. Keyed on the burst itself rather than on "the
# first one", because a listone can be drawn more than once — a league whose
# reference season is set later, a --reset, a demo re-seed — and every one of those
# bursts is the league being built, never news.
_SEEDING_BURST = 20


def _real_signings(league, limit: int) -> list[dict]:
    """Players who have JOINED THIS LEAGUE'S LISTONE since it was drawn.

    Read from LeaguePlayerRole, not from the roster stint, and that is the whole
    difficulty. A stint records when we first SAW a player in a squad, not when he
    signed: the first import of a season opens one for all 660 players on the same
    day and leaves ``transfer_kind`` empty, so "stint opened" cannot tell a January
    arrival from the initial squad load. Deriving the feed from it turned the home
    page into a list of 660 transfers dated at the season's first scrape.

    A league's own listone can tell them apart, by SHAPE rather than by date: it is
    written in bursts of hundreds at a stroke, while the market tops it up one
    player at a time. So a role row is news when it was written more or less alone
    — which is also the more useful question, since it is per league and answers
    "who is new to MY listone" rather than "who moved somewhere in Serie A".

    Seeded rows only: an admin row is a decision being answered, and the feed
    already carries those under their own kind.
    """
    if not league.reference_season_id:
        return []
    rows = list(LeaguePlayerRole.objects
                .filter(league=league, source=LeaguePlayerRole.SOURCE_SEED)
                .select_related("player")
                .order_by("-created_at")[:limit * 4])
    if not rows:
        return []

    # Count each minute over the WHOLE listone, not just the page we fetched: a
    # burst bigger than the page would otherwise look small enough to report.
    per_minute: dict = {}
    for at in (LeaguePlayerRole.objects
               .filter(league=league, source=LeaguePlayerRole.SOURCE_SEED)
               .values_list("created_at", flat=True)):
        per_minute[at.replace(second=0, microsecond=0)] = (
            per_minute.get(at.replace(second=0, microsecond=0), 0) + 1)

    fresh = [r for r in rows
             if per_minute.get(r.created_at.replace(second=0, microsecond=0), 0)
             <= _SEEDING_BURST]
    if not fresh:
        return []
    # Only the ones that are actually news. Leagues are usually drawn up in August
    # with the real market still running, so signings keep landing for weeks; the
    # floor is its own and well above the one that sends a player to the admin's
    # queue, because "worth deciding a role for" and "worth telling the league
    # about" are different questions and the second is much rarer.
    values = latest_market_values([r.player_id for r in fresh])
    out = []
    for r in fresh:
        # `or 0`, not a default: latest_market_values KEYS a player whose latest
        # quote is NULL, so .get() returns None rather than missing.
        value = values.get(r.player_id) or 0
        if value < NEWS_MIN_VALUE_EUR:
            continue
        name = r.player.short_name or r.player.full_name
        out.append({
            "kind": "mercato_reale",
            "at": r.created_at.isoformat(),
            "text": f"{name} entra nel listone",
            "detail": _eur_short(value),
            "team_id": None,
            "crest": None,
        })
    return out


# Per quanti giorni una notizia importante tiene la cima del blocco. Cinque: un
# fine settimana e i giorni intorno, cioè il tempo che passa fra due aperture
# dell'app di chi non la guarda tutti i giorni. Più a lungo e la notizia diventa
# lo sfondo su cui si smette di posare gli occhi.
NEWS_PIN_DAYS = 5


class LeagueActivityView(APIView):
    """What has happened in the league lately, newest first — salvo ciò che è IN
    EVIDENZA, che sta in cima finché è fresco (v. NEWS_PIN_DAYS).

    Merged from the records that already exist rather than from a new event table:
    a roster slot knows when it was acquired, a decision when it was settled, a
    matchday when it was concluded, a competition that it has run out of rounds.
    That keeps the feed honest — it cannot drift from what actually happened — at
    the cost of a few small queries.
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _membership_or_404(league, request.user.id)
        limit = min(int(request.query_params.get("limit") or 12), 50)

        items: list[dict] = []

        for slot in (FantasyRosterSlot.objects
                     .filter(team__league=league, released_at__isnull=True)
                     .select_related("player", "team")
                     .order_by("-acquired_at")[:limit]):
            items.append({
                "kind": "acquisto",
                "at": slot.acquired_at.isoformat() if slot.acquired_at else None,
                "text": f"{slot.player.short_name or slot.player.full_name} → {slot.team.name}",
                "detail": currency.amount(slot.purchase_price) if slot.purchase_price else None,
                "team_id": slot.team_id,
                "crest": slot.team.crest,
            })

        items.extend(_real_signings(league, limit))

        for d in (LeagueDecision.objects
                  .filter(league=league, status=LeagueDecision.STATUS_RESOLVED)
                  .select_related("player")
                  .order_by("-resolved_at")[:limit]):
            items.append({
                "kind": "decisione",
                "at": d.resolved_at.isoformat() if d.resolved_at else None,
                "text": d.title,
                "detail": d.outcome or None,
                "team_id": None,
                "crest": None,
            })

        for md in (FantasyMatchday.objects
                   .filter(league=league, status=FantasyMatchday.STATUS_CONCLUDED)
                   .order_by("-concluded_at")[:limit]):
            items.append({
                "kind": "giornata",
                "at": md.concluded_at.isoformat() if md.concluded_at else None,
                "text": f"Conclusa la giornata {md.real_matchday}",
                "detail": None,
                "team_id": None,
                "crest": None,
            })

        # The end of a competition and the honours it settled. Both derived, like
        # everything else here — and both dated from the ledger, so they appear
        # exactly where the matchday that decided them appears and not at the top
        # of the feed for ever.
        board = honours.league_honours(league)
        for row in board["finished"]:
            comp = row["competition"]
            items.append({
                "kind": "competizione",
                "at": row["at"].isoformat() if row["at"] else None,
                "text": f"{comp.name}: è finita",
                "detail": _COMPETITION_END_DETAIL.get(comp.format, "competizione conclusa"),
                "team_id": None,
                "crest": None,
            })
        for award in board["awards"]:
            # Nome e stemma di ALLORA, non di adesso: questa riga e' datata al
            # giorno in cui il premio fu vinto, e mostrarla con lo stemma che la
            # squadra porta oggi sarebbe raccontare un fatto con l'aria di un
            # altro giorno. Chi non era ancora congelato ricade sulla squadra viva.
            winners = award["winners"]
            prize = award["prize"]
            names = [w["name"] for w in winners]
            items.append({
                "kind": "premio",
                "at": award["at"].isoformat() if award["at"] else None,
                "text": f"{prize.icon or '🏆'} {prize.name}: {', '.join(names)}",
                "detail": f"{award['competition'].name} · {describe_condition(prize)}",
                # One winner gets his crest next to the line; a shared record has
                # no single face and gets none.
                "team_id": winners[0]["team_id"] if len(winners) == 1 else None,
                "crest": winners[0]["crest"] if len(winners) == 1 else None,
            })

        # IN EVIDENZA, ma non per sempre. Una notizia importante che restasse in
        # cima finché non la si archivia diventa arredamento: la si smette di
        # leggere dopo il secondo giorno e da lì in poi occupa il posto delle
        # notizie vere. Quindi la precedenza SCADE, ed è la data della notizia a
        # farla scadere — non una lettura, che nessuno registra.
        #
        # Il flag è per notizia e non per tipo, così un giorno lo si potrà
        # accendere a mano su una qualsiasi; oggi lo accende da sé solo un premio,
        # che è l'unica cosa che questa lega produce e che si aspetta davvero.
        now = timezone.now()
        for item in items:
            item.setdefault("important", item["kind"] == "premio")
            at = parse_datetime(item["at"]) if item["at"] else None
            item["pinned"] = bool(
                item["important"] and at is not None
                and (now - at) <= timedelta(days=NEWS_PIN_DAYS)
            )

        # Undated rows (older data, or a field never filled) sink rather than
        # jumping to the top of a feed sorted by a missing value.
        items.sort(key=lambda i: (i["pinned"], i["at"] or ""), reverse=True)
        return Response(items[:limit])


class LeagueHonoursView(APIView):
    """L'albo d'oro DELLA LEGA, e se la lega è finita.

    Diverso da quello del fantallenatore, che attraversa le leghe: qui la domanda
    è "come è andata questa stagione, in questa lega", e la risposta ha senso di
    esistere tutta insieme solo quando non c'è più niente da giocare.

    ``is_over`` è la condizione che vale la pena spiegare: non "l'ultima giornata
    è stata conclusa" — un campionato può finire settimane prima di una coppa che
    gli corre accanto — ma OGNI competizione della lega è chiusa. È la stessa
    bandiera che scrive ``honours.complete_competition``, quindi non può dire di
    sì mentre un tabellone ha ancora una finale da giocare.
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _membership_or_404(league, request.user.id)

        total = league.competitions.count()
        board = honours.league_honours(league)
        finished = board["finished"]
        ends = [row["at"] for row in finished if row["at"]]
        return Response({
            "is_over": bool(total and len(finished) == total),
            "competitions_total": total,
            "competitions_finished": len(finished),
            # Quando è finita l'ULTIMA: la data della fine della lega.
            "finished_at": max(ends).isoformat() if ends else None,
            "awards": [{
                "prize_id": a["prize"].id,
                "name": a["prize"].name,
                "icon": a["prize"].icon or "🏆",
                "condition_label": describe_condition(a["prize"]),
                "competition_id": a["competition"].id,
                "competition_name": a["competition"].name,
                "competition_format": a["competition"].format,
                # Identità CONGELATA all'assegnazione, come ovunque nell'albo:
                # ribattezzare la squadra non riscrive le coppe già vinte.
                "winners": a["winners"],
                "at": a["at"].isoformat() if a["at"] else None,
            } for a in board["awards"]],
        })


def _visible_leagues(viewer, manager) -> list[int] | None:
    """Le leghe di ``manager`` che ``viewer`` puo' vedere: quelle che condividono.

    None vuol dire "tutte", e capita in un caso solo: stai guardando te stesso.

    Non perche' un trofeo sia un segreto, ma perche' il nome di una lega in cui
    uno gioca altrove sono affari suoi e non di tutto il sito — e perche' "con chi
    altro gioca questa persona" e' esattamente il genere di cosa che non deve
    uscire da un'app di fantacalcio. Nessuna lega in comune: 404, come se la
    pagina non esistesse, che e' meno di quanto direbbe un rifiuto esplicito.
    """
    if manager.id == viewer.id:
        return None
    mine = set(LeagueMembership.objects.filter(user=viewer)
               .values_list("league_id", flat=True))
    theirs = set(LeagueMembership.objects.filter(user=manager)
                 .values_list("league_id", flat=True))
    shared = mine & theirs
    if not shared:
        raise Http404("No league in common")
    return list(shared)


class ManagerHonoursView(APIView):
    """L'albo d'oro di un fantallenatore — il suo, o quello di chiunque altro.

    Deliberately not league-scoped. A league lasts one season and a manager does
    not: a palmarès that started again every August would be worth nothing, and
    what makes it worth looking at is precisely that the cups pile up.

    What the VIEWER may see is another matter: see ``_visible_leagues``.
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, user_id: int):
        manager = get_object_or_404(User, id=user_id)
        visible = _visible_leagues(request.user, manager)
        awards = honours.manager_honours(manager, leagues=visible)
        return Response({
            "user_id": manager.id,
            "username": manager.username,
            "awards": [honours.serialize_award(a) for a in awards],
        })


class ManagerProfileView(APIView):
    """La scheda pubblica di un fantallenatore: chi è, e dove gioca.

    Esiste perché l'albo d'oro non appartiene a una lega. La rosa sì — è la
    proprietà di UNA squadra in UNA lega, e finisce con quella; il fantallenatore
    no, e i suoi trofei attraversano i campionati. Tenerli sulla pagina delle rose
    li faceva sembrare una cosa della lega corrente, e per vedere quelli di un
    avversario bisognava passare dalla sua rosa, che è tutt'altra domanda.

    Cosa NON c'è qui, di proposito: l'email e ogni altro dato di contatto. Questa
    pagina la può aprire chiunque condivida una lega, quindi porta solo quello che
    uno mette in campo — nome, faccia, squadre, trofei.

    Le leghe mostrate sono quelle in comune (vedi ``_visible_leagues``), con la
    squadra che ci schiera: è il ponte verso la sua rosa, ed è anche il motivo per
    cui la squadra e il fantallenatore ora si cliccano separatamente.
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, user_id: int):
        manager = get_object_or_404(User, id=user_id)
        visible = _visible_leagues(request.user, manager)

        memberships = (LeagueMembership.objects.filter(user=manager)
                       .select_related("league", "team")
                       .order_by("-joined_at", "-id"))
        if visible is not None:
            memberships = memberships.filter(league_id__in=visible)

        leagues = []
        for m in memberships:
            team = getattr(m, "team", None)
            leagues.append({
                "league_id": m.league_id,
                "name": m.league.name,
                "mode": m.league.mode,
                "role": m.role,
                "joined_at": m.joined_at.isoformat() if m.joined_at else None,
                # Nessuna squadra è uno stato normale, non un errore: si entra in
                # una lega prima di averne una.
                "team_id": team.id if team else None,
                "team_name": team.name if team else None,
                "team_crest": team.crest if team else "",
            })

        profile = getattr(manager, "profile", None)
        return Response({
            "user_id": manager.id,
            "username": manager.username,
            "avatar": profile.avatar if profile else "",
            "joined_at": manager.date_joined.isoformat() if manager.date_joined else None,
            "is_self": manager.id == request.user.id,
            "leagues": leagues,
        })


class LeagueMyTeamView(APIView):
    """The caller's OWN team inside one league: rename it, give it a crest.

    League-scoped on purpose, and deliberately not part of /auth/me. The avatar
    identifies the MANAGER and there is one per account; the name and the crest
    belong to one team in one league, and the same person fields a different team
    in every league he joins. Putting these on the profile page would have meant
    editing a league-scoped thing from a page that has no league.

    Admins get no say here: a team's own name is the manager's business. What an
    admin can already do league-wide is elsewhere.
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def patch(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        membership = _membership_or_404(league, request.user.id)
        team = getattr(membership, "team", None)
        if team is None:
            raise Http404("No team in this league")

        s = UpdateMyTeamSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        data = s.validated_data

        updated = []
        if "name" in data:
            name = data["name"]
            # unique_together is (league, name), so a clash would otherwise surface
            # as an IntegrityError — a 500 for something the user can fix himself.
            # Checked case-INSENSITIVELY: "Real Madrid" and "real madrid" in the
            # same standings table is a bug report waiting to happen, even though
            # the database would accept both.
            taken = (FantasyTeam.objects.filter(league=league, name__iexact=name)
                     .exclude(pk=team.pk).exists())
            if taken:
                return Response(
                    {"name": ["In questa lega esiste già una squadra con questo nome."]},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            team.name = name
            updated.append("name")
        if "crest" in data:
            team.crest = data["crest"]
            updated.append("crest")

        if updated:
            team.save(update_fields=updated)
        return Response({"team_id": team.id, "name": team.name, "crest": team.crest})


def _league_wide_records(league) -> dict:
    """Per-team W/D/L and goals across EVERY competition in the league.

    A league is a set of competitions of possibly different shapes (championship,
    knockout), with no designated one — so a per-competition table read out of
    context (points, rank) misleads. This sums finished fixtures league-wide, and
    reports goals rather than points, which mean the same thing in every format.
    ``home_total``/``away_total`` on a fixture ARE the goals (the readable score);
    the fantasy vote total lives on the detail as ``vfoot_home``. So the goals are
    those fields directly — no threshold conversion, which would double-count."""
    rec: dict[int, dict] = {}

    def row(tid: int) -> dict:
        return rec.setdefault(tid, {"played": 0, "wins": 0, "draws": 0, "losses": 0,
                                    "goals_for": 0, "goals_against": 0})

    for fx in (FantasyFixture.objects
               .filter(competition__league=league, status=FantasyFixture.STATUS_FINISHED)
               .values_list("home_team_id", "away_team_id", "home_total", "away_total")):
        htid, atid, ht, at = fx
        hg, ag = int(round(ht)), int(round(at))
        h, a = row(htid), row(atid)
        h["played"] += 1; a["played"] += 1
        h["goals_for"] += hg; h["goals_against"] += ag
        a["goals_for"] += ag; a["goals_against"] += hg
        if ht > at:
            h["wins"] += 1; a["losses"] += 1
        elif ht < at:
            a["wins"] += 1; h["losses"] += 1
        else:
            h["draws"] += 1; a["draws"] += 1
    return rec


class MemberRoleUpdateView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def patch(self, request, league_id: int, membership_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _ensure_admin(league, request.user.id)

        target = get_object_or_404(LeagueMembership, id=membership_id, league=league)
        s = UpdateMemberRoleSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        new_role = s.validated_data["role"]

        # Safety invariant: a league must always keep at least one admin.
        if target.role == LeagueMembership.ROLE_ADMIN and new_role != LeagueMembership.ROLE_ADMIN:
            admin_count = LeagueMembership.objects.filter(league=league, role=LeagueMembership.ROLE_ADMIN).count()
            if admin_count <= 1:
                return Response(
                    {"detail": "Non puoi rimuovere l'ultimo amministratore della lega."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        target.role = new_role
        target.save(update_fields=["role"])

        return Response({"membership_id": target.id, "role": target.role})


def _ensure_players_decided(league, player_ids):
    """Guard the money moments, PER PLAYER.

    A role settled after the bidding changes what people paid for, so a player
    whose role is still an open question cannot be auctioned or put on a roster.
    But only HE waits: freezing the whole market was tolerable for the opening
    listone and wrong for the rest of the season, where a single January signing
    would otherwise stop everyone else from trading.

    Names the players rather than only refusing — a gate that says "no" without
    saying who is a gate nobody can act on. Returns a 400 Response, or None.
    """
    blocked = unavailable_players(league, player_ids)
    if not blocked:
        return None
    names = ", ".join(b["name"] for b in blocked[:6])
    more = f" e altri {len(blocked) - 6}" if len(blocked) > 6 else ""
    return Response(
        {"detail": f"Ruolo ancora da decidere per {names}{more}: "
                   "non sono disponibili finche' l'amministratore non decide.",
         "code": "pending_decisions", "players": blocked},
        status=status.HTTP_400_BAD_REQUEST)


class MarketToggleView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def patch(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _ensure_admin(league, request.user.id)
        s = MarketToggleSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        if s.validated_data["is_open"]:
            # Catch up with the real market first. Roles are frozen but the roster
            # is not, and a player who arrived after the listone was drawn has no
            # frozen role: seeding him here is what stops a January signing from
            # slipping past the gate and being priced before anyone has agreed
            # what he is. Additive — nothing already decided is touched.
            # Catch up with the real market. Roles are frozen but the roster is
            # not, and a player who arrived after the listone was drawn has no
            # frozen role. Opening the market is NOT refused for it: only the
            # players still in limbo are, one by one, where they are used.
            if league.mode == FantasyLeague.MODE_CLASSIC:
                snapshot_league_listone(league)
        league.market_open = s.validated_data["is_open"]
        league.save(update_fields=["market_open"])
        return Response({"league_id": league.id, "market_open": league.market_open})


class LeagueSettingsUpdateView(APIView):
    """Admin-editable league settings (currently the max number of bench
    substitutions applied at scoring time)."""

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def patch(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _ensure_admin(league, request.user.id)
        fields: list[str] = []

        if "max_substitutions" in request.data:
            try:
                value = int(request.data.get("max_substitutions"))
            except (TypeError, ValueError):
                return Response({"detail": "max_substitutions non valido."}, status=status.HTTP_400_BAD_REQUEST)
            if not (0 <= value <= 11):
                return Response({"detail": "max_substitutions deve essere tra 0 e 11."}, status=status.HTTP_400_BAD_REQUEST)
            league.max_substitutions = value
            fields.append("max_substitutions")

        if "defense_bonus_enabled" in request.data:
            league.defense_bonus_enabled = bool(request.data.get("defense_bonus_enabled"))
            fields.append("defense_bonus_enabled")

        if "defense_bonus_mode" in request.data:
            mode = request.data.get("defense_bonus_mode")
            valid = {c[0] for c in FantasyLeague.DEF_BONUS_MODE_CHOICES}
            if mode not in valid:
                return Response({"detail": f"defense_bonus_mode deve essere in {sorted(valid)}."},
                                status=status.HTTP_400_BAD_REQUEST)
            league.defense_bonus_mode = mode
            fields.append("defense_bonus_mode")

        if "keeper_clean_sheet_enabled" in request.data:
            league.keeper_clean_sheet_enabled = bool(request.data.get("keeper_clean_sheet_enabled"))
            fields.append("keeper_clean_sheet_enabled")

        if "home_advantage_bonus" in request.data:
            try:
                bonus = float(request.data.get("home_advantage_bonus") or 0)
            except (TypeError, ValueError):
                return Response({"detail": "home_advantage_bonus non valido."},
                                status=status.HTTP_400_BAD_REQUEST)
            # Un tetto c'e', ed e' basso di proposito: sei punti di fantavoto sono
            # un gol pieno regalato prima del calcio d'inizio. Il fattore campo e'
            # una spinta, non un handicap.
            if not (0 <= bonus <= 6):
                return Response({"detail": "Il fattore campo deve essere tra 0 e 6."},
                                status=status.HTTP_400_BAD_REQUEST)
            league.home_advantage_bonus = bonus
            fields.append("home_advantage_bonus")

        if "enforce_lineup_deadline" in request.data:
            league.enforce_lineup_deadline = bool(request.data.get("enforce_lineup_deadline"))
            fields.append("enforce_lineup_deadline")

        if "lineup_lock_mode" in request.data:
            mode = str(request.data.get("lineup_lock_mode") or "")
            if mode not in dict(FantasyLeague.LOCK_MODE_CHOICES):
                return Response({"detail": "lineup_lock_mode non valido."},
                                status=status.HTTP_400_BAD_REQUEST)
            league.lineup_lock_mode = mode
            fields.append("lineup_lock_mode")

        # Auction economy (budget + roster slots). Frozen once an auction started:
        # a mid-auction change would rewrite the affordability of bids already made.
        econ_keys = {"initial_budget", "slots_gk", "slots_def", "slots_mid", "slots_fwd"}
        if econ_keys & set(request.data.keys()):
            if league.auction_sessions.exclude(status=AuctionSession.STATUS_DRAFT).exists():
                return Response(
                    {"detail": "L'asta e' gia' iniziata: budget e slot non sono piu' modificabili."},
                    status=status.HTTP_400_BAD_REQUEST)
            for key, lo in (("initial_budget", 1), ("slots_gk", 0), ("slots_def", 0),
                            ("slots_mid", 0), ("slots_fwd", 0)):
                if key in request.data:
                    try:
                        value = int(request.data.get(key))
                    except (TypeError, ValueError):
                        return Response({"detail": f"{key} non valido."}, status=status.HTTP_400_BAD_REQUEST)
                    if value < lo:
                        return Response({"detail": f"{key} deve essere >= {lo}."},
                                        status=status.HTTP_400_BAD_REQUEST)
                    setattr(league, key, value)
                    fields.append(key)

        if not fields:
            return Response({"detail": "Nessuna impostazione fornita."}, status=status.HTTP_400_BAD_REQUEST)
        league.save(update_fields=fields)
        return Response({
            "league_id": league.id,
            "max_substitutions": league.max_substitutions,
            "defense_bonus_enabled": league.defense_bonus_enabled,
            "defense_bonus_mode": league.defense_bonus_mode,
            "keeper_clean_sheet_enabled": league.keeper_clean_sheet_enabled,
            "home_advantage_bonus": league.home_advantage_bonus,
            "enforce_lineup_deadline": league.enforce_lineup_deadline,
            "lineup_lock_mode": league.lineup_lock_mode,
            "initial_budget": league.initial_budget,
            "roster_slots": {"POR": league.slots_gk, "DIF": league.slots_def,
                             "CEN": league.slots_mid, "ATT": league.slots_fwd},
        })


class RealSeasonListView(APIView):
    """Real competition seasons available to use as a league's reference."""

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from realdata.models import CompetitionSeason

        seasons = (
            CompetitionSeason.objects.select_related("competition", "season")
            .annotate(
                _matchdays=Count(
                    "matches__matchday",
                    filter=Q(matches__matchday__isnull=False),
                    distinct=True,
                )
            )
            .order_by("-season__code", "competition__name")
        )
        return Response(
            [
                {
                    "id": cs.id,
                    "name": str(cs),
                    "competition": cs.competition.name,
                    "season": cs.season.code,
                    "matchdays": int(cs._matchdays or 0),
                }
                for cs in seasons
            ]
        )


class LeagueReferenceSeasonView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def patch(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _ensure_admin(league, request.user.id)
        season_id = request.data.get("reference_season_id")

        # IMMUTABLE once set: rosters, frozen listone and the calendar all hang off
        # the reference season, so changing it mid-life would invalidate them. Only
        # legacy leagues that never got one can still have it assigned here.
        if league.reference_season_id is not None:
            same = (season_id not in (None, "", 0)
                    and int(season_id) == league.reference_season_id)
            if not same:
                return Response(
                    {"detail": "La stagione di riferimento non è modificabile: "
                               "rose, listone e calendario dipendono da essa."},
                    status=status.HTTP_400_BAD_REQUEST)
        elif season_id in (None, "", 0):
            return Response({"detail": "Stagione di riferimento obbligatoria."},
                            status=status.HTTP_400_BAD_REQUEST)
        else:
            league.reference_season = get_object_or_404(CompetitionSeason, id=season_id)
            league.save(update_fields=["reference_season"])
        season = league.reference_season
        return Response(
            {
                "league_id": league.id,
                "reference_season": (
                    {
                        "id": season.id,
                        "name": str(season),
                        "competition": season.competition.name,
                        "season": season.season.code,
                    }
                    if season
                    else None
                ),
            }
        )


class TeamRosterView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, league_id: int, team_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _membership_or_404(league, request.user.id)

        team = get_object_or_404(FantasyTeam, id=team_id, league=league)
        slots = FantasyRosterSlot.objects.filter(team=team, released_at__isnull=True).select_related("player")

        return Response(
            {
                "team_id": team.id,
                "team_name": team.name,
                "players": [
                    {
                        "player_id": s.player_id,
                        "name": s.player.short_name or s.player.full_name,
                        "price": s.purchase_price,
                    }
                    for s in slots
                ],
            }
        )


class TeamRosterAddView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, league_id: int, team_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _ensure_admin(league, request.user.id)
        if not league.market_open:
            return Response({"detail": "Market is closed."}, status=status.HTTP_400_BAD_REQUEST)
        team = get_object_or_404(FantasyTeam, id=team_id, league=league)
        s = AddRosterPlayerSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        data = s.validated_data
        blocked = _ensure_players_decided(league, [data["player_id"]])
        if blocked:
            return blocked

        player = get_object_or_404(Player, id=data["player_id"])

        already = FantasyRosterSlot.objects.filter(team__league=league, player=player, released_at__isnull=True).first()
        if already:
            return Response({"detail": "Player already assigned in this league."}, status=status.HTTP_400_BAD_REQUEST)

        slot = FantasyRosterSlot.objects.create(team=team, player=player, purchase_price=data["purchase_price"])
        return Response({"slot_id": slot.id, "player_id": player.id}, status=status.HTTP_201_CREATED)


class TeamRosterRemoveView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, league_id: int, team_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _ensure_admin(league, request.user.id)
        if not league.market_open:
            return Response({"detail": "Market is closed."}, status=status.HTTP_400_BAD_REQUEST)
        # No role gate here: releasing a player never depends on his role. Nor
        # should the case arise — anyone on a roster was bought, so he had a role
        # at the time, and a role settled in a league never becomes an open
        # question again (see open_role_decisions).
        team = get_object_or_404(FantasyTeam, id=team_id, league=league)
        s = RemoveRosterPlayerSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        player_id = s.validated_data["player_id"]

        slot = FantasyRosterSlot.objects.filter(team=team, player_id=player_id, released_at__isnull=True).first()
        if not slot:
            return Response({"detail": "Player not in active roster."}, status=status.HTTP_404_NOT_FOUND)

        slot.released_at = timezone.now()
        slot.save(update_fields=["released_at"])
        # Same invariant as a market settlement: a lineup that is still open must not
        # be left holding a player the team no longer has. There is no incoming
        # player here, so the slot is vacated — the manager can refill it, since a
        # locked lineup is never touched.
        vacated = lineup_repair.swap_player(league, team.id, player_id, None)
        return Response({"lineups_vacated": vacated}, status=status.HTTP_200_OK)


class LeagueRosterBulkAssignView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _ensure_admin(league, request.user.id)
        if not league.market_open:
            return Response({"detail": "Market is closed."}, status=status.HTTP_400_BAD_REQUEST)

        s = BulkAssignRosterSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        data = s.validated_data
        assignments = data.get("assignments")
        if assignments:
            # Same per-player gate as a single add: a bulk import must not be a
            # side door around a role nobody has agreed on yet.
            blocked = _ensure_players_decided(
                league, [r.get("player_id") for r in assignments if r.get("player_id")])
            if blocked:
                return blocked
            teams_by_id = {t.id: t for t in FantasyTeam.objects.filter(league=league).select_related("manager__user")}
            teams_by_name = {t.name.lower(): t for t in teams_by_id.values()}
            teams_by_manager = {t.manager.user.username.lower(): t for t in teams_by_id.values()}

            created = 0
            for row in assignments:
                target_team = None
                if "team_id" in row and str(row["team_id"]).isdigit():
                    target_team = teams_by_id.get(int(row["team_id"]))
                if not target_team and row.get("team_name"):
                    target_team = teams_by_name.get(str(row["team_name"]).strip().lower())
                if not target_team and row.get("manager_username"):
                    target_team = teams_by_manager.get(str(row["manager_username"]).strip().lower())
                if not target_team:
                    continue

                try:
                    player_id = int(row.get("player_id"))
                except (TypeError, ValueError):
                    continue
                player = Player.objects.filter(id=player_id).first()
                if not player:
                    continue

                if FantasyRosterSlot.objects.filter(team__league=league, player=player, released_at__isnull=True).exists():
                    continue

                price_raw = row.get("purchase_price", row.get("price", data.get("purchase_price", 1)))
                try:
                    price = max(1, int(price_raw))
                except (TypeError, ValueError):
                    price = 1

                FantasyRosterSlot.objects.create(team=target_team, player=player, purchase_price=price)
                created += 1

            return Response({"assigned_players": created, "mode": "explicit"})

        if not data.get("player_ids"):
            return Response(
                {
                    "detail": "Provide deterministic assignments using team_name or manager_username. "
                    "Random distribution is available only via player_ids fallback."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        count = bulk_assign_players_to_teams(
            league_id=league.id,
            player_ids=data["player_ids"],
            purchase_price=data.get("purchase_price", 1),
            random_seed=data.get("random_seed", 42),
        )
        return Response({"assigned_players": count, "mode": "random"})


class LeagueRosterImportCSVView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _ensure_admin(league, request.user.id)

        s = ImportRosterCSVSerializer(data=request.data)
        s.is_valid(raise_exception=True)

        csv_text = s.validated_data.get("csv_text", "")
        if not csv_text and "file" in request.FILES:
            csv_text = request.FILES["file"].read().decode("utf-8")
        if not csv_text.strip():
            return Response({"detail": "No CSV content provided."}, status=status.HTTP_400_BAD_REQUEST)

        reader = csv.DictReader(io.StringIO(csv_text))
        headers = set(reader.fieldnames or [])
        if "player_id" not in headers:
            return Response({"detail": "CSV headers must include player_id."}, status=status.HTTP_400_BAD_REQUEST)
        if "team_name" not in headers and "manager_username" not in headers:
            return Response(
                {"detail": "CSV headers must include team_name or manager_username (plus player_id, optional price)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        rows = [
            {
                "team_name": row.get("team_name"),
                "manager_username": row.get("manager_username"),
                "player_id": row.get("player_id"),
                "price": row.get("price") or row.get("purchase_price"),
            }
            for row in reader
        ]
        created, skipped = _apply_roster_assignments(league, rows)
        return Response({"imported": created, "skipped": skipped})


def _apply_roster_assignments(league, rows):
    """Apply a list of {team_name|manager_username, player_id, price} assignments.

    Shared by the CSV and xlsx importers. Skips rows that don't resolve to a team
    or player, and players already on a roster in this league. Returns (created,
    skipped) counts.
    """
    teams = {t.name: t for t in FantasyTeam.objects.filter(league=league)}
    teams_by_manager = {
        t.manager.user.username: t
        for t in FantasyTeam.objects.filter(league=league).select_related("manager__user")
    }
    created = 0
    skipped = 0
    for row in rows:
        team = teams.get((row.get("team_name") or "").strip())
        if not team:
            team = teams_by_manager.get((row.get("manager_username") or "").strip())
        if not team:
            skipped += 1
            continue
        try:
            player_id = int(row.get("player_id") or 0)
            price = int(row.get("price") or 1)
        except (ValueError, TypeError):
            skipped += 1
            continue
        if not player_id:
            skipped += 1
            continue

        player = Player.objects.filter(id=player_id).first()
        if not player:
            skipped += 1
            continue

        if FantasyRosterSlot.objects.filter(
            team__league=league, player=player, released_at__isnull=True
        ).exists():
            skipped += 1
            continue

        FantasyRosterSlot.objects.create(team=team, player=player, purchase_price=max(1, price))
        created += 1

    return created, skipped


class LeagueRosterImportXLSXView(APIView):
    """Re-upload the filled listone .xlsx (see the frontend listoneXlsx exporter) to
    assign rosters in one shot. Reads the "Listone" sheet, mapping the player_id,
    "Assegnato a" (team name) and "Prezzo" columns to roster slots."""

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    LISTONE_SHEET = "Listone"
    COL_PLAYER_ID = "player_id"
    COL_ASSIGNED = "Assegnato a"
    COL_PRICE = "Prezzo"

    @transaction.atomic
    def post(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _ensure_admin(league, request.user.id)

        upload = request.FILES.get("file")
        if not upload:
            return Response({"detail": "Carica un file .xlsx."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            import openpyxl  # lazy: only needed for this endpoint
        except ImportError:
            return Response(
                {"detail": "Import xlsx non disponibile sul server (manca openpyxl)."},
                status=status.HTTP_501_NOT_IMPLEMENTED,
            )

        try:
            wb = openpyxl.load_workbook(upload, data_only=True, read_only=True)
        except Exception:
            return Response({"detail": "File .xlsx non valido."}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb[self.LISTONE_SHEET] if self.LISTONE_SHEET in wb.sheetnames else wb.worksheets[0]

        # Find the header row (the export puts an instruction on row 1, headers on
        # row 2) by scanning the first few rows for the player_id header.
        header_map = None
        header_row_idx = None
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
            labels = {str(v).strip(): i for i, v in enumerate(row) if v is not None}
            if self.COL_PLAYER_ID in labels:
                header_map = labels
                header_row_idx = r_idx
                break

        if not header_map or self.COL_ASSIGNED not in header_map:
            return Response(
                {"detail": f"Il foglio deve contenere le colonne '{self.COL_PLAYER_ID}' e '{self.COL_ASSIGNED}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        pid_i = header_map[self.COL_PLAYER_ID]
        team_i = header_map[self.COL_ASSIGNED]
        price_i = header_map.get(self.COL_PRICE)

        rows = []
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if pid_i >= len(row):
                continue
            pid = row[pid_i]
            team_name = row[team_i] if team_i < len(row) else None
            if pid is None or not team_name or not str(team_name).strip():
                continue  # unassigned player — skip
            price = row[price_i] if (price_i is not None and price_i < len(row)) else None
            rows.append({
                "player_id": pid,
                "team_name": str(team_name).strip(),
                "price": price,
            })

        created, skipped = _apply_roster_assignments(league, rows)
        return Response({"imported": created, "skipped": skipped})


class PlayerSearchView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        query = (request.query_params.get("q") or "").strip()
        if len(query) < 2:
            return Response([])

        try:
            limit = max(1, min(50, int(request.query_params.get("limit", "20"))))
        except ValueError:
            limit = 20

        league_id = request.query_params.get("league_id")
        assigned_in_league = []
        if league_id and str(league_id).isdigit():
            assigned_in_league = FantasyRosterSlot.objects.filter(
                team__league_id=int(league_id),
                released_at__isnull=True,
            ).values_list("player_id", flat=True)

        # Two passes. The database narrows with what SQL can do cheaply — an
        # icontains on either name — and if that finds nothing we fall back to the
        # forgiving matcher over the (bounded) candidate set. icontains is neither
        # accent- nor typo-tolerant, so "Leao" and "Mkitarian" used to come back
        # empty from the auction room while the listone, which filters in the
        # browser, found them. Same rules on both sides now.
        base = Player.objects.exclude(id__in=assigned_in_league)
        players = list(
            base.filter(Q(full_name__icontains=query) | Q(short_name__icontains=query))
            .order_by("short_name", "full_name")[:limit]
        )
        if not players:
            candidates = base.order_by("short_name", "full_name").values_list(
                "id", "short_name", "full_name")
            hit_ids = [pid for pid, short, full in candidates
                       if name_matches(query, short, full)][:limit]
            if hit_ids:
                by_id = Player.objects.in_bulk(hit_ids)
                players = [by_id[i] for i in hit_ids if i in by_id]

        return Response(
            [
                {
                    "player_id": p.id,
                    "name": p.short_name or p.full_name,
                    "full_name": p.full_name,
                }
                for p in players
            ]
        )


class CompetitionTemplateCreateView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _ensure_admin(league, request.user.id)

        s = CompetitionTemplateSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        data = s.validated_data

        if FantasyCompetition.objects.filter(league=league, name=data["name"]).exists():
            return Response(
                {"detail": "A competition with this name already exists in this league."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        comp = FantasyCompetition.objects.create(
            league=league,
            name=data["name"],
            competition_type=data["competition_type"],
            status=FantasyCompetition.STATUS_DRAFT if data.get("container_only", False) else FantasyCompetition.STATUS_ACTIVE,
            starts_at=data.get("starts_at"),
            ends_at=data.get("ends_at"),
        )

        fixtures = 0
        team_ids = []
        if not data.get("container_only", False):
            team_ids = data.get("team_ids")
            if team_ids is None:
                team_ids = list(FantasyTeam.objects.filter(league=league).values_list("id", flat=True))
            entries = [CompetitionTeam(competition=comp, team_id=tid) for tid in team_ids]
            CompetitionTeam.objects.bulk_create(entries)

            if comp.competition_type == FantasyCompetition.TYPE_ROUND_ROBIN:
                fixtures = generate_round_robin_fixtures(comp)
            else:
                fixtures = generate_knockout_fixtures(comp)
        else:
            team_ids = []

        return Response(
            {
                "competition_id": comp.id,
                "name": comp.name,
                "competition_type": comp.competition_type,
                "status": comp.status,
                "starts_at": comp.starts_at.isoformat() if comp.starts_at else None,
                "ends_at": comp.ends_at.isoformat() if comp.ends_at else None,
                "participants": len(team_ids),
                "fixtures_created": fixtures,
                "container_only": bool(data.get("container_only", False)),
            },
            status=status.HTTP_201_CREATED,
        )


class CompetitionWizardCreateView(APIView):
    """Create a competition whole: shape, field, calendar and honours in one call."""

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _ensure_admin(league, request.user.id)

        s = CompetitionWizardSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        data = s.validated_data

        try:
            result = create_competition(
                league,
                name=data["name"],
                fmt=data["format"],
                team_ids=data.get("team_ids"),
                qualification=data.get("qualification"),
                legs=data.get("legs", 1),
                knockout_legs=data.get("knockout_legs", 1),
                final_legs=data.get("final_legs"),
                groups=data.get("groups", 1),
                advance_per_group=data.get("advance_per_group", 2),
                points=data.get("points"),
                start_matchday=data.get("start_matchday"),
                end_matchday=data.get("end_matchday"),
                prizes=data.get("prizes"),
                seed=data.get("random_seed", 42),
            )
        except WizardError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        comp = result["competition"]
        return Response(
            {
                "competition": _serialize_competition(comp),
                "stages": _serialize_stages(comp),
                "schedule": result["schedule"],
                "resolution": result["resolution"],
            },
            status=status.HTTP_201_CREATED,
        )


class CompetitionWizardPreviewView(APIView):
    """What a spec would produce, before anything is written.

    The wizard needs to answer "how many rounds is that, and from which real
    matchday can it start" while the user is still choosing — and the honest way
    to answer is with the same arithmetic that will build it.
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _ensure_admin(league, request.user.id)
        s = CompetitionWizardPreviewSerializer(data=request.data or {})
        s.is_valid(raise_exception=True)
        data = s.validated_data

        fmt = data["format"]
        legs = max(1, min(MAX_LEGS, data.get("legs", 1)))
        groups = max(1, data.get("groups", 1))
        advance = max(1, data.get("advance_per_group", 2))

        qualification = data.get("qualification")
        floor = None
        constraint = None
        if qualification:
            source_stage = (
                CompetitionStage.objects.filter(id=qualification.get("source_stage_id"))
                .select_related("competition")
                .first()
            )
            if source_stage is None or source_stage.competition.league_id != league.id:
                return Response({"detail": "Stage di qualificazione non valido."}, status=status.HTTP_400_BAD_REQUEST)
            n = qualified_team_count(
                source_stage,
                qualification.get("mode") or CompetitionStageRule.MODE_TABLE_RANGE,
                qualification.get("rank_from"),
                qualification.get("rank_to"),
            )
            source_comp = source_stage.competition
            cut = qualification.get("source_round")
            rows = competition_round_rows(source_comp)
            valid = {r["round_no"] for r in rows}
            if cut is not None and int(cut) not in valid:
                return Response(
                    {"detail": f"«{source_comp.name}» arriva alla giornata {max(valid) if valid else 0}."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            cut = cut or (max(valid) if valid else None)
            source_md = (source_comp.round_calendar or {}).get(str(cut)) if cut else None
            if source_md:
                floor = int(source_md) + 1
                constraint = (
                    f"i partecipanti si decidono alla giornata {cut} di «{source_comp.name}» "
                    f"(giornata reale {source_md})"
                )
        else:
            n = len(
                list(FantasyTeam.objects.filter(league=league, id__in=data.get("team_ids") or []).values_list("id", flat=True))
            )

        if n < 2:
            return Response({"detail": "Servono almeno 2 squadre."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            stages = _plan_preview(fmt, n, legs, groups, advance,
                                   data.get("knockout_legs", 1), data.get("final_legs"))
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        total_rounds = 0
        order_spans: dict[int, int] = {}
        for st in stages:
            order_spans[st["order_index"]] = max(order_spans.get(st["order_index"], 0), st["rounds"])
        total_rounds = sum(order_spans.values())

        season = league.reference_season
        season_matchdays: list[int] = []
        if season is not None:
            season_matchdays = [
                int(md)
                for md in Match.objects.filter(competition_season_id=season.id, matchday__isnull=False)
                .order_by("matchday")
                .values_list("matchday", flat=True)
                .distinct()
                if md is not None
            ]

        return Response(
            {
                "teams": n,
                "stages": stages,
                "total_rounds": total_rounds,
                "min_start_matchday": floor,
                "constraint": constraint,
                "season_real_matchdays": season_matchdays,
            }
        )


def _plan_preview(fmt: str, n: int, legs: int, groups: int, advance: int,
                  knockout_legs: int = 1, final_legs: int | None = None) -> list[dict]:
    """Stage-by-stage shape of a spec, with no side effects.

    ``knockout_legs``/``final_legs`` must move in lockstep with ``_knockout_chain``:
    this is the number the wizard shows before building, and a preview that
    promises three giornate for something that will take five is worse than no
    preview at all."""
    from vfoot.services.competition_stages import knockout_stage_name, rounds_per_leg

    ko_legs = 2 if knockout_legs == 2 else 1
    fin_legs = ko_legs if final_legs is None else (2 if final_legs == 2 else 1)

    def bracket(entry: int, start_order: int) -> list[dict]:
        out: list[dict] = []
        order = start_order
        base = 1
        while base * 2 <= entry:
            base *= 2
        if entry != base:
            out.append(
                {
                    "name": "Turno preliminare",
                    "type": "knockout",
                    "order_index": order,
                    "teams": (entry - base) * 2,
                    "rounds": ko_legs,
                    "matches": (entry - base) * ko_legs,
                }
            )
            order += 1
        size = base
        while size >= 2:
            this_legs = fin_legs if size == 2 else ko_legs
            out.append(
                {
                    "name": knockout_stage_name(size),
                    "type": "knockout",
                    "order_index": order,
                    "teams": size,
                    "rounds": this_legs,
                    "matches": (size // 2) * this_legs,
                }
            )
            order += 1
            size //= 2
        return out

    if fmt == FantasyCompetition.FORMAT_LEAGUE:
        rounds = rounds_per_leg(n) * legs
        return [
            {
                "name": "Campionato",
                "type": "round_robin",
                "order_index": 1,
                "teams": n,
                "rounds": rounds,
                "matches": (n * (n - 1) // 2) * legs,
            }
        ]

    if fmt == FantasyCompetition.FORMAT_CUP:
        return bracket(n, 1)

    if groups > n // 2:
        raise ValueError("Troppi gironi per il numero di squadre.")
    qualified = groups * advance
    if qualified < 2:
        raise ValueError("Devono qualificarsi almeno 2 squadre.")
    if qualified > n:
        raise ValueError("Non possono qualificarsi più squadre di quante partecipano.")
    out: list[dict] = []
    for gi in range(groups):
        size = n // groups + (1 if gi < n % groups else 0)
        if advance > size:
            raise ValueError("Da ogni girone non possono passare più squadre di quante lo compongono.")
        out.append(
            {
                "name": "Girone unico" if groups == 1 else f"Girone {chr(ord('A') + gi)}",
                "type": "round_robin",
                "order_index": 1,
                "teams": size,
                "rounds": rounds_per_leg(size) * legs,
                "matches": (size * (size - 1) // 2) * legs,
            }
        )
    return out + bracket(qualified, 2)


def _result_view(comp: FantasyCompetition) -> str:
    """Which results view a competition needs: a round-robin → 'classifica' (table),
    a knockout → 'tabellone' (bracket), a mix of stages → 'risultati' (both)."""
    types = set(comp.stages.values_list("stage_type", flat=True))
    if not types:
        return "tabellone" if comp.competition_type == FantasyCompetition.TYPE_KNOCKOUT else "classifica"
    if len(types) == 1:
        return "tabellone" if next(iter(types)) == CompetitionStage.TYPE_KNOCKOUT else "classifica"
    return "risultati"


def _team_names(team_ids: list[int]) -> list[str]:
    if not team_ids:
        return []
    by_id = dict(FantasyTeam.objects.filter(id__in=team_ids).values_list("id", "name"))
    return [by_id.get(tid, str(tid)) for tid in team_ids]


def _serialize_competition(comp: FantasyCompetition) -> dict:
    participants = list(comp.participants.select_related("team", "team__manager__user"))
    rules = list(comp.qualification_rules.select_related("source_competition"))
    prizes = list(comp.prizes.select_related("source_stage"))
    awarded = honours.prize_winners(comp)
    fixture_count = comp.fixtures.count()
    finished_count = comp.fixtures.filter(status=FantasyFixture.STATUS_FINISHED).count()
    calendar = {str(k): int(v) for k, v in (comp.round_calendar or {}).items() if str(k).isdigit()}
    # Every round the PLAN foresees, not only the ones that have fixtures: a cup's
    # final exists as a reserved matchday long before the semifinal names its two
    # teams, and a calendar built from fixtures alone cannot show it at all. The
    # stage plan is computed ONCE and shared with the rounds — explaining a blockage
    # reads the ledger, and this serializer runs for every competition of the league
    # on every page.
    plans = competition_plan.stage_plan(comp)
    rounds = competition_plan.round_plan_rows(comp, plans=plans)
    return {
        "competition_id": comp.id,
        "name": comp.name,
        "competition_type": comp.competition_type,
        "format": comp.format,
        "result_view": _result_view(comp),
        "status": comp.status,
        # Structure is frozen once a result exists: a redraw would erase games that
        # have been played. The UI reads this to know which edits to offer.
        "structure_locked": finished_count > 0,
        "rounds": rounds,
        # Per PHASE, for the case where naming every undrawn round would be N copies
        # of the same sentence: a group stage entered by the top four of a
        # championship is one rule, not six placeholder matches.
        "stage_plan": plans,
        "round_calendar": calendar,
        "dependencies": competition_calendar.dependencies(comp),
        "points": {
            "win": comp.points_win,
            "draw": comp.points_draw,
            "loss": comp.points_loss,
        },
        "starts_at": comp.starts_at.isoformat() if comp.starts_at else None,
        "ends_at": comp.ends_at.isoformat() if comp.ends_at else None,
        "start_matchday": comp.start_matchday,
        "end_matchday": comp.end_matchday,
        "participants": [
            {
                "team_id": p.team_id,
                "team_name": p.team.name,
                "source": p.source,
                "manager_username": p.team.manager.user.username,
                "seed": p.seed,
            }
            for p in participants
        ],
        "qualification_rules": [
            {
                "rule_id": r.id,
                "source_competition_id": r.source_competition_id,
                "source_competition_name": r.source_competition.name,
                "source_stage": r.source_stage,
                "source_round": r.source_round,
                "mode": r.mode,
                "rank_from": r.rank_from,
                "rank_to": r.rank_to,
            }
            for r in rules
        ],
        "prizes": [_serialize_prize(p, awarded) for p in prizes],
        "fixtures": {"total": fixture_count, "finished": finished_count},
    }


def _serialize_stage(stage: CompetitionStage, bounds: dict | None = None) -> dict:
    participants = list(stage.participants.select_related("team", "team__manager__user"))
    rules = list(stage.rules_in.select_related("source_stage", "source_stage__competition"))
    fixtures_total = stage.fixtures.count()
    fixtures_finished = stage.fixtures.filter(status=FantasyFixture.STATUS_FINISHED).count()
    span = bounds or {}
    return {
        "stage_id": stage.id,
        "competition_id": stage.competition_id,
        "name": stage.name,
        "stage_type": stage.stage_type,
        "status": stage.status,
        "order_index": stage.order_index,
        "legs": stage.legs,
        "round_offset": stage.round_offset,
        "planned_rounds": stage.planned_rounds,
        "expected_participants": stage.expected_participants,
        "first_matchday": span.get("first_matchday"),
        "last_matchday": span.get("last_matchday"),
        "participants": [
            {
                "team_id": p.team_id,
                "team_name": p.team.name,
                "source": p.source,
                "manager_username": p.team.manager.user.username,
                "seed": p.seed,
            }
            for p in participants
        ],
        "rules_in": [
            {
                "rule_id": r.id,
                "source_stage_id": r.source_stage_id,
                "source_stage_name": r.source_stage.name,
                "source_competition_id": r.source_stage.competition_id,
                "source_competition_name": r.source_stage.competition.name,
                "mode": r.mode,
                "source_round": r.source_round,
                "rank_from": r.rank_from,
                "rank_to": r.rank_to,
            }
            for r in rules
        ],
        "fixtures": {"total": fixtures_total, "finished": fixtures_finished},
    }


def _serialize_stages(comp: FantasyCompetition) -> list[dict]:
    bounds = competition_calendar.stage_round_bounds(comp)
    stages = (
        CompetitionStage.objects.filter(competition=comp)
        .prefetch_related("participants__team__manager__user", "rules_in__source_stage__competition", "fixtures")
        .order_by("order_index", "id")
    )
    return [_serialize_stage(s, bounds.get(s.id)) for s in stages]


class CompetitionStageListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, competition_id: int):
        comp = get_object_or_404(FantasyCompetition, id=competition_id)
        _membership_or_404(comp.league, request.user.id)
        return Response(_serialize_stages(comp))


class CompetitionStageBuildDefaultView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, competition_id: int):
        comp = get_object_or_404(FantasyCompetition, id=competition_id)
        _ensure_admin(comp.league, request.user.id)
        s = CompetitionStageBuildSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        data = s.validated_data
        result = build_default_stage_graph(
            comp,
            allow_repechage=data.get("allow_repechage", False),
            seed=data.get("random_seed", 42),
            legs=data.get("legs", 1),
        )
        return Response(result, status=status.HTTP_201_CREATED)


class CompetitionStageCreateView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, competition_id: int):
        comp = get_object_or_404(FantasyCompetition, id=competition_id)
        _ensure_admin(comp.league, request.user.id)
        s = CompetitionStageCreateSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        data = s.validated_data

        stage = CompetitionStage.objects.create(
            competition=comp,
            name=data["name"],
            stage_type=data["stage_type"],
            order_index=data.get("order_index", 1),
            legs=max(1, min(MAX_LEGS, data.get("legs", 1))),
        )
        team_ids = data.get("team_ids") or []
        valid_team_ids = list(FantasyTeam.objects.filter(league=comp.league, id__in=team_ids).values_list("id", flat=True))
        CompetitionStageParticipant.objects.bulk_create(
            [
                CompetitionStageParticipant(stage=stage, team_id=tid, source=CompetitionStageParticipant.SOURCE_MANUAL)
                for tid in valid_team_ids
            ]
        )
        if data.get("expected_participants"):
            stage.expected_participants = int(data["expected_participants"])
            stage.save(update_fields=["expected_participants"])

        seed_raw = request.data.get("random_seed", 42)
        try:
            seed = int(seed_raw)
        except (TypeError, ValueError):
            seed = 42
        # A new stage renumbers everything after it: layout first, fixtures second.
        recompute_round_layout(comp)
        stage.refresh_from_db()
        resolve_stage(stage, seed=seed)
        competition_calendar.schedule(comp)
        stage.refresh_from_db()
        return Response(_serialize_stage(stage), status=status.HTTP_201_CREATED)


class CompetitionStageDetailUpdateView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def patch(self, request, stage_id: int):
        stage = get_object_or_404(CompetitionStage, id=stage_id)
        _ensure_admin(stage.competition.league, request.user.id)
        s = CompetitionStageUpdateSerializer(data=request.data or {})
        s.is_valid(raise_exception=True)
        data = s.validated_data

        structural = {"stage_type", "order_index", "legs", "team_ids", "expected_participants"} & set(data)
        if structural and stage_has_results(stage):
            return Response(
                {
                    "detail": (
                        "Questo turno ha già risultati: si può cambiare solo il nome. "
                        "Per rifarne la struttura elimina la competizione e ricreala."
                    ),
                    "locked_fields": sorted(structural),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        changed_fields: list[str] = []
        for field in ["name", "stage_type", "order_index", "legs", "expected_participants"]:
            if field in data:
                value = max(1, min(MAX_LEGS, data[field])) if field == "legs" else data[field]
                setattr(stage, field, value)
                changed_fields.append(field)
        if changed_fields:
            stage.save(update_fields=changed_fields)
        # Rounds move when the number of legs or the stage order changes, and every
        # later stage moves with them.
        if {"legs", "order_index", "stage_type", "expected_participants"} & set(data):
            recompute_round_layout(stage.competition)
            stage.refresh_from_db()
            if "team_ids" not in data:
                resolve_stage(stage, seed=int(data.get("random_seed", 42)))

        if "team_ids" in data:
            team_ids = data.get("team_ids") or []
            valid_team_ids = list(
                FantasyTeam.objects.filter(league=stage.competition.league, id__in=team_ids).values_list("id", flat=True)
            )
            CompetitionStageParticipant.objects.filter(
                stage=stage,
                source=CompetitionStageParticipant.SOURCE_MANUAL,
            ).exclude(team_id__in=valid_team_ids).delete()

            existing_manual = set(
                CompetitionStageParticipant.objects.filter(
                    stage=stage,
                    source=CompetitionStageParticipant.SOURCE_MANUAL,
                ).values_list("team_id", flat=True)
            )
            CompetitionStageParticipant.objects.bulk_create(
                [
                    CompetitionStageParticipant(
                        stage=stage,
                        team_id=tid,
                        source=CompetitionStageParticipant.SOURCE_MANUAL,
                    )
                    for tid in valid_team_ids
                    if tid not in existing_manual
                ]
            )

            seed = int(data.get("random_seed", 42))
            recompute_round_layout(stage.competition)
            stage.refresh_from_db()
            resolve_stage(stage, seed=seed)

        competition_calendar.schedule(stage.competition)
        stage.refresh_from_db()
        bounds = competition_calendar.stage_round_bounds(stage.competition)
        return Response(_serialize_stage(stage, bounds.get(stage.id)))

    @transaction.atomic
    def delete(self, request, stage_id: int):
        stage = get_object_or_404(CompetitionStage, id=stage_id)
        _ensure_admin(stage.competition.league, request.user.id)

        dependent_rules = list(
            CompetitionStageRule.objects.filter(source_stage=stage)
            .select_related("target_stage", "target_stage__competition")
            .order_by("target_stage__competition_id", "target_stage__order_index", "target_stage_id")
        )
        if dependent_rules:
            return Response(
                {
                    "detail": "Cannot delete stage: it is used to derive participants for other stages.",
                    "dependent_targets": [
                        {
                            "target_stage_id": r.target_stage_id,
                            "target_stage_name": r.target_stage.name,
                            "target_competition_id": r.target_stage.competition_id,
                            "target_competition_name": r.target_stage.competition.name,
                            "mode": r.mode,
                            "rank_from": r.rank_from,
                            "rank_to": r.rank_to,
                        }
                        for r in dependent_rules
                    ],
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        external_prizes = list(
            CompetitionPrize.objects.filter(source_stage=stage)
            .exclude(competition=stage.competition)
            .select_related("competition")
        )
        if external_prizes:
            return Response(
                {
                    "detail": "Cannot delete stage: it is referenced by prizes in other competitions.",
                    "dependent_prizes": [
                        {
                            "prize_id": p.id,
                            "prize_name": p.name,
                            "competition_id": p.competition_id,
                            "competition_name": p.competition.name,
                        }
                        for p in external_prizes
                    ],
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        competition = stage.competition
        stage.delete()
        recompute_round_layout(competition)
        competition_calendar.schedule(competition)
        return Response(status=status.HTTP_204_NO_CONTENT)


class CompetitionStageAddRuleView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, stage_id: int):
        stage = get_object_or_404(CompetitionStage, id=stage_id)
        _ensure_admin(stage.competition.league, request.user.id)
        s = CompetitionStageRuleCreateSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        data = s.validated_data

        source_stage = get_object_or_404(CompetitionStage, id=data["source_stage_id"])
        if source_stage.competition.league_id != stage.competition.league_id:
            return Response({"detail": "Source and target stages must belong to the same league."}, status=status.HTTP_400_BAD_REQUEST)
        if source_stage.id == stage.id:
            return Response({"detail": "Uno turno non può qualificare se stesso."}, status=status.HTTP_400_BAD_REQUEST)

        source_round = data.get("source_round")
        if source_round is not None:
            valid = {r["round_no"] for r in competition_round_rows(source_stage.competition)}
            if int(source_round) not in valid:
                return Response(
                    {
                        "detail": (
                            f"«{source_stage.competition.name}» arriva alla giornata "
                            f"{max(valid) if valid else 0}: la {source_round}ª non esiste."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        rule = CompetitionStageRule.objects.create(
            target_stage=stage,
            source_stage=source_stage,
            mode=data["mode"],
            source_round=source_round,
            rank_from=data.get("rank_from"),
            rank_to=data.get("rank_to"),
        )
        result = resolve_stage(stage, seed=data.get("random_seed", 42))
        # A dependency moves the earliest matchday this competition may start on.
        schedule_result = competition_calendar.schedule(stage.competition)
        return Response(
            {
                "rule_id": rule.id,
                "target_stage_id": stage.id,
                "source_stage_id": source_stage.id,
                "mode": rule.mode,
                "source_round": rule.source_round,
                "rank_from": rule.rank_from,
                "rank_to": rule.rank_to,
                "resolve": result,
                "schedule": schedule_result,
            },
            status=status.HTTP_201_CREATED,
        )


class CompetitionStageResolveView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, stage_id: int):
        stage = get_object_or_404(CompetitionStage, id=stage_id)
        _ensure_admin(stage.competition.league, request.user.id)
        seed_raw = request.data.get("random_seed", 42)
        try:
            seed = int(seed_raw)
        except (TypeError, ValueError):
            seed = 42
        result = resolve_stage(stage, seed=seed)
        return Response(result)


def _serialize_prize(prize: CompetitionPrize, winners_by_prize: dict | None = None) -> dict:
    """Un premio come lo legge il browser: la regola, e chi l'ha vinta.

    I vincitori sono quelli ASSEGNATI, non quelli che soddisferebbero la
    condizione adesso: prima della fine della competizione un premio è "ancora da
    assegnare" anche quando l'aritmetica lo ha già deciso. ``winners_by_prize``
    evita una query per premio a chi ne serializza una lista intera.
    """
    winners = (winners_by_prize if winners_by_prize is not None
               else honours.prize_winners(prize.competition)).get(prize.id, [])
    return {
        "prize_id": prize.id,
        "name": prize.name,
        "icon": prize.icon or "🏆",
        "condition_type": prize.condition_type,
        "condition_label": describe_condition(prize),
        "stat": prize.stat,
        "source_stage_id": prize.source_stage_id,
        "source_stage_name": prize.source_stage.name if prize.source_stage_id else None,
        "rank_from": prize.rank_from,
        "rank_to": prize.rank_to,
        "winner_team_ids": winners,
        "winner_team_names": _team_names(winners),
    }


class CompetitionPrizeListCreateView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, competition_id: int):
        comp = get_object_or_404(FantasyCompetition, id=competition_id)
        _membership_or_404(comp.league, request.user.id)
        prizes = CompetitionPrize.objects.filter(competition=comp).select_related("source_stage")
        awarded = honours.prize_winners(comp)
        return Response([_serialize_prize(p, awarded) for p in prizes])

    @transaction.atomic
    def post(self, request, competition_id: int):
        comp = get_object_or_404(FantasyCompetition, id=competition_id)
        _ensure_admin(comp.league, request.user.id)
        s = CompetitionPrizeCreateSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        data = s.validated_data

        cond = data["condition_type"]
        source_stage_id = data.get("source_stage_id")
        rank_from = data.get("rank_from")
        rank_to = data.get("rank_to")

        source_stage = None
        if cond in [CompetitionPrize.CONDITION_STAGE_TABLE_RANGE, CompetitionPrize.CONDITION_STAGE_WINNER, CompetitionPrize.CONDITION_STAGE_LOSER]:
            if not source_stage_id:
                return Response({"detail": "source_stage_id is required for stage-based prize conditions."}, status=status.HTTP_400_BAD_REQUEST)
            source_stage = get_object_or_404(CompetitionStage, id=source_stage_id, competition=comp)

        if cond in [CompetitionPrize.CONDITION_FINAL_TABLE_RANGE, CompetitionPrize.CONDITION_STAGE_TABLE_RANGE]:
            if rank_from is None:
                return Response({"detail": "rank_from is required for table range conditions."}, status=status.HTTP_400_BAD_REQUEST)
            if rank_to is None:
                rank_to = rank_from
            if rank_from <= 0 or rank_to <= 0 or rank_to < rank_from:
                return Response({"detail": "Invalid rank range."}, status=status.HTTP_400_BAD_REQUEST)
        else:
            rank_from = None
            rank_to = None

        stat = data.get("stat") or ""
        if cond in [CompetitionPrize.CONDITION_STAT_TOP, CompetitionPrize.CONDITION_STAT_BOTTOM]:
            if not stat:
                return Response({"detail": "Serve la misura su cui si vince il primato."},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            stat = ""

        prize = CompetitionPrize.objects.create(
            competition=comp,
            name=data["name"],
            icon=data.get("icon") or "🏆",
            condition_type=cond,
            stat=stat,
            source_stage=source_stage,
            rank_from=rank_from,
            rank_to=rank_to,
        )
        return Response(_serialize_prize(prize), status=status.HTTP_201_CREATED)


class CompetitionPrizeDeleteView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def delete(self, request, prize_id: int):
        prize = get_object_or_404(CompetitionPrize, id=prize_id)
        _ensure_admin(prize.competition.league, request.user.id)
        prize.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


def _source_fixtures_for_stage(source_comp: FantasyCompetition, stage: str, source_round: int | None = None):
    qs = FantasyFixture.objects.filter(competition=source_comp, status=FantasyFixture.STATUS_FINISHED)
    if source_round is not None:
        # Explicit round cut-off: snapshot the table after this round.
        qs = qs.filter(round_no__lte=max(1, source_round))
    elif stage == CompetitionQualificationRule.STAGE_HALF:
        max_round = FantasyFixture.objects.filter(competition=source_comp).order_by("-round_no").values_list("round_no", flat=True).first()
        if max_round:
            qs = qs.filter(round_no__lte=max(1, max_round // 2))
    return qs


def _table_ranking_team_ids(source_comp: FantasyCompetition, stage: str, source_round: int | None = None) -> list[int]:
    rows: dict[int, dict] = {}
    fixtures = _source_fixtures_for_stage(source_comp, stage, source_round)
    for fx in fixtures:
        for tid in [fx.home_team_id, fx.away_team_id]:
            rows.setdefault(tid, {"pts": 0, "gf": 0.0, "ga": 0.0})
        ht, at = fx.home_team_id, fx.away_team_id
        hs, as_ = fx.home_total, fx.away_total
        rows[ht]["gf"] += hs
        rows[ht]["ga"] += as_
        rows[at]["gf"] += as_
        rows[at]["ga"] += hs
        if hs > as_:
            rows[ht]["pts"] += source_comp.points_win
            rows[at]["pts"] += source_comp.points_loss
        elif hs < as_:
            rows[at]["pts"] += source_comp.points_win
            rows[ht]["pts"] += source_comp.points_loss
        else:
            rows[ht]["pts"] += source_comp.points_draw
            rows[at]["pts"] += source_comp.points_draw

    ranking = sorted(rows.items(), key=lambda kv: (kv[1]["pts"], kv[1]["gf"] - kv[1]["ga"], kv[1]["gf"]), reverse=True)
    return [tid for tid, _ in ranking]


def _winner_loser_from_source(source_comp: FantasyCompetition, stage: str, mode: str, source_round: int | None = None) -> list[int]:
    ranking = _table_ranking_team_ids(source_comp, stage, source_round)
    if not ranking:
        return []
    if mode == CompetitionQualificationRule.MODE_WINNER:
        return [ranking[0]]
    if mode == CompetitionQualificationRule.MODE_LOSER:
        return [ranking[-1]]
    return []


@transaction.atomic
def _resolve_rule_participants_and_regenerate(competition: FantasyCompetition) -> dict:
    manual_ids = set(
        CompetitionTeam.objects.filter(competition=competition, source=CompetitionTeam.SOURCE_MANUAL).values_list("team_id", flat=True)
    )
    CompetitionTeam.objects.filter(competition=competition, source=CompetitionTeam.SOURCE_RULE).delete()

    resolved_ids: set[int] = set()
    unresolved_rules = 0
    for rule in CompetitionQualificationRule.objects.filter(competition=competition).select_related("source_competition"):
        source = rule.source_competition
        if rule.mode == CompetitionQualificationRule.MODE_TABLE_RANGE:
            ranking = _table_ranking_team_ids(source, rule.source_stage, rule.source_round)
            if not ranking:
                unresolved_rules += 1
                continue
            rf = max(1, rule.rank_from or 1)
            rt = max(rf, rule.rank_to or rf)
            ids = ranking[rf - 1 : rt]
        else:
            ids = _winner_loser_from_source(source, rule.source_stage, rule.mode, rule.source_round)
            if not ids:
                unresolved_rules += 1
                continue
        for tid in ids:
            if tid not in manual_ids and tid not in resolved_ids:
                resolved_ids.add(tid)
                CompetitionTeam.objects.create(competition=competition, team_id=tid, source=CompetitionTeam.SOURCE_RULE)

    participants = list(CompetitionTeam.objects.filter(competition=competition).values_list("team_id", flat=True))
    fixtures_created = 0
    if len(participants) >= 2:
        # Build the full stage graph (bracket + progression rules for knockout,
        # regular-season stage for round-robin) rather than flat single-round
        # fixtures — so a rule-fed competition (e.g. cup fed by championship
        # top-N) gets a proper structure once its participants resolve.
        result = build_default_stage_graph(competition)
        fixtures_created = result.get("fixtures_created", 0)
        competition_calendar.schedule(competition)

    return {
        "competition_id": competition.id,
        "resolved_rule_participants": len(resolved_ids),
        "unresolved_rules": unresolved_rules,
        "fixtures_created": fixtures_created,
    }


class CompetitionListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _membership_or_404(league, request.user.id)
        comps = (
            FantasyCompetition.objects.filter(league=league)
            .prefetch_related(
                "participants__team__manager__user",
                "qualification_rules__source_competition",
                "prizes__source_stage",
                "fixtures",
            )
            .order_by("id")
        )
        return Response([_serialize_competition(c) for c in comps])


def _current_matchday(league: FantasyLeague):
    """The matchday the LEDGER is due to score next — never a statement about what is
    being played. See services/matchday_state for why the two are separate."""
    return matchday_state.ledger_matchday(league)


def _fixture_phase(fx: FantasyFixture, current_real_md: int | None,
                   awaiting_mds: set[int] | None = None) -> str:
    """concluded | current | awaiting | future | unscheduled — drives the UI badge."""
    if fx.status == FantasyFixture.STATUS_FINISHED:
        return "concluded"
    if fx.fantasy_matchday_id is None:
        return "unscheduled"
    real_md = fx.fantasy_matchday.real_matchday
    if awaiting_mds and real_md in awaiting_mds:
        return "awaiting"
    if current_real_md is None:
        return "future"
    if real_md == current_real_md:
        return "current"
    if real_md < current_real_md:
        return "concluded"
    return "future"


def _live_totals(league: FantasyLeague, fixtures, locked_mds: set[int]) -> dict:
    """{fixture_id: {"home_total", "away_total", "provisional"}} for the rounds that
    have begun and are not concluded.

    Computed here rather than left to the client because it is the same number the
    tabellino shows, from the same functions: a calendar that said "vs" while the
    tabellino behind it said 66-72 would be two answers to one question. One
    scorer per matchday — the half-second index is per ROUND, not per fixture.
    """
    if not locked_mds:
        return {}
    from vfoot.services.classic_matchday_scoring import live_scorer
    from vfoot.services.classic_scoring import Ruleset

    by_md: dict[int, list] = {}
    for fx in fixtures:
        md = fx.fantasy_matchday if fx.fantasy_matchday_id else None
        if (md is None or md.status == FantasyMatchday.STATUS_CONCLUDED
                or fx.status == FantasyFixture.STATUS_FINISHED
                or md.real_matchday not in locked_mds):
            continue
        by_md.setdefault(md.id, []).append(fx)
    out: dict[int, dict] = {}
    for group in by_md.values():
        md = group[0].fantasy_matchday
        ruleset = (Ruleset.from_snapshot(md.ruleset_snapshot) if md.ruleset_snapshot
                   else Ruleset.from_league(league))
        try:
            score = live_scorer(league, md, ruleset)
        except Exception:  # noqa: BLE001 — a calendar must render without the extra
            log.exception("Punteggi provvisori non calcolabili per la giornata %s", md.id)
            continue
        for fx in group:
            p = score(fx)
            out[fx.id] = {"home_total": p["home_goals"], "away_total": p["away_goals"],
                          "provisional": bool(p.get("provisional"))}
    return out


def _serialize_fixture_row(fx: FantasyFixture, my_team_id: int | None, current_real_md: int | None = None,
                           my_roster_ready: bool = False, awaiting_mds: set[int] | None = None,
                           locked_mds: set[int] | None = None,
                           live_totals: dict | None = None,
                           closed_mds: set[int] | None = None) -> dict:
    mine = bool(my_team_id and (fx.home_team_id == my_team_id or fx.away_team_id == my_team_id))
    played = fx.status == FantasyFixture.STATUS_FINISHED
    real_md = fx.fantasy_matchday.real_matchday if fx.fantasy_matchday_id else None
    locked = bool(locked_mds is not None and real_md is not None and real_md in locked_mds)
    # Two different closures, and they coincide only under the matchday-wide
    # deadline. `locked` = the round has begun, which is what makes a live tabellino
    # exist. `closed` = the manager has nothing left to decide — under the per-player
    # lock that is the LAST kickoff, and defaulting it to `locked` would take the
    # Formazione button away from him on Saturday afternoon.
    closed = locked if closed_mds is None else (real_md is not None and real_md in closed_mds)
    live = (live_totals or {}).get(fx.id)
    return {
        "fixture_id": fx.id,
        "competition_id": fx.competition_id,
        "competition_name": fx.competition.name,
        "stage_id": fx.stage_id,
        "stage_name": fx.stage.name if fx.stage_id else None,
        # TURNO per l'unità dentro una competizione della lega, GIORNATA solo per
        # quella del campionato vero. Sono due orologi e vogliono due parole: qui
        # c'era scritto "Giornata", e una riga di risultato finiva per leggersi
        # "Giornata 21 · giornata 21" — il conto interno della competizione
        # chiamato col nome del calendario su cui è giocato. Il nome di una fase
        # scelto dall'admin ("Semifinali") batte comunque il numero.
        "round_label": fx.stage.name if fx.stage_id else f"Turno {fx.round_no}",
        "fantasy_matchday_id": fx.fantasy_matchday_id,
        "real_matchday": fx.fantasy_matchday.real_matchday if fx.fantasy_matchday_id else None,
        "round_no": fx.round_no,
        "leg_no": fx.leg_no,
        "kickoff": fx.kickoff.isoformat() if fx.kickoff else None,
        "status": fx.status,
        "phase": _fixture_phase(fx, current_real_md, awaiting_mds),
        # Whether this matchday's lineups have locked (its first confirmed kickoff
        # has passed). Read from the REAL calendar, so it stays true no matter how
        # far behind the admin is with his conclusions.
        "lineup_locked": locked,
        "home_team": {"team_id": fx.home_team_id, "name": fx.home_team.name,
                      "crest": fx.home_team.crest},
        "away_team": {"team_id": fx.away_team_id, "name": fx.away_team.name,
                      "crest": fx.away_team.crest},
        "score": ({"home_total": fx.home_total, "away_total": fx.away_total} if played
                  else {"home_total": live["home_total"], "away_total": live["away_total"]}
                  if live else None),
        # The score above is a PARTIAL one: the round has begun and nobody has
        # counted it. Said explicitly rather than left to be inferred from the
        # status, because "0-0 because it has not started" and "0-0 at the
        # twentieth minute" are the same two numbers.
        "score_provisional": bool(live and live["provisional"]) if not played else False,
        # I PUNTEGGI, che non sono i gol: la somma dei fantavoto delle due
        # formazioni. Servono perché sono il primo spareggio di un turno secco
        # (v. services/knockout) e quindi il NUMERO che decide chi passa quando
        # una finale finisce in parità — dirlo senza mostrarlo lasciava «passa ai
        # punteggi» come un verdetto senza prova. Solo dove il tabellino c'è: una
        # lega aura non li ha, e lì lo spareggio scende al gradino dopo.
        "totals": ({"home": detail.vfoot_home, "away": detail.vfoot_away}
                   if (detail := getattr(fx, "detail", None)) is not None else None),
        "is_user_involved": mine,
        # Whether a lineup can be set for THIS fixture, decided here rather than in
        # the UI: it also depends on the roster, which the calendar does not load.
        # An empty roster has nothing to field, and offering the button anyway sent
        # people to a formation page with no players in it.
        # It also depends on the DEADLINE, which is the part that used to be missing:
        # a fixture is 'not played' until the admin concludes its matchday, so a
        # forgotten conclusion kept offering "Formazione" on a round played weeks
        # earlier — a button the save endpoint could only answer with a 409.
        "can_set_lineup": bool(
            mine and my_roster_ready and not played and fx.fantasy_matchday_id is not None
            and not closed
        ),
        # Whether /fixtures/<id> has anything to show. A concluded fixture has its
        # frozen tabellino; a LOCKED one has the live computation, which is the whole
        # point of following your own matchday while it is played. Only a round that
        # has not kicked off has nothing, and there the link would end on a 404.
        "has_detail": played or locked,
    }


class LeagueFixturesView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        membership = _membership_or_404(league, request.user.id)
        my_team_id = membership.team.id if hasattr(membership, "team") else None

        qs = (
            FantasyFixture.objects.filter(competition__league=league)
            # `detail` INCLUSO: la riga porta i punteggi di squadra, e senza
            # caricarlo qui ognuna delle duecento righe del calendario se lo
            # sarebbe andato a prendere da sola. `payload` resta fuori — sono due
            # numeri, non il referto di venticinque giocatori.
            .select_related("competition", "stage", "fantasy_matchday", "home_team",
                            "away_team", "detail")
            .defer("detail__payload")
            .order_by("-kickoff", "-id")
        )
        competition_id = request.query_params.get("competition_id")
        if competition_id and str(competition_id).isdigit():
            qs = qs.filter(competition_id=int(competition_id))

        current = _current_matchday(league)
        current_real_md = current.real_matchday if current else None
        awaiting_mds = {md.real_matchday for md in matchday_state.awaiting_matchdays(league)}
        locked_mds = (matchday_state.locked_matchdays(league.reference_season_id)
                      if league.reference_season_id else set())
        closed_mds = matchday_state.closed_matchdays(league)
        # One query for the whole calendar, not one per fixture.
        my_roster_ready = bool(my_team_id) and FantasyRosterSlot.objects.filter(
            team_id=my_team_id, released_at__isnull=True).exists()
        rows = list(qs[:200])
        live_totals = _live_totals(league, rows, locked_mds)
        items = [_serialize_fixture_row(fx, my_team_id, current_real_md, my_roster_ready,
                                        awaiting_mds, locked_mds, live_totals, closed_mds)
                 for fx in rows]
        return Response(items)


def _competition_round_nos(comp: FantasyCompetition) -> list[int]:
    return [row["round_no"] for row in competition_round_rows(comp)]


def _real_matchday_stats_bulk(real_competition_season_id: int, matchdays: list[int],
                              league: FantasyLeague | None = None) -> dict[int, dict]:
    """Lo stesso conto di ``_real_matchday_stats``, per TUTTE le giornate in una volta.

    Chiesto una giornata alla volta costava quattro interrogazioni ciascuna, e la
    pagina che le elenca ne chiede trentotto: erano centosettanta query per
    disegnare un calendario che non cambia mai. La regola (una partita conta una
    volta per accoppiamento, ed e' risolta se ``data_ready`` o se la lega ci ha
    messo un voto d'ufficio) e' identica — qui e' solo raggruppata per giornata.
    """
    if not matchdays:
        return {}
    office_match_ids = set()
    if league is not None:
        office_match_ids = set(
            OfficeOverride.objects.filter(league=league, is_active=True)
            .values_list("match_id", flat=True)
        )
    by_md: dict[int, dict[tuple[int, int], bool]] = {md: {} for md in matchdays}
    for mid, md, home_id, away_id, ready in Match.objects.filter(
        competition_season_id=real_competition_season_id, matchday__in=matchdays
    ).values_list("id", "matchday", "home_team_id", "away_team_id", "data_ready"):
        pairs = by_md.setdefault(int(md), {})
        key = (home_id, away_id)
        settled = bool(ready) or mid in office_match_ids
        pairs[key] = pairs.get(key, False) or settled

    out = {}
    for md, pairs in by_md.items():
        total = len(pairs)
        completed = sum(1 for v in pairs.values() if v)
        out[md] = {"total": total, "completed": completed,
                   "is_completed": total > 0 and completed == total}
    return out


def _real_matchday_stats(real_competition_season_id: int, real_matchday: int,
                         league: FantasyLeague | None = None) -> dict:
    # A postponed-and-replayed match appears TWICE in a real matchday: a stale
    # 'postponed' placeholder with no score, plus the rescheduled row that was
    # actually played (a different external_id, played weeks later — see Serie A
    # 2025-26 md24 Milan-Como, md16 with four postponements). Counting raw rows
    # makes the matchday look forever incomplete ("reale 10/11", "10/14"). Collapse by
    # the team pairing so each real fixture counts once and is 'completed' when ANY
    # of its rows has a final score. Within one matchday a (home, away) pairing is
    # unique, and a replay keeps the same home/away, so the pairing is a safe key.
    #
    # ``league`` makes the answer league-specific: a match this league has ruled on
    # with an office vote is settled AS FAR AS THIS LEAGUE IS CONCERNED, even though
    # it was never played. That is the whole point of the ruling — it is what unblocks
    # the conclusion — and it is per league by construction, so the league next door
    # still sees the round as incomplete and goes on waiting for the recovery.
    office_match_ids = set()
    if league is not None:
        office_match_ids = set(
            OfficeOverride.objects.filter(league=league, is_active=True)
            .values_list("match_id", flat=True)
        )
    #
    # Settled is DATA_READY, not "has a score". A live match has a score — the light
    # poll writes it every couple of minutes — so counting a score made a round with
    # three matches in progress read as complete, and the home said "la giornata 22 è
    # finita, puoi calcolare i punteggi" directly above "intanto si gioca la giornata
    # 22". Not only contradictory: the conclusion would have been refused anyway,
    # because scoring keys on data_ready too (see match_resolver.pending_matches). The
    # two now agree, which is the property that matters — the button is offered
    # exactly when it works.
    done_by_pair: dict[tuple[int, int], bool] = {}
    for mid, home_id, away_id, ready in Match.objects.filter(
        competition_season_id=real_competition_season_id, matchday=real_matchday
    ).values_list("id", "home_team_id", "away_team_id", "data_ready"):
        key = (home_id, away_id)
        settled = bool(ready) or mid in office_match_ids
        done_by_pair[key] = done_by_pair.get(key, False) or settled
    total = len(done_by_pair)
    completed = sum(1 for v in done_by_pair.values() if v)
    return {
        "total": total,
        "completed": completed,
        "is_completed": total > 0 and completed == total,
    }


def _stage_is_done(stage: CompetitionStage) -> bool:
    total = stage.fixtures.count()
    if total == 0:
        return False
    finished = stage.fixtures.filter(status=FantasyFixture.STATUS_FINISHED).count()
    return finished == total


class LeagueMatchdaySyncView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _ensure_admin(league, request.user.id)
        result = _sync_matchdays_for_league(league)
        return Response(result)


def _sync_matchdays_for_league(league: FantasyLeague) -> dict:
    fixtures = list(
        FantasyFixture.objects.filter(competition__league=league, source_real_match__isnull=False)
        .select_related("source_real_match")
        .order_by("id")
    )
    cache: dict[tuple[int, int], FantasyMatchday] = {}
    updates: list[FantasyFixture] = []
    linked = 0

    for fx in fixtures:
        if not fx.source_real_match:
            continue
        md = fx.source_real_match.matchday
        csid = fx.source_real_match.competition_season_id
        if md is None or not csid:
            continue
        key = (csid, int(md))
        fmd = cache.get(key)
        if not fmd:
            fmd, _ = FantasyMatchday.objects.get_or_create(
                league=league,
                real_competition_season_id=csid,
                real_matchday=int(md),
            )
            cache[key] = fmd
        if fx.fantasy_matchday_id != fmd.id:
            fx.fantasy_matchday_id = fmd.id
            updates.append(fx)
            linked += 1

    if updates:
        FantasyFixture.objects.bulk_update(updates, ["fantasy_matchday"], batch_size=500)

    return {"fixtures_linked": linked, "matchdays_touched": len(cache)}


class LeagueMatchdayListView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _membership_or_404(league, request.user.id)
        _sync_matchdays_for_league(league)

        rows = (
            FantasyMatchday.objects.filter(league=league)
            .select_related("real_competition_season__competition", "real_competition_season__season", "concluded_by")
            .order_by("real_matchday", "id")
        )
        current = _current_matchday(league)
        current_id = current.id if current else None
        # The other clock, read from the real calendar: what can still be fielded and
        # what is on the pitch right now. Deliberately independent of the conclusions
        # above, so an admin who is three matchdays behind does not move it.
        fieldable = matchday_state.next_fieldable_matchday(league)
        playing = matchday_state.playing_matchday(league)
        locks = (matchday_state.matchday_locks(league.reference_season_id)
                 if league.reference_season_id else {})
        locked = (matchday_state.locked_matchdays(league.reference_season_id)
                  if league.reference_season_id else set())
        # Which of these matchdays another competition is WAITING ON. The league
        # advances past an unclosed round on purpose; a cup fed by its table cannot,
        # and until this was said out loud a competition could sit undrawn for weeks
        # with nothing on any screen connecting the two.
        decides = competition_plan.matchday_impacts(league)
        # Tre conti che erano dentro il ciclo, uno per giornata: su una stagione
        # intera facevano centosettanta interrogazioni per un elenco che quasi
        # sempre non cambia. Sono gli stessi conti, fatti una volta sola.
        stats_by_md: dict[int, dict[int, dict]] = {}
        for csid in {md.real_competition_season_id for md in rows}:
            stats_by_md[csid] = _real_matchday_stats_bulk(
                csid, [md.real_matchday for md in rows if md.real_competition_season_id == csid],
                league)
        fx_counts: dict[int, list[int]] = {}
        for md_id, st in FantasyFixture.objects.filter(fantasy_matchday__in=rows).values_list(
                "fantasy_matchday_id", "status"):
            row = fx_counts.setdefault(md_id, [0, 0])
            row[0] += 1
            if st == FantasyFixture.STATUS_FINISHED:
                row[1] += 1

        payload = []
        for md in rows:
            real_stats = stats_by_md.get(md.real_competition_season_id, {}).get(
                md.real_matchday, {"total": 0, "completed": 0, "is_completed": False})
            fx_total, fx_finished = fx_counts.get(md.id, (0, 0))
            if md.status == FantasyMatchday.STATUS_CONCLUDED:
                phase = "concluded"
            elif md.status == FantasyMatchday.STATUS_AWAITING:
                phase = "awaiting"
            elif md.id == current_id:
                phase = "current"
            else:
                phase = "future"
            allowed, blocked_reason = matchday_state.can_conclude(league, md)
            if allowed and not real_stats["is_completed"]:
                allowed, blocked_reason = False, "La giornata reale non è ancora completata."
            elif allowed and fx_total == 0:
                allowed, blocked_reason = False, "Nessuna fixture fantasy associata."
            lock_at = locks.get(md.real_matchday)
            payload.append(
                {
                    "fantasy_matchday_id": md.id,
                    "league_id": league.id,
                    "status": md.status,
                    "phase": phase,
                    "is_fieldable": md.real_matchday == fieldable,
                    "is_playing": md.real_matchday == playing,
                    # THREE states, not two, and the middle one is the one that was
                    # missing. `has_kicked_off` says the round has begun — its first
                    # confirmed kickoff has passed — and it stays true for the whole
                    # round, including the Saturday night when nothing is on the
                    # pitch and the Monday when everything is over but the admin has
                    # not counted it yet. `is_playing` is the narrower "there is a
                    # ball rolling RIGHT NOW". A page that keyed on the second one
                    # made your own match come and go between kick-offs.
                    "has_kicked_off": md.real_matchday in locked,
                    "lineup_lock_at": lock_at.isoformat() if lock_at else None,
                    # Two different questions. `can_conclude` is "may this one be
                    # closed RIGHT NOW" (it enforces the order, so behind an unclosed
                    # matchday it is False). `awaits_conclusion` is "does the league
                    # owe this one" — true for every arrear, which is what the count
                    # of a forgotten admin's backlog is made of.
                    "can_conclude": allowed,
                    "conclude_blocked_reason": "" if allowed else blocked_reason,
                    "awaits_conclusion": (md.status != FantasyMatchday.STATUS_CONCLUDED
                                          and real_stats["is_completed"] and fx_total > 0),
                    "awaiting_since": md.awaiting_since.isoformat() if md.awaiting_since else None,
                    "awaiting_reason": md.awaiting_reason,
                    "real_competition_season": {
                        "id": md.real_competition_season_id,
                        "name": str(md.real_competition_season),
                        "competition": md.real_competition_season.competition.name,
                        "season": md.real_competition_season.season.code,
                    },
                    "real_matchday": md.real_matchday,
                    "real_completion": real_stats,
                    "fixtures": {"total": fx_total, "finished": fx_finished},
                    # The phases this matchday decides the field of, if any.
                    "decides": decides.get(md.real_matchday, []),
                    "concluded_at": md.concluded_at.isoformat() if md.concluded_at else None,
                    "concluded_by": md.concluded_by.username if md.concluded_by_id else None,
                }
            )
        return Response(payload)


class LeagueMatchdayConcludeView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, league_id: int, fantasy_matchday_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _ensure_admin(league, request.user.id)
        _sync_matchdays_for_league(league)
        md = get_object_or_404(FantasyMatchday, id=fantasy_matchday_id, league=league)

        s = MatchdayConcludeSerializer(data=request.data or {})
        s.is_valid(raise_exception=True)
        force = s.validated_data.get("force", False)

        # Conclude in order — with one deliberate exception: a matchday parked as
        # AWAITING may be closed whenever its postponed match is finally played,
        # without first having to close everything played since.
        current = _current_matchday(league)
        allowed, reason = matchday_state.can_conclude(league, md)
        if not allowed and not force:
            return Response(
                {"detail": reason,
                 "current_matchday_id": current.id if current else None},
                status=status.HTTP_400_BAD_REQUEST,
            )

        real_stats = _real_matchday_stats(md.real_competition_season_id, md.real_matchday, league)
        if not real_stats["is_completed"] and not force:
            return Response(
                {
                    "detail": "Real matchday is not completed yet.",
                    "real_completion": real_stats,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        fixtures = list(
            FantasyFixture.objects.filter(fantasy_matchday=md)
            .select_related("source_real_match", "stage", "competition")
            .order_by("id")
        )

        missing_source = 0
        missing_goals = 0
        updated = 0
        stage_ids: set[int] = set()

        if league.mode == FantasyLeague.MODE_CLASSIC:
            # Classic: the H2H result is the sum of the lineup's fantavoti converted to
            # goals (66/+6), NOT the real match scoreline. Freeze totals + per-player
            # payload + the ruleset used, all in this transaction.
            from vfoot.services.classic_scoring import Ruleset
            from vfoot.services.classic_matchday_scoring import score_and_persist_matchday

            ruleset = Ruleset.from_league(league)
            result = score_and_persist_matchday(
                md, league, ruleset, fixtures,
                s.validated_data.get("lineup_resolutions", {}), force, update_snapshot=True)
            if result["missing_teams"]:
                return Response(
                    {
                        "detail": "Alcune squadre non hanno la formazione: scegli 'forfait' o 'previous' per ognuna.",
                        "teams_without_lineup": result["missing_teams"],
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if result["pending_matches"]:
                # Somebody in a fielded XI plays a match that has not been played.
                # Scoring now would silently treat a postponement as a senza voto:
                # the honest options are to park the matchday (awaiting) or, once it
                # exists, to impose an office vote on the missing match.
                return Response(
                    {
                        "detail": "Ci sono partite della giornata non ancora giocate: "
                                  "metti la giornata in attesa oppure concludila forzando.",
                        "pending_matches": result["pending_matches"],
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            updated = result["updated"]
            stage_ids = result["stage_ids"]
        else:
            # Aura (and any non-classic mode): keep the real-scoreline behaviour.
            for fx in fixtures:
                src = fx.source_real_match
                if not src:
                    missing_source += 1
                    continue
                if src.home_goals is None or src.away_goals is None:
                    missing_goals += 1
                    continue
                fx.home_total = float(src.home_goals)
                fx.away_total = float(src.away_goals)
                fx.status = FantasyFixture.STATUS_FINISHED
                updated += 1
                if fx.stage_id:
                    stage_ids.add(fx.stage_id)

            if (missing_source > 0 or missing_goals > 0) and not force:
                return Response(
                    {
                        "detail": "Some fixtures are not scoreable yet (missing mapping or final real score).",
                        "missing_source_real_match": missing_source,
                        "missing_real_scores": missing_goals,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if fixtures:
                FantasyFixture.objects.bulk_update(
                    fixtures, ["home_total", "away_total", "status"], batch_size=500)

        stage_ids_to_resolve: set[int] = set()
        done_stages = 0
        for sid in stage_ids:
            stage = CompetitionStage.objects.filter(id=sid).first()
            if not stage:
                continue
            if _stage_is_done(stage):
                # Le sfide rimaste in parita' si decidono ai rigori, QUI: prima che
                # il turno successivo chieda chi e' passato, e una volta sola.
                knockout.settle_shootouts(list(
                    FantasyFixture.objects.filter(stage=stage)
                    .select_related("detail", "fantasy_matchday")))
                if stage.status != CompetitionStage.STATUS_DONE:
                    stage.status = CompetitionStage.STATUS_DONE
                    stage.save(update_fields=["status"])
                done_stages += 1
                targets = stage.rules_out.values_list("target_stage_id", flat=True)
                for tid in targets:
                    stage_ids_to_resolve.add(int(tid))

        resolved_targets = []
        for tid in sorted(stage_ids_to_resolve):
            target = CompetitionStage.objects.filter(id=tid).first()
            if not target:
                continue
            result = resolve_stage(target, seed=42)
            resolved_targets.append(result)

        # A cup fed by "the table after round 7" hangs off a source stage that is
        # NOT done — the championship keeps going. So after every conclusion, ask
        # every competition of the league whether anything it was waiting for has
        # now happened; the ones with nothing pending cost a single query.
        for other in FantasyCompetition.objects.filter(league=league):
            filled = resolve_pending_stages(other, seed=42)
            if filled["stages_filled"]:
                resolved_targets.append({"competition_id": other.id, **filled})

        md.status = FantasyMatchday.STATUS_CONCLUDED
        md.concluded_at = timezone.now()
        md.concluded_by = request.user
        # Concluding a parked matchday is how the wait ends; the parking marks go.
        md.awaiting_since = None
        md.awaiting_reason = ""
        md.nudged_at = None
        md.save(update_fields=["status", "concluded_at", "concluded_by",
                               "awaiting_since", "awaiting_reason", "nudged_at"])

        # LAST, and after the matchday is marked concluded: the honours are dated
        # from the ledger, so asking before this save would find the competition
        # complete and its date missing.
        finished = _competitions_that_just_ended(league)

        return Response(
            {
                "fantasy_matchday_id": md.id,
                "status": md.status,
                "real_completion": real_stats,
                "fixtures_scored": updated,
                "fixtures_total": len(fixtures),
                "missing_source_real_match": missing_source,
                "missing_real_scores": missing_goals,
                "done_stages": done_stages,
                "resolved_target_stages": resolved_targets,
                # What this conclusion ENDED, if anything. The admin clicks
                # "concludi" thirty-eight times and the thirty-eighth is the one
                # that assigns the scudetto; without this the screen said the same
                # thing as the other thirty-seven.
                "finished_competitions": finished,
            }
        )


def _describe_prize_changes(changes: list[dict]) -> list[dict]:
    """Un trofeo che cambia mano, detto in modo che l'admin capisca cos'ha fatto."""
    out = []
    for c in changes:
        prize = c["prize"]
        out.append({
            "prize_id": prize.id,
            "name": prize.name,
            "icon": prize.icon or "🏆",
            "competition_name": c["competition"].name,
            "now": _team_names(c["added"]),
            "before": _team_names(c["removed"]),
        })
    return out


def _competitions_that_just_ended(league) -> list[dict]:
    """Le competizioni che QUESTA conclusione ha chiuso, coi premi che ha assegnato.

    Solo quelle ancora aperte prima del clic: è ciò che rende questo l'annuncio di
    un evento e non un elenco permanente. Concludi un'altra giornata e non si
    ripete — i premi ormai sono scritti, e riassegnarli non cambierebbe nulla.
    """
    out = []
    for comp in FantasyCompetition.objects.filter(league=league).exclude(
            status=FantasyCompetition.STATUS_DONE):
        result = honours.complete_competition(comp)
        if result is None:
            continue
        winners = honours.prize_winners(comp)
        out.append({
            "competition_id": comp.id,
            "name": comp.name,
            "format": comp.format,
            "prizes": [{"name": p.name,
                        "icon": p.icon or "🏆",
                        "condition_label": describe_condition(p),
                        "winner_team_ids": winners.get(p.id, []),
                        "winner_team_names": _team_names(winners.get(p.id, []))}
                       for p in comp.prizes.all() if winners.get(p.id)],
        })
    return out


class LeagueMatchdayAwaitView(APIView):
    """Admin: park the current matchday as AWAITING, or bring it back.

    The gesture behind "una partita è stata rinviata: la lega va avanti, questa
    giornata la chiudo quando si recupera". Deliberately manual: whether to wait for
    the recovery or to impose an office vote is a decision of the league, not a fact
    of the calendar, and it is the one moment where the two are told apart.

    Body: {"awaiting": true|false (default true), "reason": "..."}.
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, league_id: int, fantasy_matchday_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _ensure_admin(league, request.user.id)
        md = get_object_or_404(FantasyMatchday, id=fantasy_matchday_id, league=league)

        awaiting = request.data.get("awaiting", True)
        if not isinstance(awaiting, bool):
            awaiting = str(awaiting).lower() not in ("false", "0", "")

        if md.status == FantasyMatchday.STATUS_CONCLUDED:
            return Response({"detail": "La giornata è già conclusa."},
                            status=status.HTTP_400_BAD_REQUEST)

        if awaiting:
            # Only the matchday the ledger is actually on can be parked: parking a
            # future one would leave a hole in the middle of the ledger rather than
            # letting the league step over the one it is stuck on.
            pointer = matchday_state.ledger_matchday(league)
            if pointer is None or pointer.id != md.id:
                return Response(
                    {"detail": "Si può mettere in attesa solo la giornata che la lega "
                               "deve ancora conteggiare."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            md.status = FantasyMatchday.STATUS_AWAITING
            md.awaiting_since = timezone.now()
            md.awaiting_reason = str(request.data.get("reason", ""))[:200]
        else:
            md.status = FantasyMatchday.STATUS_PLANNED
            md.awaiting_since = None
            md.awaiting_reason = ""
        md.save(update_fields=["status", "awaiting_since", "awaiting_reason"])

        new_pointer = matchday_state.ledger_matchday(league)
        return Response({
            "fantasy_matchday_id": md.id,
            "status": md.status,
            "awaiting_since": md.awaiting_since.isoformat() if md.awaiting_since else None,
            "awaiting_reason": md.awaiting_reason,
            "ledger_matchday": new_pointer.real_matchday if new_pointer else None,
            # Parking a matchday is the one gesture that can silently stall another
            # competition: the league steps over the round, and a cup reading its
            # table does not. Say which, at the moment the decision is made.
            "decides": competition_plan.matchday_impacts(league).get(md.real_matchday, []),
        })


class LeagueMatchdayOfficeVotesView(APIView):
    """Admin: the league's ruling on the matches of this matchday it will not wait for.

    GET  -> the matches of the round that have no final data, each with the office
            vote this league has already imposed on it (if any).
    POST -> {"match_ids": [...], "voto": 6.0, "reason": "..."} imposes (or updates)
            the ruling; {"remove": true} withdraws it.

    Per league by construction: the rows are keyed on (league, match), so one league
    imposing the 6 on Como-Milan leaves every other league free to wait for the
    recovery. Classic only — an imposed VOTE has no meaning in aura, where a fixture
    takes the real scoreline.
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def _rows(self, league, md):
        overrides = {o.match_id: o for o in OfficeOverride.objects.filter(
            league=league, fantasy_matchday=md)}
        out = []
        for m in (Match.objects
                  .filter(competition_season_id=md.real_competition_season_id,
                          matchday=md.real_matchday)
                  .select_related("home_team__team", "away_team__team")
                  .order_by("kickoff", "id")):
            o = overrides.get(m.id)
            # A postponed shell whose replay has already been played is not missing
            # data — it is a duplicate, and ruling on it would be ruling on nothing.
            superseded = m.status == Match.STATUS_POSTPONED and Match.objects.filter(
                competition_season_id=md.real_competition_season_id,
                matchday=md.real_matchday, home_team_id=m.home_team_id,
                away_team_id=m.away_team_id, data_ready=True).exists()
            if m.data_ready or superseded:
                continue
            out.append({
                "match_id": m.id,
                "home": m.home_team.team.name,
                "away": m.away_team.team.name,
                "status": m.status,
                "kickoff": m.kickoff.isoformat() if m.kickoff else None,
                "office_vote": o.voto if (o and o.is_active) else None,
                "reason": o.reason if o else "",
            })
        return out

    def get(self, request, league_id: int, fantasy_matchday_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _membership_or_404(league, request.user.id)
        md = get_object_or_404(FantasyMatchday, id=fantasy_matchday_id, league=league)
        return Response({"fantasy_matchday_id": md.id, "matches": self._rows(league, md)})

    @transaction.atomic
    def post(self, request, league_id: int, fantasy_matchday_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _ensure_admin(league, request.user.id)
        md = get_object_or_404(FantasyMatchday, id=fantasy_matchday_id, league=league)
        if league.mode != FantasyLeague.MODE_CLASSIC:
            return Response({"detail": "Il voto d'ufficio è disponibile solo per le leghe classic."},
                            status=status.HTTP_400_BAD_REQUEST)

        match_ids = [int(x) for x in (request.data.get("match_ids") or [])]
        if not match_ids:
            return Response({"detail": "Nessuna partita indicata."},
                            status=status.HTTP_400_BAD_REQUEST)
        matches = list(Match.objects.filter(
            id__in=match_ids, competition_season_id=md.real_competition_season_id,
            matchday=md.real_matchday))
        if len(matches) != len(set(match_ids)):
            return Response({"detail": "Alcune partite non appartengono a questa giornata."},
                            status=status.HTTP_400_BAD_REQUEST)

        if request.data.get("remove"):
            OfficeOverride.objects.filter(league=league, match_id__in=match_ids).delete()
            return Response({"removed": len(match_ids), "matches": self._rows(league, md)})

        try:
            voto = float(request.data.get("voto", 6.0))
        except (TypeError, ValueError):
            return Response({"detail": "Voto non valido."}, status=status.HTTP_400_BAD_REQUEST)
        if not 0.0 <= voto <= 10.0:
            return Response({"detail": "Il voto deve stare fra 0 e 10."},
                            status=status.HTTP_400_BAD_REQUEST)
        reason = str(request.data.get("reason", ""))[:200]

        for m in matches:
            OfficeOverride.objects.update_or_create(
                league=league, match=m,
                defaults={"fantasy_matchday": md, "voto": voto, "reason": reason,
                          "is_active": True, "created_by": request.user},
            )
        return Response({"applied": len(matches), "voto": voto,
                         "matches": self._rows(league, md)})


class LeagueMatchdayRecomputeView(APIView):
    """Admin: re-score an already CONCLUDED classic matchday, rewriting totals + payload
    (+ the ruleset snapshot when recomputing with current rules). This is the operative
    answer to "I changed the rules / fixed a vote after the matchday": the result and the
    tabellino are rewritten together, atomically, never leaving a stale detail.

    Body: use = "current" (default; re-read the league's live ruleset, update the
    snapshot) | "snapshot" (re-run with the frozen ruleset, e.g. after a vote fix);
    lineup_resolutions {team_id: forfait|previous}; force.
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, league_id: int, fantasy_matchday_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _ensure_admin(league, request.user.id)
        md = get_object_or_404(FantasyMatchday, id=fantasy_matchday_id, league=league)

        if league.mode != FantasyLeague.MODE_CLASSIC:
            return Response({"detail": "Ricalcolo disponibile solo per le leghe classic."},
                            status=status.HTTP_400_BAD_REQUEST)
        if md.status != FantasyMatchday.STATUS_CONCLUDED:
            return Response({"detail": "La giornata non è conclusa: usa 'Concludi'."},
                            status=status.HTTP_400_BAD_REQUEST)

        from vfoot.services.classic_scoring import Ruleset
        from vfoot.services.classic_matchday_scoring import score_and_persist_matchday

        use = request.data.get("use", "current")
        force = bool(request.data.get("force", False))
        resolutions = request.data.get("lineup_resolutions", {}) or {}

        if use == "snapshot" and md.ruleset_snapshot:
            ruleset = Ruleset.from_snapshot(md.ruleset_snapshot)
            update_snapshot = False
        else:
            ruleset = Ruleset.from_league(league)
            update_snapshot = True

        fixtures = list(
            FantasyFixture.objects.filter(fantasy_matchday=md)
            .select_related("home_team", "away_team", "competition", "stage")
            .order_by("id")
        )
        result = score_and_persist_matchday(
            md, league, ruleset, fixtures, resolutions, force, update_snapshot=update_snapshot)
        if result["missing_teams"]:
            return Response(
                {
                    "detail": "Alcune squadre non hanno la formazione: scegli 'forfait' o 'previous' per ognuna.",
                    "teams_without_lineup": result["missing_teams"],
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        # Un ricalcolo puo' spostare un trofeo gia' assegnato. Prima veniva
        # riscritto in silenzio (l'albo era derivato, quindi cambiava da solo e
        # nessuno lo vedeva accadere); ora si dice a chi ha premuto il pulsante.
        moved = honours.review_league(league)
        return Response({
            "fantasy_matchday_id": md.id,
            "recomputed_with": "snapshot" if not update_snapshot else "current",
            "fixtures_scored": result["updated"],
            "fixtures_total": len(fixtures),
            "prizes_changed": _describe_prize_changes(moved),
        })


class CompetitionDetailUpdateView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, competition_id: int):
        comp = get_object_or_404(FantasyCompetition, id=competition_id)
        _membership_or_404(comp.league, request.user.id)
        return Response(_serialize_competition(comp))

    @transaction.atomic
    def patch(self, request, competition_id: int):
        comp = get_object_or_404(FantasyCompetition, id=competition_id)
        _ensure_admin(comp.league, request.user.id)
        s = CompetitionUpdateSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        data = s.validated_data

        for field in [
            "name", "status", "points_win", "points_draw", "points_loss",
            "starts_at", "ends_at", "start_matchday", "end_matchday",
        ]:
            if field in data:
                setattr(comp, field, data[field])
        comp.save()

        # If a source competition advances/completes, refresh dependents.
        if "status" in data and data["status"] in [FantasyCompetition.STATUS_ACTIVE, FantasyCompetition.STATUS_DONE]:
            dependents = FantasyCompetition.objects.filter(qualification_rules__source_competition=comp).distinct()
            for dep in dependents:
                _resolve_rule_participants_and_regenerate(dep)
        return Response(_serialize_competition(comp))

    @transaction.atomic
    def delete(self, request, competition_id: int):
        comp = get_object_or_404(FantasyCompetition, id=competition_id)
        _ensure_admin(comp.league, request.user.id)

        ext_stage_rules = list(
            CompetitionStageRule.objects.filter(source_stage__competition=comp)
            .exclude(target_stage__competition=comp)
            .select_related("source_stage", "target_stage", "target_stage__competition")
            .order_by("target_stage__competition_id", "target_stage__order_index", "target_stage_id")
        )
        if ext_stage_rules:
            return Response(
                {
                    "detail": "Cannot delete competition: some stages are used as qualification sources by other competitions.",
                    "dependent_targets": [
                        {
                            "source_stage_id": r.source_stage_id,
                            "source_stage_name": r.source_stage.name,
                            "target_stage_id": r.target_stage_id,
                            "target_stage_name": r.target_stage.name,
                            "target_competition_id": r.target_stage.competition_id,
                            "target_competition_name": r.target_stage.competition.name,
                            "mode": r.mode,
                            "rank_from": r.rank_from,
                            "rank_to": r.rank_to,
                        }
                        for r in ext_stage_rules
                    ],
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        ext_qual_rules = list(
            CompetitionQualificationRule.objects.filter(source_competition=comp)
            .exclude(competition=comp)
            .select_related("competition")
        )
        if ext_qual_rules:
            return Response(
                {
                    "detail": "Cannot delete competition: it is referenced by competition-level qualification rules.",
                    "dependent_competitions": [
                        {
                            "competition_id": r.competition_id,
                            "competition_name": r.competition.name,
                            "mode": r.mode,
                            "source_stage": r.source_stage,
                            "rank_from": r.rank_from,
                            "rank_to": r.rank_to,
                        }
                        for r in ext_qual_rules
                    ],
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        ext_prizes = list(
            CompetitionPrize.objects.filter(source_stage__competition=comp)
            .exclude(competition=comp)
            .select_related("competition", "source_stage")
        )
        if ext_prizes:
            return Response(
                {
                    "detail": "Cannot delete competition: some of its stages are referenced by prizes in other competitions.",
                    "dependent_prizes": [
                        {
                            "prize_id": p.id,
                            "prize_name": p.name,
                            "competition_id": p.competition_id,
                            "competition_name": p.competition.name,
                            "source_stage_id": p.source_stage_id,
                            "source_stage_name": p.source_stage.name if p.source_stage_id else None,
                        }
                        for p in ext_prizes
                    ],
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        comp.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class CompetitionScheduleView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, competition_id: int):
        comp = get_object_or_404(FantasyCompetition, id=competition_id)
        _ensure_admin(comp.league, request.user.id)
        s = CompetitionScheduleSerializer(data=request.data or {})
        s.is_valid(raise_exception=True)
        data = s.validated_data

        changed = []
        if "starts_at" in data:
            comp.starts_at = data["starts_at"]
            changed.append("starts_at")
        if "ends_at" in data:
            comp.ends_at = data["ends_at"]
            changed.append("ends_at")
        if changed:
            comp.save(update_fields=changed)

        start_md = data.get("start_matchday")
        end_md = data.get("end_matchday")

        # Scheduling needs a real-matchday source: either the league reference
        # season (preferred) or the legacy date window.
        if comp.league.reference_season is None and (not comp.starts_at or not comp.ends_at):
            return Response(
                {"detail": "Set the league reference season (or competition dates) before scheduling."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        parsed_mapping: dict[int, int] = {}
        raw_mapping = data.get("round_mapping") or {}
        if isinstance(raw_mapping, dict):
            for raw_round, raw_matchday in raw_mapping.items():
                try:
                    rno = int(raw_round)
                    md = int(raw_matchday)
                except (TypeError, ValueError):
                    continue
                parsed_mapping[rno] = md
        result = competition_calendar.schedule(
            comp, round_mapping=parsed_mapping or None, start_md=start_md, end_md=end_md
        )
        return Response(result)


class CompetitionSchedulePreviewView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, competition_id: int):
        comp = get_object_or_404(FantasyCompetition, id=competition_id)
        _ensure_admin(comp.league, request.user.id)
        s = CompetitionSchedulePreviewSerializer(data=request.data or {})
        s.is_valid(raise_exception=True)
        data = s.validated_data
        return Response(
            competition_calendar.preview(
                comp, start_md=data.get("start_matchday"), end_md=data.get("end_matchday")
            )
        )


class CompetitionAddQualificationRuleView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, competition_id: int):
        comp = get_object_or_404(FantasyCompetition, id=competition_id)
        _ensure_admin(comp.league, request.user.id)
        s = QualificationRuleCreateSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        data = s.validated_data

        source = get_object_or_404(FantasyCompetition, id=data["source_competition_id"], league=comp.league)
        rule = CompetitionQualificationRule.objects.create(
            competition=comp,
            source_competition=source,
            source_stage=data["source_stage"],
            source_round=data.get("source_round"),
            mode=data["mode"],
            rank_from=data.get("rank_from"),
            rank_to=data.get("rank_to"),
        )
        _resolve_rule_participants_and_regenerate(comp)
        return Response(
            {
                "rule_id": rule.id,
                "competition_id": comp.id,
                "source_competition_id": source.id,
                "source_competition_name": source.name,
                "source_stage": rule.source_stage,
                "source_round": rule.source_round,
                "mode": rule.mode,
                "rank_from": rule.rank_from,
                "rank_to": rule.rank_to,
            },
            status=status.HTTP_201_CREATED,
        )


class CompetitionResolveDependenciesView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, competition_id: int):
        comp = get_object_or_404(FantasyCompetition, id=competition_id)
        _ensure_admin(comp.league, request.user.id)
        result = _resolve_rule_participants_and_regenerate(comp)
        return Response(result)


# ==========================================================================
# Auction room
# ==========================================================================
#
# The REST endpoints are the single write path (auth + transaction here); each
# mutation records an append-only AuctionEvent and then nudges the WebSocket room
# so every watcher re-fetches this exact state. The legality of every purchase is
# enforced by services/auction_engine.py, at bid time AND again at close time.

def _team_for_membership(m: LeagueMembership) -> FantasyTeam | None:
    return FantasyTeam.objects.filter(manager=m).first()


def _player_label(player: Player) -> str:
    return player.short_name or player.full_name


def _record_auction_event(session, event_type, actor, payload, nomination=None):
    return AuctionEvent.objects.create(
        session=session, nomination=nomination, event_type=event_type,
        actor=actor, payload=payload)


def _called_player_ids(session) -> set[int]:
    """Players no longer in the pool: currently up for auction (open) or already
    assigned (closed). A cancelled nomination returns its player to the pool."""
    return set(
        AuctionNomination.objects.filter(
            session=session,
            status__in=[AuctionNomination.STATUS_OPEN, AuctionNomination.STATUS_CLOSED],
        ).values_list("player_id", flat=True))


def _pool_remaining_ids(session) -> list[int]:
    called = _called_player_ids(session)
    return [pid for pid in (session.nomination_order or []) if pid not in called]


def _nominatable_ids(session) -> list[int]:
    """Everyone the admin may CALL BY NAME, which is more than the planned order.

    ``nomination_order`` is fixed when the auction is created and drives the random
    draws. Manual calls should not be bound by it: a player signed after the
    auction started is in the league's listone but not in that list, and a roster
    left short at the end has no way to be completed. What must stay refused is
    someone already owned, or the one currently up for auction.
    """
    owned = set(FantasyRosterSlot.objects
                .filter(team__league=session.league, released_at__isnull=True)
                .values_list("player_id", flat=True))
    open_ids = set(AuctionNomination.objects
                   .filter(session=session, status=AuctionNomination.STATUS_OPEN)
                   .values_list("player_id", flat=True))
    listone = (LeaguePlayerRole.objects
               .filter(league=session.league)
               .values_list("player_id", flat=True))
    planned = session.nomination_order or []
    order = {pid: i for i, pid in enumerate(planned)}
    # Those still to be drawn first, in the drawn order; then the extras.
    return sorted((pid for pid in listone if pid not in owned and pid not in open_ids),
                  key=lambda pid: (order.get(pid, len(planned)), pid))


def _open_nomination(session):
    return (AuctionNomination.objects
            .filter(session=session, status=AuctionNomination.STATUS_OPEN)
            .select_related("player", "nominator__user").first())


def _top_bid(nomination):
    return (nomination.bids.filter(is_void=False)
            .order_by("-amount", "created_at").first())


def _serialize_auction_state(session) -> dict:
    league = session.league
    budgets = team_budgets(league)
    teams = list(FantasyTeam.objects.filter(league=league).select_related("manager__user"))
    team_by_membership = {t.manager_id: t for t in teams}

    pool = _pool_remaining_ids(session)
    pool_roles = league_role_map(league, pool)
    remaining_by_role = {r: 0 for r in AUCTION_ROLES}
    for pid in pool:
        r = pool_roles.get(pid)
        if r in remaining_by_role:
            remaining_by_role[r] += 1

    # Open nomination detail, incl. per-team affordability for THIS player's role.
    open_nom = _open_nomination(session)
    open_payload = None
    if open_nom:
        role = player_role(league, open_nom.player)
        top = _top_bid(open_nom)
        top_amount = top.amount if top else 0
        min_next = top_amount + 1 if top else 1
        top_team = team_by_membership.get(top.bidder_id) if top else None
        bids = list(open_nom.bids.filter(is_void=False)
                    .select_related("bidder__user").order_by("-amount", "created_at")[:25])
        options = []
        for tb in budgets.values():
            mb = tb.max_bid_for_role(role) if role else 0
            options.append({
                "team_id": tb.team_id, "team_name": tb.team_name,
                "max_bid": mb, "eligible": mb >= min_next,
            })
        open_payload = {
            "nomination_id": open_nom.id,
            "player_id": open_nom.player_id,
            "player_name": _player_label(open_nom.player),
            "player_role": role,
            "call_mode": open_nom.call_mode,
            "nominator": open_nom.nominator.user.username,
            "top_bid": top_amount,
            "top_bidder_team_id": top_team.id if top_team else None,
            "top_bidder_team_name": top_team.name if top_team else None,
            "min_next_bid": min_next,
            "bids": [{
                "bid_id": b.id,
                "team_id": (team_by_membership.get(b.bidder_id).id
                            if team_by_membership.get(b.bidder_id) else None),
                "team_name": (team_by_membership.get(b.bidder_id).name
                              if team_by_membership.get(b.bidder_id) else None),
                "manager": b.bidder.user.username,
                "amount": b.amount,
            } for b in bids],
            "team_options": options,
        }

    # History of nominations (incl. cancelled) for the room log.
    nominations = list(
        AuctionNomination.objects.filter(session=session)
        .select_related("player", "nominator__user", "closed_winner_team")
        .order_by("-created_at")[:40])
    rows = [{
        "nomination_id": n.id, "status": n.status,
        "player_id": n.player_id, "player_name": _player_label(n.player),
        "call_mode": n.call_mode, "nominator": n.nominator.user.username,
        "winner_team_id": n.closed_winner_team_id,
        "winner_team_name": n.closed_winner_team.name if n.closed_winner_team_id else None,
        "winning_amount": n.winning_amount,
    } for n in nominations]

    events = list(AuctionEvent.objects.filter(session=session)
                  .select_related("actor")[:40])
    feed = [{
        "id": e.id, "type": e.event_type,
        "actor": e.actor.username if e.actor_id else None,
        "payload": e.payload, "created_at": e.created_at.isoformat(),
    } for e in events]

    team_budgets_out = [{
        "team_id": tb.team_id, "team_name": tb.team_name,
        "manager_username": tb.manager_username,
        "initial_budget": tb.initial_budget,
        "spent_budget": tb.spent, "available_budget": tb.remaining,
        "slots": tb.slots, "slots_remaining_total": tb.slots_remaining_total,
        "max_bid_any": tb.max_bid_any,
    } for tb in budgets.values()]

    return {
        "auction_id": session.id, "name": session.name, "status": session.status,
        "league_id": league.id,
        "roster_slots": league.roster_quota(),
        "initial_budget": league.initial_budget,
        "pool_total": len(session.nomination_order or []),
        "pool_remaining": len(pool),
        "remaining_by_role": remaining_by_role,
        "open_nomination": open_payload,
        "recent_nominations": rows,
        "events": feed,
        "team_budgets": team_budgets_out,
    }


# --- Mutating helpers (shared by explicit endpoints and undo-last) ---------

def _void_bid(bid, actor):
    if bid.is_void:
        raise ValueError("Offerta gia' annullata.")
    nom = bid.nomination
    bid.is_void = True
    bid.save(update_fields=["is_void"])
    _record_auction_event(
        nom.session, AuctionEvent.TYPE_BID_VOIDED, actor,
        {"bid_id": bid.id, "amount": bid.amount, "player_name": _player_label(nom.player)},
        nomination=nom)
    return {"bid_id": bid.id}


def _cancel_nomination(nom, actor):
    if nom.status != AuctionNomination.STATUS_OPEN:
        raise ValueError("La chiamata non e' aperta.")
    nom.bids.filter(is_void=False).update(is_void=True)
    nom.status = AuctionNomination.STATUS_CANCELLED
    nom.save(update_fields=["status"])
    _record_auction_event(
        nom.session, AuctionEvent.TYPE_NOMINATION_CANCELLED, actor,
        {"player_name": _player_label(nom.player)}, nomination=nom)
    return {"nomination_id": nom.id}


def _revert_assignment(nom, actor):
    if nom.status != AuctionNomination.STATUS_CLOSED:
        raise ValueError("La chiamata non e' assegnata.")
    team_name = nom.closed_winner_team.name if nom.closed_winner_team_id else None
    amount = nom.winning_amount
    if nom.roster_slot_id:
        FantasyRosterSlot.objects.filter(id=nom.roster_slot_id).delete()
    # Reopen so the player can be re-auctioned; keep his (now un-void) bids so the
    # room can simply re-close, or the admin can cancel/re-assign.
    nom.bids.update(is_void=False)
    nom.roster_slot = None
    nom.closed_winner_team = None
    nom.winning_amount = None
    nom.status = AuctionNomination.STATUS_OPEN
    nom.save(update_fields=["roster_slot", "closed_winner_team", "winning_amount", "status"])
    _record_auction_event(
        nom.session, AuctionEvent.TYPE_ASSIGNMENT_REVERTED, actor,
        {"player_name": _player_label(nom.player), "team_name": team_name, "amount": amount},
        nomination=nom)
    return {"nomination_id": nom.id}


class LeagueActiveAuctionView(APIView):
    """Discover the league's live auction (if any) — used by the room entry point."""

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        m = _membership_or_404(league, request.user.id)
        session = (AuctionSession.objects
                   .filter(league=league, status=AuctionSession.STATUS_ACTIVE)
                   .order_by("-created_at").first())
        return Response({
            "auction_id": session.id if session else None,
            "status": session.status if session else None,
            "is_admin": m.role == LeagueMembership.ROLE_ADMIN,
            "mode": league.mode,
        })


class AuctionCreateView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _ensure_admin(league, request.user.id)

        # The auction economy (budget/slots/roles) is a classic-mode concept; aura
        # is deliberately not finalised for the auction yet.
        if league.mode != FantasyLeague.MODE_CLASSIC:
            return Response(
                {"detail": "L'asta e' disponibile solo per le leghe in modalita' classic."},
                status=status.HTTP_400_BAD_REQUEST)

        existing = AuctionSession.objects.filter(
            league=league, status=AuctionSession.STATUS_ACTIVE).first()
        if existing:
            return Response(
                {"auction_id": existing.id, "detail": "Un'asta e' gia' in corso."},
                status=status.HTTP_200_OK)

        s = CreateAuctionSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        data = s.validated_data

        # Pool = the players passed in, or the whole frozen listone by default.
        if data.get("player_ids"):
            player_ids = data["player_ids"]
        else:
            player_ids = list(LeaguePlayerRole.objects.filter(league=league)
                              .values_list("player_id", flat=True))
        if not player_ids:
            return Response(
                {"detail": "Listone vuoto: congela prima il listone della lega."},
                status=status.HTTP_400_BAD_REQUEST)

        blocked = _ensure_players_decided(league, player_ids)
        if blocked:
            return blocked
        players = list(Player.objects.filter(id__in=player_ids).values_list("id", flat=True))

        session = AuctionSession.objects.create(
            league=league, name=data["name"], status=AuctionSession.STATUS_ACTIVE,
            nomination_order=players, nomination_index=0, created_by=request.user)
        _record_auction_event(session, AuctionEvent.TYPE_SESSION_CREATED, request.user,
                              {"name": session.name, "pool": len(players)})
        broadcast_auction(session.id)
        return Response({"auction_id": session.id, "players": len(players)},
                        status=status.HTTP_201_CREATED)


class AuctionStateView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, auction_id: int):
        session = get_object_or_404(AuctionSession, id=auction_id)
        _membership_or_404(session.league, request.user.id)
        return Response(_serialize_auction_state(session))


class AuctionPoolView(APIView):
    """Everyone still callable in this auction, so the room can search offline.

    The banditore's search used to hit /players/search on every keystroke, which
    was both a request per letter and WRONG: that endpoint looks across every
    Player in the database (1706 here) while only the league's frozen listone can
    actually be nominated (660) — so it offered names that would be refused.

    Kept OUT of the auction state on purpose: the socket only says "something
    changed" and the client re-fetches the state on every single bid. Carrying the
    pool there would re-download it each time. It changes only when a nomination
    is settled, which `pool_remaining` in the state already signals.
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, auction_id: int):
        session = get_object_or_404(AuctionSession, id=auction_id)
        _membership_or_404(session.league, request.user.id)

        # Everyone CALLABLE, not just those left in the drawn order: the manual
        # call is the way to finish a short roster once the order has run out.
        callable_ids = _nominatable_ids(session)
        drawn = set(_pool_remaining_ids(session))
        roles = league_role_map(session.league, callable_ids)
        players = Player.objects.filter(id__in=callable_ids).values_list(
            "id", "short_name", "full_name")
        by_id = {pid: (short, full) for pid, short, full in players}
        return Response([
            {
                "player_id": pid,
                "name": by_id[pid][0] or by_id[pid][1] or str(pid),
                "full_name": by_id[pid][1] or "",
                "role": roles.get(pid),
                # False => outside the draw order (added to the listone later, or
                # already gone round once). Callable, but the UI can say so.
                "in_draw_order": pid in drawn,
            }
            for pid in callable_ids
            if pid in by_id
        ])


class AuctionNominateView(APIView):
    """Put a player up for auction, in one of three admin-chosen ways: pick him
    manually, draw one at random, or draw one at random within a role."""

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, auction_id: int):
        session = get_object_or_404(AuctionSession, id=auction_id)
        league = session.league
        m = _ensure_admin(league, request.user.id)

        if session.status != AuctionSession.STATUS_ACTIVE:
            return Response({"detail": "L'asta non e' attiva."}, status=status.HTTP_400_BAD_REQUEST)
        if _open_nomination(session):
            return Response({"detail": "C'e' gia' un giocatore in chiamata: chiudilo o annullalo prima."},
                            status=status.HTTP_400_BAD_REQUEST)

        s = NominateSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        data = s.validated_data
        mode = data["mode"]

        pool = _pool_remaining_ids(session)
        # An exhausted draw order ends the auction — but only for a DRAW. Calling
        # by name is exactly what is left to do when the order has run out and a
        # roster is still short, so it must not be met with "tutti i giocatori
        # sono stati chiamati" and a closed session.
        if not pool and mode != "manual":
            session.status = AuctionSession.STATUS_CLOSED
            session.save(update_fields=["status"])
            _record_auction_event(session, AuctionEvent.TYPE_SESSION_CLOSED, request.user,
                                  {"reason": "pool_empty"})
            broadcast_auction(session.id)
            return Response({"detail": "Tutti i giocatori sono stati chiamati."},
                            status=status.HTTP_200_OK)

        if mode == "manual":
            player_id = data["player_id"]
            # Deliberately NOT `pool`: a call by name may reach outside the drawn
            # order (see _nominatable_ids). Only owned players and the one on the
            # block are refused.
            if player_id not in _nominatable_ids(session):
                return Response({"detail": "Giocatore non disponibile (gia' in una rosa o gia' in chiamata)."},
                                status=status.HTTP_400_BAD_REQUEST)
            call_mode = AuctionNomination.CALL_MANUAL
        else:
            candidates = pool
            if mode == "random_role":
                role = data["role"]
                roles = league_role_map(league, pool)
                candidates = [pid for pid in pool if roles.get(pid) == role]
                if not candidates:
                    return Response({"detail": f"Nessun giocatore disponibile per il ruolo {role}."},
                                    status=status.HTTP_400_BAD_REQUEST)
                call_mode = AuctionNomination.CALL_RANDOM_ROLE
            else:
                call_mode = AuctionNomination.CALL_RANDOM
            seed = data.get("random_seed")
            rng = Random(seed) if seed is not None else Random()
            player_id = rng.choice(candidates)

        player = get_object_or_404(Player, id=player_id)
        nom = AuctionNomination.objects.create(
            session=session, player=player, nominator=m, call_mode=call_mode)
        _record_auction_event(
            session, AuctionEvent.TYPE_NOMINATED, request.user,
            {"player_name": _player_label(player), "player_id": player.id,
             "call_mode": call_mode, "role": player_role(league, player)},
            nomination=nom)
        broadcast_auction(session.id)
        return Response({
            "nomination_id": nom.id, "player_id": player.id,
            "player_name": _player_label(player), "call_mode": call_mode,
        }, status=status.HTTP_201_CREATED)


class AuctionPlaceBidView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, nomination_id: int):
        nomination = get_object_or_404(
            AuctionNomination.objects.select_related("player", "session__league"),
            id=nomination_id)
        if nomination.status != AuctionNomination.STATUS_OPEN:
            return Response({"detail": "La chiamata e' chiusa."}, status=status.HTTP_400_BAD_REQUEST)

        league = nomination.session.league
        m = _membership_or_404(league, request.user.id)
        is_admin = m.role == LeagueMembership.ROLE_ADMIN

        s = PlaceBidSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        amount = s.validated_data["amount"]

        # Bidder team: normally the caller's own; an admin may bid for another team
        # (verbal auctions where managers call out their raises).
        team_id = s.validated_data.get("team_id")
        if team_id and is_admin:
            team = get_object_or_404(FantasyTeam, id=team_id, league=league)
            bidder = team.manager
        else:
            team = _team_for_membership(m)
            bidder = m
            if team is None:
                return Response({"detail": "Non hai una squadra in questa lega."},
                                status=status.HTTP_400_BAD_REQUEST)

        top = _top_bid(nomination)
        min_required = (top.amount + 1) if top else 1
        if amount < min_required:
            return Response({"detail": f"L'offerta deve essere almeno {min_required}."},
                            status=status.HTTP_400_BAD_REQUEST)

        # Legality: role slot free + at least 1 credit reserved per other open slot.
        role = player_role(league, nomination.player)
        legality = check_purchase(league, team.id, role, amount)
        if not legality.ok:
            return Response({"detail": legality.reason, "max_bid": legality.max_bid},
                            status=status.HTTP_400_BAD_REQUEST)

        bid = AuctionBid.objects.create(nomination=nomination, bidder=bidder, amount=amount)
        _record_auction_event(
            nomination.session, AuctionEvent.TYPE_BID, request.user,
            {"bid_id": bid.id, "player_name": _player_label(nomination.player),
             "team_name": team.name, "team_id": team.id, "amount": amount,
             "by_admin": is_admin and bool(team_id)},
            nomination=nomination)
        broadcast_auction(nomination.session_id)
        return Response({"bid_id": bid.id, "amount": bid.amount}, status=status.HTTP_201_CREATED)


class AuctionCloseNominationView(APIView):
    """Assign the open player to the best bid (or, with no bid, refuse — use cancel)."""

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, nomination_id: int):
        nomination = get_object_or_404(
            AuctionNomination.objects.select_related("player", "session__league"),
            id=nomination_id)
        league = nomination.session.league
        _ensure_admin(league, request.user.id)

        if nomination.status == AuctionNomination.STATUS_CLOSED:
            return Response({"detail": "Gia' assegnato."}, status=status.HTTP_200_OK)
        if nomination.status != AuctionNomination.STATUS_OPEN:
            return Response({"detail": "La chiamata non e' aperta."}, status=status.HTTP_400_BAD_REQUEST)

        top = _top_bid(nomination)
        if not top:
            return Response(
                {"detail": "Nessuna offerta: usa 'Annulla chiamata' per rimettere il giocatore in lista."},
                status=status.HTTP_400_BAD_REQUEST)

        winner_team = _team_for_membership(top.bidder)
        if winner_team is None:
            return Response({"detail": "La squadra vincente non esiste piu'."},
                            status=status.HTTP_400_BAD_REQUEST)

        role = player_role(league, nomination.player)
        legality = check_purchase(league, winner_team.id, role, top.amount)
        if not legality.ok:
            return Response({"detail": f"Assegnazione non valida: {legality.reason}"},
                            status=status.HTTP_400_BAD_REQUEST)

        slot = FantasyRosterSlot.objects.create(
            team=winner_team, player=nomination.player, purchase_price=top.amount)
        nomination.status = AuctionNomination.STATUS_CLOSED
        nomination.closed_winner_team = winner_team
        nomination.winning_amount = top.amount
        nomination.roster_slot = slot
        nomination.save(update_fields=["status", "closed_winner_team", "winning_amount", "roster_slot"])
        _record_auction_event(
            nomination.session, AuctionEvent.TYPE_ASSIGNED, request.user,
            {"player_name": _player_label(nomination.player), "team_name": winner_team.name,
             "team_id": winner_team.id, "amount": top.amount, "via": "bid"},
            nomination=nomination)
        broadcast_auction(nomination.session_id)
        return Response({"nomination_id": nomination.id, "winner_team_id": winner_team.id,
                         "amount": top.amount}, status=status.HTTP_200_OK)


class AuctionAssignView(APIView):
    """Admin direct-assign shortcut: hand a player to a team at a fixed price, no
    bidding. For in-person / verbal auctions. Still fully legality-checked."""

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, auction_id: int):
        session = get_object_or_404(AuctionSession, id=auction_id)
        league = session.league
        m = _ensure_admin(league, request.user.id)

        if session.status != AuctionSession.STATUS_ACTIVE:
            return Response({"detail": "L'asta non e' attiva."}, status=status.HTTP_400_BAD_REQUEST)

        s = AuctionAssignSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        data = s.validated_data
        player = get_object_or_404(Player, id=data["player_id"])
        team = get_object_or_404(FantasyTeam, id=data["team_id"], league=league)
        price = data["price"]

        # Player must be available: either currently up for auction, or still in pool.
        open_for_player = AuctionNomination.objects.filter(
            session=session, player=player, status=AuctionNomination.STATUS_OPEN).first()
        if open_for_player is None and player.id not in _pool_remaining_ids(session):
            return Response({"detail": "Giocatore non disponibile (gia' assegnato o fuori dal listone)."},
                            status=status.HTTP_400_BAD_REQUEST)

        role = player_role(league, player)
        legality = check_purchase(league, team.id, role, price)
        if not legality.ok:
            return Response({"detail": legality.reason, "max_bid": legality.max_bid},
                            status=status.HTTP_400_BAD_REQUEST)

        slot = FantasyRosterSlot.objects.create(team=team, player=player, purchase_price=price)
        if open_for_player is not None:
            open_for_player.bids.filter(is_void=False).update(is_void=True)
            nom = open_for_player
            nom.call_mode = AuctionNomination.CALL_ASSIGN
        else:
            nom = AuctionNomination(session=session, player=player, nominator=m,
                                    call_mode=AuctionNomination.CALL_ASSIGN)
        nom.status = AuctionNomination.STATUS_CLOSED
        nom.closed_winner_team = team
        nom.winning_amount = price
        nom.roster_slot = slot
        nom.save()
        _record_auction_event(
            session, AuctionEvent.TYPE_ASSIGNED, request.user,
            {"player_name": _player_label(player), "team_name": team.name,
             "team_id": team.id, "amount": price, "via": "assign"},
            nomination=nom)
        broadcast_auction(session.id)
        return Response({"nomination_id": nom.id, "winner_team_id": team.id, "amount": price},
                        status=status.HTTP_201_CREATED)


class AuctionCancelNominationView(APIView):
    """Withdraw the open player without assigning — he returns to the pool (undo a
    wrong nomination)."""

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, nomination_id: int):
        nomination = get_object_or_404(
            AuctionNomination.objects.select_related("player", "session__league"),
            id=nomination_id)
        _ensure_admin(nomination.session.league, request.user.id)
        try:
            _cancel_nomination(nomination, request.user)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        broadcast_auction(nomination.session_id)
        return Response({"nomination_id": nomination.id, "status": nomination.status})


class AuctionRevertNominationView(APIView):
    """Undo a completed purchase: refund the credits, free the slot, reopen the
    player for auction."""

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, nomination_id: int):
        nomination = get_object_or_404(
            AuctionNomination.objects.select_related("player", "session__league"),
            id=nomination_id)
        _ensure_admin(nomination.session.league, request.user.id)
        try:
            _revert_assignment(nomination, request.user)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        broadcast_auction(nomination.session_id)
        return Response({"nomination_id": nomination.id, "status": nomination.status})


class AuctionVoidBidView(APIView):
    """Undo a mistaken bid (admin)."""

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, bid_id: int):
        bid = get_object_or_404(
            AuctionBid.objects.select_related("nomination__session__league"), id=bid_id)
        _ensure_admin(bid.nomination.session.league, request.user.id)
        try:
            _void_bid(bid, request.user)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        broadcast_auction(bid.nomination.session_id)
        return Response({"bid_id": bid.id, "is_void": bid.is_void})


class AuctionUndoLastView(APIView):
    """Undo the last state-changing action in the room (bid / nomination / purchase)."""

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, auction_id: int):
        session = get_object_or_404(AuctionSession, id=auction_id)
        _ensure_admin(session.league, request.user.id)

        # Walk back over events, skipping ones that are themselves undos or that
        # refer to state already reverted, to find the last undoable action.
        UNDOABLE = {
            AuctionEvent.TYPE_BID, AuctionEvent.TYPE_NOMINATED, AuctionEvent.TYPE_ASSIGNED,
        }
        for event in AuctionEvent.objects.filter(session=session)[:50]:
            if event.event_type not in UNDOABLE:
                continue
            try:
                if event.event_type == AuctionEvent.TYPE_BID:
                    bid = AuctionBid.objects.filter(
                        id=event.payload.get("bid_id"), is_void=False).first()
                    if not bid:
                        continue
                    _void_bid(bid, request.user)
                    result = {"undone": "bid", "bid_id": bid.id}
                elif event.event_type == AuctionEvent.TYPE_NOMINATED:
                    if not event.nomination_id:
                        continue
                    nom = AuctionNomination.objects.get(id=event.nomination_id)
                    if nom.status != AuctionNomination.STATUS_OPEN:
                        continue
                    _cancel_nomination(nom, request.user)
                    result = {"undone": "nomination", "nomination_id": nom.id}
                else:  # ASSIGNED
                    if not event.nomination_id:
                        continue
                    nom = AuctionNomination.objects.get(id=event.nomination_id)
                    if nom.status != AuctionNomination.STATUS_CLOSED:
                        continue
                    _revert_assignment(nom, request.user)
                    result = {"undone": "assignment", "nomination_id": nom.id}
            except ValueError:
                continue
            broadcast_auction(session.id)
            return Response(result, status=status.HTTP_200_OK)

        return Response({"detail": "Nessuna azione da annullare."}, status=status.HTTP_400_BAD_REQUEST)


class AuctionCloseSessionView(APIView):
    """End the auction (admin). Any open nomination is cancelled first."""

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request, auction_id: int):
        session = get_object_or_404(AuctionSession, id=auction_id)
        _ensure_admin(session.league, request.user.id)
        open_nom = _open_nomination(session)
        if open_nom:
            _cancel_nomination(open_nom, request.user)
        session.status = AuctionSession.STATUS_CLOSED
        session.save(update_fields=["status"])
        _record_auction_event(session, AuctionEvent.TYPE_SESSION_CLOSED, request.user,
                              {"reason": "manual"})
        broadcast_auction(session.id)
        return Response({"auction_id": session.id, "status": session.status})


def _compute_standings(fixtures, pw: int, pd: int, pl: int,
                       live_totals: dict | None = None) -> list[dict]:
    """Ranked standings (fixtures with .detail prefetched).

    LE PARTITE IN CORSO CONTANO, quando ``live_totals`` le porta. Una classifica
    che si fermava all'ultima giornata conclusa dava, per tutta la domenica, la
    risposta sbagliata alla sola domanda che le si fa in quel momento — "come sta
    andando" — e la dava accanto a un calendario che i punteggi provvisori li
    mostrava già: due risposte alla stessa domanda, e nessuna delle due marcata.
    Le righe toccate escono con ``provisional`` acceso, perché "sesto a 41" e
    "sesto a 41 con una partita da finire" non sono la stessa frase.

    Senza ``live_totals`` il conto è quello di prima, sulle sole partite chiuse:
    è ciò che serve a chi la classifica la vuole definitiva (un premio, un
    verdetto), e resta il comportamento predefinito.
    """
    rows: dict[int, dict] = {}
    names: dict[int, str] = {}
    crests: dict[int, str] = {}

    def row(team_id: int) -> dict:
        return rows.setdefault(
            team_id, {"pts": 0, "played": 0, "w": 0, "d": 0, "l": 0, "gf": 0.0, "ga": 0.0,
                      "score_sum": 0.0, "scored": 0, "live": False}
        )

    for fx in fixtures:
        played = fx.status == FantasyFixture.STATUS_FINISHED
        live = None if played else (live_totals or {}).get(fx.id)
        if not played and live is None:
            continue
        hs, as_ = ((fx.home_total, fx.away_total) if played
                   else (live["home_total"], live["away_total"]))
        names[fx.home_team_id] = fx.home_team.name
        names[fx.away_team_id] = fx.away_team.name
        crests[fx.home_team_id] = fx.home_team.crest
        crests[fx.away_team_id] = fx.away_team.crest
        h, a = row(fx.home_team_id), row(fx.away_team_id)
        h["played"] += 1
        a["played"] += 1
        h["gf"] += hs
        h["ga"] += as_
        a["gf"] += as_
        a["ga"] += hs
        if not played:
            h["live"] = a["live"] = True
        if hs > as_:
            h["pts"] += pw; a["pts"] += pl; h["w"] += 1; a["l"] += 1
        elif hs < as_:
            a["pts"] += pw; h["pts"] += pl; a["w"] += 1; h["l"] += 1
        else:
            h["pts"] += pd; a["pts"] += pd; h["d"] += 1; a["d"] += 1
        detail = getattr(fx, "detail", None)
        if detail is not None:
            h["score_sum"] += detail.vfoot_home
            a["score_sum"] += detail.vfoot_away
            h["scored"] += 1
            a["scored"] += 1

    ranked = sorted(rows.items(), key=lambda kv: (kv[1]["pts"], kv[1]["gf"] - kv[1]["ga"], kv[1]["gf"]), reverse=True)
    return [
        {
            "rank": i + 1, "team_id": tid, "team": names.get(tid, str(tid)),
            "crest": crests.get(tid, ""),
            "played": r["played"], "wins": r["w"], "draws": r["d"], "losses": r["l"],
            "goals_for": int(r["gf"]), "goals_against": int(r["ga"]),
            "goal_diff": int(r["gf"] - r["ga"]), "points": r["pts"],
            # La media sulle partite che un tabellino ce l'hanno: una in corso non
            # ne ha ancora uno, e dividerla per le giocate abbassava la media di
            # tutti appena cominciava la giornata.
            "avg_score_for": round(r["score_sum"] / r["scored"], 3) if r["scored"] else 0.0,
            # Questa riga contiene una partita ancora da finire.
            "provisional": r["live"],
        }
        for i, (tid, r) in enumerate(ranked)
    ]


_KO_ROUND_LABELS = {1: "Finale", 2: "Semifinali", 4: "Quarti di finale", 8: "Ottavi di finale"}


def _highlighted_ranks(stage) -> tuple[list[int], list[int]]:
    """Table positions that lead somewhere: (prize places, qualifying places).

    Kept apart because they are not the same promise. A prize is won there and
    then; qualifying only means you carry on. Merging them painted the same green
    on the first four of a championship whose prizes go to the first three, so the
    fourth looked like it won something.

    Both come from the competition's own data — prize bands and table-range
    qualification rules — not from a number borrowed from real football.
    """
    if stage is None:
        return [], []

    prize_ranks: set[int] = set()
    for prize in stage.competition.prizes.all():
        if prize.condition_type == CompetitionPrize.CONDITION_STAGE_TABLE_RANGE:
            if prize.source_stage_id != stage.id:
                continue
        elif prize.condition_type == CompetitionPrize.CONDITION_FINAL_TABLE_RANGE:
            # The final table is the last stage's; a group table is not it.
            if stage.competition.stages.order_by("-order_index", "-id").first() != stage:
                continue
        else:
            continue
        lo = prize.rank_from or 1
        hi = prize.rank_to or lo
        prize_ranks.update(range(lo, hi + 1))

    qualify_ranks: set[int] = set()
    for rule in stage.rules_out.filter(mode=CompetitionStageRule.MODE_TABLE_RANGE):
        lo = rule.rank_from or 1
        hi = rule.rank_to or lo
        qualify_ranks.update(range(lo, hi + 1))

    # A place that wins something is not also merely "through".
    return sorted(prize_ranks), sorted(qualify_ranks - prize_ranks)


def _section(name, stage_type, order, fixtures, my_team_id, current_md, pw, pd, pl,
             stage=None, awaiting_mds=None, locked_mds=None, live_totals=None) -> dict:
    """One results section: a standings table (round-robin) or a bracket (knockout)."""
    fixtures = list(fixtures)
    prize_ranks, qualify_ranks = _highlighted_ranks(stage)
    base = {"name": name, "type": stage_type, "order": order,
            "prize_ranks": prize_ranks, "qualify_ranks": qualify_ranks}
    if stage_type == CompetitionStage.TYPE_KNOCKOUT:
        by_round: dict[int, list] = {}
        for f in fixtures:
            by_round.setdefault(f.round_no, []).append(f)
        rounds = []
        for rno in sorted(by_round):
            fs = by_round[rno]
            # Chi è passato, e perché. Su un tabellone il risultato non basta a
            # dirlo: un 1-1 da solo non dice niente, e prima il turno seguente
            # compariva con dentro una squadra che a schermo non aveva vinto
            # nulla. Un turno alla volta — dentro un turno una squadra gioca una
            # sfida sola, quindi il vincitore la identifica.
            passed = {t.winner_id: t for t in knockout.tie_outcomes(fs)}
            rows = []
            for f in fs:
                row = _serialize_fixture_row(f, my_team_id, current_md, False,
                                             awaiting_mds, locked_mds, live_totals)
                winner = next((tid for tid in (f.home_team_id, f.away_team_id)
                               if tid in passed), None)
                row["advanced_team_id"] = winner
                # Solo quando il risultato da solo non lo spiega: "passa perché ha
                # segnato di più" non è una notizia.
                row["advanced_reason"] = (
                    passed[winner].reason
                    if winner is not None and passed[winner].reason != knockout.BY_GOALS
                    else None
                )
                # La serie, tiro per tiro. Solo dove c'e' stata: dire "rigori" e
                # non mostrarli sarebbe la parte peggiore di entrambe le scelte.
                row["shootout"] = f.shootout or None
                rows.append(row)
            rounds.append({
                "round_no": rno,
                "label": _KO_ROUND_LABELS.get(len(fs), f"Turno {rno}"),
                "fixtures": rows,
            })
        base["rounds"] = rounds
    else:
        # Tutte le partite, non solo le finite: quelle in corso entrano in
        # classifica se ``live_totals`` ha un punteggio per loro (vedi
        # _compute_standings), e senza cambiano niente.
        standings = _compute_standings(fixtures, pw, pd, pl, live_totals)
        # Before the first match there are no finished fixtures, so the table came
        # back empty and a group showed its name over nothing — for a group stage
        # that reads as "the draw failed", when in fact nobody has played yet.
        # Every team that HAS a fixture in this section belongs in the table, at
        # zero.
        if not standings:
            seen: dict[int, tuple[str, str]] = {}
            for f in fixtures:
                seen.setdefault(f.home_team_id, (f.home_team.name, f.home_team.crest))
                seen.setdefault(f.away_team_id, (f.away_team.name, f.away_team.crest))
            standings = [
                {
                    "rank": i + 1, "team_id": tid, "team": name, "crest": crest,
                    "played": 0, "wins": 0, "draws": 0, "losses": 0,
                    "goals_for": 0, "goals_against": 0, "goal_diff": 0, "points": 0,
                    "avg_score_for": 0.0, "provisional": False,
                }
                # Alphabetical: with everyone on zero any other order would suggest
                # a ranking that does not exist yet.
                for i, (tid, (name, crest)) in enumerate(
                    sorted(seen.items(), key=lambda kv: kv[1][0].lower()))
            ]
        base["standings"] = standings
    return base


class CompetitionStructureView(APIView):
    """Stage-aware results for ONE competition: an ordered list of SECTIONS, each a
    standings table (round-robin) or a bracket (knockout). A flat competition (no
    stages) yields a single section from its own type. Handles group+KO cups."""

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, league_id: int, competition_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        membership = _membership_or_404(league, request.user.id)
        my_team_id = membership.team.id if hasattr(membership, "team") else None
        comp = get_object_or_404(FantasyCompetition, id=competition_id, league=league)
        current = _current_matchday(league)
        current_md = current.real_matchday if current else None
        pw, pd, pl = comp.points_win, comp.points_draw, comp.points_loss
        rel = ("competition", "stage", "fantasy_matchday", "home_team", "away_team", "detail")
        # Il tabellino serve per due numeri (i punteggi di squadra), ma si porta
        # dietro `payload`: l'intero referto, venticinque giocatori con tutte le
        # loro statistiche, in JSON. Caricarlo per ogni partita significava
        # deserializzare centinaia di referti per disegnare una classifica —
        # l'80% del tempo di questa pagina se ne andava in `json.loads`, e non
        # per calcolare qualcosa, solo per leggere righe che nessuno guardava.
        defer = ("detail__payload",)

        awaiting_mds = {md.real_matchday for md in matchday_state.awaiting_matchdays(league)}
        locked_mds = (matchday_state.locked_matchdays(league.reference_season_id)
                      if league.reference_season_id else set())
        # I punteggi provvisori della competizione, in un colpo solo: `_live_totals`
        # tiene un solo scorer per giornata, e chiederli fase per fase avrebbe
        # rifatto quel lavoro per ogni girone.
        live_totals = _live_totals(
            league, list(comp.fixtures.select_related("fantasy_matchday")), locked_mds)

        stages = list(comp.stages.order_by("order_index", "id"))
        if stages:
            sections = [
                _section(s.name, s.stage_type, s.order_index,
                         s.fixtures.select_related(*rel).defer(*defer), my_team_id, current_md, pw, pd, pl,
                         stage=s, awaiting_mds=awaiting_mds, locked_mds=locked_mds,
                         live_totals=live_totals)
                for s in stages
            ]
        else:
            sections = [
                _section(comp.name, comp.competition_type, 1,
                         comp.fixtures.select_related(*rel).defer(*defer), my_team_id, current_md, pw, pd, pl,
                         awaiting_mds=awaiting_mds, locked_mds=locked_mds,
                         live_totals=live_totals)
            ]
        return Response({
            "competition_id": comp.id, "name": comp.name,
            "result_view": _result_view(comp), "sections": sections,
        })


class LeagueStandingsView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _membership_or_404(league, request.user.id)
        # A standings table is COMPETITION-scoped, not league-scoped. Use the given
        # competition; without one, the league's principal competition — the
        # round-robin with the most matches (services/league_competitions), not
        # whichever happened to be created first. The chosen id goes back in the
        # response, so a caller that cares can say which table it is showing.
        comp_param = request.query_params.get("competition_id")
        comp = None
        if comp_param:
            comp = league.competitions.filter(id=comp_param).first()
        if comp is None:
            comp = main_competition(league)
        pw, pd, pl = (comp.points_win, comp.points_draw, comp.points_loss) if comp else (3, 1, 0)

        # LA STESSA FUNZIONE della vista struttura, non una seconda copia. Erano
        # due conti identici scritti a mano in due punti, con un commento che
        # chiedeva di cambiarli insieme: la prima colonna aggiunta a uno solo dei
        # due sarebbe stata il bug che quel commento temeva, e i punteggi
        # provvisori sarebbero stati esattamente quella colonna.
        fixtures = list(
            FantasyFixture.objects
            .filter(competition=comp)
            .select_related("fantasy_matchday", "home_team", "away_team", "detail")
            # Stessa ragione della vista struttura: della `detail` servono due
            # numeri, non il referto intero.
            .defer("detail__payload")
        ) if comp else []
        locked_mds = (matchday_state.locked_matchdays(league.reference_season_id)
                      if league.reference_season_id else set())
        standings = _compute_standings(
            fixtures, pw, pd, pl, _live_totals(league, fixtures, locked_mds))
        return Response({"competition_id": comp.id if comp else None, "standings": standings})


class FixtureDetailView(APIView):
    """The tabellino of a league fixture — frozen if the matchday is concluded,
    computed on the spot if it is still being played.

    The two are the same numbers by construction: the live branch runs the very
    functions the conclusion runs, in the same order. What it does NOT do is
    persist. ``FantasyFixtureDetail`` is born at the conclusion and nowhere else,
    so reopening a closed matchday stays pure reading.
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, fixture_id: int):
        fx = get_object_or_404(
            FantasyFixture.objects.select_related(
                "competition__league", "detail", "fantasy_matchday",
                "home_team", "away_team"),
            id=fixture_id,
        )
        league = fx.competition.league
        _membership_or_404(league, request.user.id)
        detail = getattr(fx, "detail", None)
        if detail is not None:
            return Response(detail.payload)

        md = fx.fantasy_matchday
        if md is None or md.status == FantasyMatchday.STATUS_CONCLUDED:
            # Concluded without a payload is the genuinely empty case (a legacy
            # matchday, or one scored before details were kept).
            return Response({"detail": "No rich detail for this fixture."},
                            status=status.HTTP_404_NOT_FOUND)

        from vfoot.services.classic_matchday_scoring import score_fixture_live
        from vfoot.services.classic_scoring import Ruleset

        # The snapshot if the round already has one (it is what the conclusion will
        # use), otherwise the league's current rules.
        ruleset = (Ruleset.from_snapshot(md.ruleset_snapshot) if md.ruleset_snapshot
                   else Ruleset.from_league(league))
        return Response(score_fixture_live(fx, league, md, ruleset))


def _zone_grid_keys(cols: int = 5, rows: int = 4) -> list[str]:
    return [f"Z_{c}_{r}" for c in range(cols) for r in range(rows)]


@_lru_cache(maxsize=1)
def _vector_calibration() -> dict:
    path = _os.path.join(_os.path.dirname(str(_settings.BASE_DIR)), "calibration/vector_zone_duel_v1.json")
    try:
        return load_calibration(path)
    except Exception:
        return {"params": {}, "feature_scales": {}}


class LeagueTeamLineupView(APIView):
    """Real lineup context for a team: its roster with spatial profiles
    (role/footprint/minutes), the league matchdays, and — for the caller's OWN
    team — the saved lineup for the chosen matchday.

    With ``?team_id=`` any league member can read ANOTHER participant's structured
    roster (the same view the Squad page renders), so squads are no longer only
    visible in the flat name+price list. The saved lineup is withheld for other
    people's teams: the roster is public within the league, the chosen XI is not."""

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        membership = _membership_or_404(league, request.user.id)
        own_team = getattr(membership, "team", None)

        team_param = request.query_params.get("team_id")
        if team_param:
            # Any member may see any team in the league — but only the roster, and
            # the client is told whose team it is and whether it is the caller's.
            team = get_object_or_404(FantasyTeam, id=team_param, league=league)
        else:
            team = own_team
        if team is None:
            return Response({"detail": "Nessuna squadra associata in questa lega."}, status=status.HTTP_404_NOT_FOUND)
        is_own = own_team is not None and team.id == own_team.id

        # A lineup is referred to a COMPETITION (a set of league fixtures mapped to
        # real matchdays). Default to the first competition; its matchdays come
        # from its fixtures.
        competitions = list(league.competitions.order_by("id").values("id", "name"))
        comp_param = request.query_params.get("competition")
        competition_id: int | None = None
        if comp_param:
            try:
                competition_id = int(comp_param)
            except (TypeError, ValueError):
                competition_id = None
        if competition_id is None and competitions:
            competition_id = competitions[0]["id"]

        matchdays: list[int] = []
        if competition_id is not None:
            matchdays = sorted(
                {
                    int(md)
                    for md in FantasyFixture.objects.filter(
                        competition_id=competition_id, fantasy_matchday__isnull=False
                    ).values_list("fantasy_matchday__real_matchday", flat=True)
                    if md is not None
                }
            )
        if not matchdays:
            matchdays = list(
                FantasyMatchday.objects.filter(league=league).order_by("real_matchday").values_list("real_matchday", flat=True)
            )
        # When a matchday is given we treat it as an "as-of" cutoff: profiles use
        # only matches BEFORE it, so setting a lineup mid-season never peeks at
        # future results (no leakage). No matchday -> full-season profiles.
        matchday_param = request.query_params.get("matchday")
        as_of: int | None = None
        if matchday_param is not None:
            try:
                matchday = int(matchday_param)
            except (TypeError, ValueError):
                matchday = matchdays[0] if matchdays else 1
            as_of = matchday
        else:
            matchday = matchdays[0] if matchdays else 1

        slots = list(
            FantasyRosterSlot.objects.filter(team=team, released_at__isnull=True).select_related("player")
        )
        player_ids = [s.player_id for s in slots]
        cal = _vector_calibration()

        # Which season should the playing-time stats describe? The reference season
        # while it is under way, but BEFORE it starts it has no games at all — and
        # reporting "poco impiegato" for everybody because nobody has played yet is
        # simply wrong. In that case fall back to the last season with data.
        ref_cs = league.reference_season
        stats_cs, stats_as_of = None, as_of
        if ref_cs is not None:
            played_here = MatchAppearance.objects.filter(
                player_id__in=player_ids,
                match__competition_season=ref_cs,
                **({"match__matchday__lt": as_of} if as_of is not None else {}),
            ).exists()
            if played_here:
                stats_cs = ref_cs
            else:
                prev = previous_season_with_data(ref_cs)
                stats_cs, stats_as_of = (prev, None) if prev else (ref_cs, as_of)

        total_matches = (
            Match.objects.filter(competition_season=stats_cs)
            .values("matchday").distinct().count()
            if stats_cs is not None
            else Match.objects.values("matchday").distinct().count()
        )
        profiles = player_profiles(
            player_ids,
            total_matches=total_matches,
            as_of_matchday=stats_as_of,
            params=cal.get("params", {}),
            scales=cal.get("feature_scales", {}),
            competition_season_id=stats_cs.id if stats_cs is not None else None,
        )

        # Average voto (measured, or estimated from the market): a number a manager
        # can actually read. The zone-duel "form" stays for aura, where it belongs.
        values_by_player: dict[int, dict] = {}
        if ref_cs is not None:
            # Calibrate on the WHOLE pool, not just this roster: 25 players are too
            # few an overlap to fit the market->voto relation, which would leave
            # every newcomer without a value.
            values_by_player, _pcs, _fit = player_values(
                ref_cs, latest_market_values(eligible_player_ids(ref_cs.id)))

        # The REAL fixture each player's club plays on this matchday — far more useful
        # than a zone map when picking a lineup (who plays, against whom, and when).
        next_match_by_player: dict[int, dict] = {}
        if ref_cs is not None and as_of is not None:
            fixtures = matchday_fixtures_by_team(ref_cs.id, as_of)
            stints = dict(PlayerTeamStint.objects
                          .filter(player_id__in=player_ids,
                                  team_season__competition_season=ref_cs,
                                  end_date__isnull=True)
                          .values_list("player_id", "team_season_id"))
            for pid, ts_id in stints.items():
                m = fixtures.get(ts_id)
                if m is None:
                    continue
                at_home = m.home_team_id == ts_id
                opp = (m.away_team if at_home else m.home_team).team
                own = (m.home_team if at_home else m.away_team).team
                next_match_by_player[pid] = {
                    "team": own.short_name or own.name,
                    "opponent": opp.short_name or opp.name,
                    "home": at_home,
                    "kickoff": m.kickoff.isoformat() if m.kickoff else None,
                    "kickoff_provisional": m.kickoff_provisional,
                    "status": m.status,
                }

        # In CLASSIC mode the role that governs the formation is the FROZEN listone
        # role (LeaguePlayerRole), not the spatially-inferred one — classic fantacalcio
        # pins roles at season start. Fall back to the player's global seed, then to
        # the spatial guess, so a roster is never roleless.
        is_classic = league.mode == FantasyLeague.MODE_CLASSIC
        frozen_roles: dict[int, str] = {}
        if is_classic:
            frozen_roles = {
                lpr.player_id: _CLASSIC_ROLE_TO_LINEUP.get(lpr.role, "MID")
                for lpr in LeaguePlayerRole.objects.filter(league=league, player_id__in=player_ids)
            }
            seed_roles = dict(
                Player.objects.filter(id__in=player_ids).exclude(classic_role_seed="").values_list("id", "classic_role_seed")
            )

        # Real club each player belongs to, in the season the stats come from — so a
        # player row can name his team, not just his fantasy price.
        real_team = dict(PlayerTeamStint.objects
                         .filter(player_id__in=player_ids,
                                 team_season__competition_season=stats_cs,
                                 end_date__isnull=True)
                         .values_list("player_id", "team_season__team__name")) \
            if stats_cs is not None else {}

        # The deadline, as it applies to THIS matchday. Sent whole rather than left
        # to be re-derived client-side from the kickoffs: the page has to grey out
        # the right players, and "which ones are already playing" is a question only
        # the league's own lock mode can answer.
        locked_pids = matchday_state.locked_players(league, matchday, player_ids)
        lock_closes_at = None
        if league.enforce_lineup_deadline and league.reference_season_id is not None:
            lock_closes_at = (
                matchday_state.matchday_last_kickoffs(league.reference_season_id).get(matchday)
                if league.lineup_lock_mode == FantasyLeague.LOCK_PLAYER
                else matchday_state.lineup_lock_at(league.reference_season_id, matchday)
            )
        lineup_lock = {
            "mode": league.lineup_lock_mode,
            "enforced": bool(league.enforce_lineup_deadline),
            "closes_at": lock_closes_at.isoformat() if lock_closes_at else None,
            "closed": bool(lock_closes_at is not None and lock_closes_at <= timezone.now()),
            "locked_player_ids": sorted(locked_pids),
        }

        roster = []
        for s in slots:
            p = profiles.get(s.player_id, {})
            role = p.get("role", "MID")
            if is_classic:
                role = frozen_roles.get(s.player_id) or _CLASSIC_ROLE_TO_LINEUP.get(
                    seed_roles.get(s.player_id, ""), role
                )
            roster.append(
                {
                    "player_id": s.player_id,
                    "name": s.player.short_name or s.player.full_name,
                    "price": s.purchase_price,
                    "role": role,
                    "avg_col": p.get("avg_col", 0.0),
                    "footprint": p.get("footprint", {}),
                    "appearances": p.get("appearances", 0),
                    "starts": p.get("starts", 0),
                    "avg_minutes": p.get("avg_minutes", 0.0),
                    "minutes_label": p.get("minutes_label", "unknown"),
                    # Le ultime giornate: sono la base del tag, e servono a schermo
                    # per mostrarla invece di chiedere fiducia.
                    "recent_appearances": p.get("recent_appearances", 0),
                    "recent_avg_minutes": p.get("recent_avg_minutes", 0.0),
                    "recent_window": p.get("recent_window", 0),
                    "real_team": real_team.get(s.player_id),
                    "form": p.get("form", 0.0),
                    "stats_season": str(stats_cs) if stats_cs is not None else None,
                    "next_match": next_match_by_player.get(s.player_id),
                    "value": (values_by_player.get(s.player_id) or {}).get("estimated_value"),
                    "value_basis": (values_by_player.get(s.player_id) or {}).get("basis"),
                    # Frozen where he stands: his club is already playing. Always
                    # False under the matchday-wide deadline, where the lineup locks
                    # as a block and no single player is the reason.
                    "locked": s.player_id in locked_pids,
                }
            )
        roster.sort(key=lambda r: (r["avg_col"], -r["price"]))

        lineup_key = f"team{team.id}" + (f":comp{competition_id}" if competition_id is not None else "")
        snap = SavedLineupSnapshot.objects.filter(
            league_id=str(league_id), matchday_id=str(matchday), lineup_id=lineup_key
        ).first()
        saved_lineup = (
            {
                "gk_player_id": int(snap.gk_player_id) if snap.gk_player_id else None,
                "starter_player_ids": snap.starter_player_ids,
                "bench_player_ids": snap.bench_player_ids,
                "starter_backups": snap.starter_backups,
            }
            if snap
            else None
        )

        return Response(
            {
                "team": {"team_id": team.id, "name": team.name,
                         "crest": team.crest,
                         "manager": team.manager.user.username,
                         # The account behind the team, so the page can ask for
                         # his albo d'oro — which belongs to the MANAGER and not
                         # to the team he happens to field in this league.
                         "manager_user_id": team.manager.user_id},
                "is_own": is_own,
                "competitions": [{"competition_id": c["id"], "name": c["name"]} for c in competitions],
                "competition": competition_id,
                "matchdays": matchdays,
                "matchday": matchday,
                "as_of_matchday": as_of,
                "prior_matches": (as_of - 1) if as_of is not None else total_matches,
                "zone_grid": {"cols": 5, "rows": 4, "zone_keys": _zone_grid_keys()},
                "rules": {
                    "starters": 11,
                    "gk_separate_slot": True,
                    "mode": league.mode,
                    # classic role constraints (also used client-side to validate);
                    # null in aura where any shape is legal.
                    "classic_constraints": CLASSIC_CONSTRAINTS if is_classic else None,
                },
                "mode": league.mode,
                "roster": roster,
                # Spending summary: a fixed 500 budget (as used elsewhere), what
                # this squad cost, and per-role breakdown — so the manager reads
                # where his money went without adding it up by hand.
                "budget": _roster_budget(roster, league.initial_budget),
                # Which season the appearances/minutes/label describe. The client
                # must say so: pre-season these are LAST year's, and a silent
                # "poco impiegato" from stale data is exactly the confusion to avoid.
                "stats_season": str(stats_cs) if stats_cs is not None else None,
                "stats_is_reference": bool(stats_cs is not None
                                           and ref_cs is not None
                                           and stats_cs.id == ref_cs.id),
                "saved_lineup": saved_lineup if is_own else None,
                "lineup_lock": lineup_lock,
            }
        )


def _roster_budget(roster: list, initial: int) -> dict:
    """Spending summary against the LEAGUE's budget.

    Was hardcoded to 500 while FantasyLeague.initial_budget defaults to 1000 and
    is configurable per league, so the Squadra page reported a residue of zero to
    anyone who had spent more than 500 — which, on the default economy, is most
    of a full roster.
    """
    spent = sum(r["price"] for r in roster)
    by_role: dict[str, int] = {}
    for r in roster:
        by_role[r["role"]] = by_role.get(r["role"], 0) + r["price"]
    return {"initial": initial, "spent": spent, "remaining": max(0, initial - spent),
            "by_role": by_role}


class LeagueTeamLineupSaveView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        membership = _membership_or_404(league, request.user.id)
        team = getattr(membership, "team", None)
        if team is None:
            return Response({"detail": "Nessuna squadra associata in questa lega."}, status=status.HTTP_400_BAD_REQUEST)

        matchday = request.data.get("matchday")
        if matchday is None:
            return Response({"detail": "matchday richiesto."}, status=status.HTTP_400_BAD_REQUEST)
        gk = request.data.get("gk_player_id")

        # The deadline. WHICH deadline is the league's own choice (see
        # FantasyLeague.lineup_lock_mode): the whole XI at the round's first kickoff,
        # or each player at his own. Both answers are checked here and not only in
        # the UI, because a hand-crafted request is exactly the way one would set a
        # lineup with the results already in hand.
        try:
            md_int = int(matchday)
        except (TypeError, ValueError):
            return Response({"detail": "matchday non valido."}, status=status.HTTP_400_BAD_REQUEST)
        per_player = (league.enforce_lineup_deadline
                      and league.lineup_lock_mode == FantasyLeague.LOCK_PLAYER)
        if league.enforce_lineup_deadline and league.reference_season_id is not None:
            if per_player:
                # The round only closes when the LAST club has kicked off; until then
                # it is partly open and the per-player check below does the work.
                last_kick = matchday_state.matchday_last_kickoffs(
                    league.reference_season_id).get(md_int)
                closed = last_kick is not None and last_kick <= timezone.now()
            else:
                first_kick = matchday_state.lineup_lock_at(league.reference_season_id, md_int)
                closed = first_kick is not None and first_kick <= timezone.now()
            if closed:
                return Response(
                    {"detail": f"Formazione bloccata: la giornata {md_int} è già iniziata."},
                    status=status.HTTP_409_CONFLICT,
                )

        # Classic mode: enforce the role constraints server-side using the FROZEN
        # listone roles, so a hand-crafted request can't bypass the client validator.
        outfield_ids = [int(x) for x in request.data.get("starter_player_ids", []) if x is not None]
        role_of: dict[int, str] = {}
        if league.mode == FantasyLeague.MODE_CLASSIC:
            starter_ids = ([int(gk)] if gk else []) + outfield_ids
            frozen = {
                lpr.player_id: _CLASSIC_ROLE_TO_LINEUP.get(lpr.role, "MID")
                for lpr in LeaguePlayerRole.objects.filter(league=league, player_id__in=starter_ids)
            }
            seed = dict(
                Player.objects.filter(id__in=starter_ids).exclude(classic_role_seed="").values_list("id", "classic_role_seed")
            )
            role_of = {
                pid: frozen.get(pid) or _CLASSIC_ROLE_TO_LINEUP.get(seed.get(pid, ""), "MID")
                for pid in starter_ids
            }
            errors = validate_classic_lineup([role_of[pid] for pid in starter_ids])
            if errors:
                return Response(
                    {"detail": "Formazione non valida (classic).", "errors": errors},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # The lineup is referred to a competition; optionally apply it to all the
        # league's competitions at once (same matchday) to streamline.
        if request.data.get("all_competitions"):
            target_comp_ids: list[int | None] = list(league.competitions.values_list("id", flat=True)) or [None]
        else:
            comp = request.data.get("competition")
            target_comp_ids = [int(comp)] if comp else [None]

        defaults = {
            "gk_player_id": str(gk) if gk else None,
            "starter_player_ids": request.data.get("starter_player_ids", []),
            "bench_player_ids": request.data.get("bench_player_ids", []),
            "starter_backups": request.data.get("starter_backups", []),
        }
        keys = [f"team{team.id}" + (f":comp{cid}" if cid is not None else "")
                for cid in target_comp_ids]

        # Each competition keeps its OWN lineup, so each is judged — and normalised —
        # against its own previous one: the same submission can be legal for the cup,
        # where the manager had never saved anything, and illegal for the league,
        # where he is moving a player who is already playing.
        previous = {
            s.lineup_id: {"gk_player_id": s.gk_player_id,
                          "starter_player_ids": s.starter_player_ids,
                          "bench_player_ids": s.bench_player_ids}
            for s in SavedLineupSnapshot.objects.filter(
                league_id=str(league_id), matchday_id=str(matchday), lineup_id__in=keys)
        }
        locked_ids: set[int] = set()

        if per_player:
            touched = set(lineup_deadline.placement(defaults))
            for prev in previous.values():
                touched |= set(lineup_deadline.placement(prev))
            locked_ids = matchday_state.locked_players(league, md_int, touched)
            if locked_ids:
                # short_name is often empty in the provider's data — the same
                # fallback the roster is serialised with, or the refusal names a
                # database id at the manager.
                names = {
                    pid: (short or full or f"giocatore {pid}")
                    for pid, short, full in Player.objects.filter(id__in=locked_ids)
                    .values_list("id", "short_name", "full_name")
                }
                errors: list[str] = []
                for key in keys:
                    for msg in lineup_deadline.violations(
                            previous.get(key), defaults, locked_ids, names):
                        if msg not in errors:
                            errors.append(msg)
                if errors:
                    return Response(
                        {"detail": "Formazione bloccata per i giocatori la cui partita è iniziata.",
                         "errors": errors},
                        status=status.HTTP_409_CONFLICT,
                    )

        for cid, key in zip(target_comp_ids, keys):
            payload = dict(defaults)
            if role_of:
                # The XI order is not the manager's — the page groups the eleven by
                # role and never offers a way to reorder them — so it is DERIVED
                # here rather than taken on trust: P-D-C-A, with anyone already
                # playing kept at his number inside his own role. Storing what the
                # clicks happened to produce is what put a promoted substitute at
                # the end of the list while the page showed him among his own.
                payload["starter_player_ids"] = lineup_deadline.normalise_xi(
                    outfield_ids, role_of,
                    (previous.get(key) or {}).get("starter_player_ids"),
                    locked_ids,
                )
            SavedLineupSnapshot.objects.update_or_create(
                league_id=str(league_id),
                matchday_id=str(matchday),
                lineup_id=key,
                defaults=payload,
            )
        return Response({"ok": True, "saved_competitions": len([c for c in target_comp_ids if c is not None]) or 1})


# -- Real reference-championship calendar & results ---------------------------


class LeagueRealFixturesView(APIView):
    """Calendar + results of the league's REAL reference championship (e.g. Serie
    A), grouped by matchday. Read model over the Match rows the calendar-sync
    keeps fresh. Optional ?matchday=N to fetch a single round."""

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _membership_or_404(league, request.user.id)
        cs = league.reference_season
        if cs is None:
            return Response({"season": None, "matchdays": []})

        qs = (Match.objects.filter(competition_season=cs)
              .select_related("home_team__team", "away_team__team")
              .annotate(_apps=Count("appearances"))
              .order_by("matchday", "kickoff", "id"))
        md_param = request.query_params.get("matchday")
        if md_param:
            try:
                qs = qs.filter(matchday=int(md_param))
            except ValueError:
                return Response({"detail": "matchday must be an integer"},
                                status=status.HTTP_400_BAD_REQUEST)

        # SofaScore keeps a postponed fixture AND its rescheduled replay as TWO
        # events with different ids. Hide a postponed row once a non-postponed
        # sibling (same teams, i.e. same leg) exists — but keep a genuinely
        # postponed-and-not-yet-replayed match visible.
        matches = list(qs)
        played_legs = {(m.home_team_id, m.away_team_id)
                       for m in matches if m.status != Match.STATUS_POSTPONED}
        matches = [m for m in matches
                   if not (m.status == Match.STATUS_POSTPONED
                           and (m.home_team_id, m.away_team_id) in played_legs)]

        groups: dict = {}
        for m in matches:
            # Appearances, not the final whistle: the pagella already rates a
            # match in progress (whoever is on the pitch is judged on what he has
            # done so far, see classic_pagella.match_in_progress), so requiring
            # FINISHED here withheld the one detail worth opening while it is
            # being played. The votes are provisional and say so; no votes at all
            # is what a 404 would be.
            has_detail = m._apps > 0
            item = {
                "id": m.id,
                "matchday": m.matchday,
                "kickoff": m.kickoff.isoformat() if m.kickoff else None,
                "kickoff_provisional": m.kickoff_provisional,
                "status": m.status,
                "home_team": m.home_team.team.name,
                "away_team": m.away_team.team.name,
                "home_short": m.home_team.team.short_name or m.home_team.team.name,
                "away_short": m.away_team.team.short_name or m.away_team.team.name,
                "home_goals": m.home_goals,
                "away_goals": m.away_goals,
                "has_detail": has_detail,
            }
            groups.setdefault(m.matchday, []).append(item)

        matchdays = [{"matchday": md, "fixtures": fx}
                     for md, fx in sorted(groups.items(),
                                          key=lambda kv: (kv[0] is None, kv[0]))]
        # A rough "current matchday": the earliest with any non-finished fixture,
        # else the last one — lets the frontend open on the live round.
        current = None
        for g in matchdays:
            if any(f["status"] != Match.STATUS_FINISHED for f in g["fixtures"]):
                current = g["matchday"]
                break
        if current is None and matchdays:
            current = matchdays[-1]["matchday"]

        return Response({
            "season": {"id": cs.id, "name": str(cs),
                       "competition": cs.competition.name},
            "current_matchday": current,
            "matchdays": matchdays,
        })


class LeagueRealMatchDetailView(APIView):
    """Vote-relevant detail of a single REAL match: the per-player pagella
    (voto puro + bonus/malus = fantavoto) for both squads, shaped as a classic
    fixture detail so the frontend ClassicMatchDetail renders it. (Aura zone
    breakdown enrichment is a planned follow-up.)"""

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, league_id: int, match_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _membership_or_404(league, request.user.id)
        cs = league.reference_season
        match = get_object_or_404(
            Match.objects.select_related("home_team__team", "away_team__team",
                                         "competition_season"),
            id=match_id)
        if cs is not None and match.competition_season_id != cs.id:
            raise Http404("Match is not in this league's reference season")
        apps = list(MatchAppearance.objects.filter(match=match))
        if not apps:
            return Response({"detail": "Nessun dato disponibile per questa partita."},
                            status=status.HTTP_404_NOT_FOUND)

        pag = pagella_for_match(match, get_reference(match.competition_season_id),
                                league=league)
        hg, ag = int(match.home_goals or 0), int(match.away_goals or 0)
        result = "home" if hg > ag else "away" if ag > hg else "draw"
        # Two different kinds of "not settled", and collapsing them lies either way.
        #
        # IN PROGRESS: the ball is rolling. Every line can still move, and someone
        # who has not played yet may still come on — so each line is marked too.
        #
        # OVER but not confirmed: nobody else is coming on and nobody's afternoon
        # will be re-judged, but the provider can still move a number by a tenth
        # until the +1h confirmation sets data_ready. The totals say so; the lines
        # do not, because "may shift by a tenth" is not the same claim as "this
        # player might still play".
        live = match_in_progress(match)
        provisional = (not match.data_ready
                       and match.status in (Match.STATUS_LIVE, Match.STATUS_FINISHED))
        for side in ("home", "away"):
            pag[side]["provisional"] = provisional
            if live:
                for line in pag[side].get("starters", []) + pag[side].get("bench", []):
                    line["provisional"] = True
        return Response({
            "live": live,
            "provisional": provisional,
            # The clock, only while it is running: on a match that is over the
            # number would be the final whistle dressed up as news.
            "minute": elapsed_minutes(apps) if live else None,
            "mode": "classic",
            "fixture_id": match.id,
            "fantasy_round": match.matchday,
            "real_matchday": match.matchday,
            "stage": None,
            "home_team": match.home_team.team.name,
            "away_team": match.away_team.team.name,
            "home_goals": hg,
            "away_goals": ag,
            "home_total": pag["home"]["total"],
            "away_total": pag["away"]["total"],
            "defense_bonus_mode": None,
            "result": result,
            "home": pag["home"],
            "away": pag["away"],
        })


class LeagueChampionshipPlayersView(APIView):
    """Full player pool of the league's reference championship (the 'listone').

    One row per currently-eligible player (open real-club stint), with role, real
    club, ownership in THIS league (free agent vs owned + owner), and a value
    signal (average voto puro from the latest season with data). The frontend does
    role / free-agent / search filtering and value sorting over this list."""

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, league_id: int):
        league = get_object_or_404(FantasyLeague, id=league_id)
        _membership_or_404(league, request.user.id)
        cs = league.reference_season
        if cs is None:
            return Response({"value_season": None, "players": []})

        pool = eligible_player_ids(cs.id)

        # real club per player (open stint on this season)
        team_by_player = {}
        for pid, tname in (PlayerTeamStint.objects
                           .filter(team_season__competition_season=cs,
                                   end_date__isnull=True, player_id__in=pool)
                           .select_related("team_season__team")
                           .values_list("player_id", "team_season__team__name")):
            team_by_player[pid] = tname

        # frozen listone role, fallback to the global classic role
        lpr = dict(LeaguePlayerRole.objects.filter(league=league)
                   .values_list("player_id", "role"))
        # ownership in this league
        owner_by_player = dict(
            FantasyRosterSlot.objects
            .filter(team__league=league, released_at__isnull=True)
            .values_list("player_id", "team__name"))

        # Value blends last season's average with current-season form as the
        # championship progresses (see player_values).
        market = latest_market_values(pool)
        values, prev_cs, fit = player_values(cs, market)

        players = (Player.objects.filter(id__in=pool)
                   .values("id", "full_name", "short_name", "classic_role_seed"))
        # Players whose role is still an open question: shown, but marked, so
        # nobody plans an auction around someone they cannot actually buy.
        undecided = undecided_player_ids(league)
        rows = []
        for p in players:
            pid = p["id"]
            v = values.get(pid)
            rows.append({
                "market_value": market.get(pid),
                "player_id": pid,
                "name": p["short_name"] or p["full_name"] or str(pid),
                # The list shows the short name ("L. Martinez"); searching for
                # "Lautaro" found nothing because that string is all the client
                # had. Sent alongside so the search can match either.
                "full_name": p["full_name"] or "",
                "role": lpr.get(pid) or p["classic_role_seed"] or "",
                "team": team_by_player.get(pid),
                "owned": pid in owner_by_player,
                "owner": owner_by_player.get(pid),
                "role_undecided": pid in undecided,
                "value": v["value"] if v else None,
                "estimated_value": v["estimated_value"] if v else None,
                "value_basis": v["basis"] if v else None,
                "appearances": v["n_cur"] if v else 0,
                "prev_appearances": v["n_prev"] if v else 0,
            })
        # Default order = the HOMOGENEOUS estimated value, so newcomers rank among
        # the rated players instead of forming an alphabetical tail. The frontend
        # also offers the measured-voto-then-market order.
        rows.sort(key=lambda x: (x["estimated_value"] is None,
                                 -(x["estimated_value"] or 0),
                                 -(x["market_value"] or 0), x["name"]))
        return Response({
            "value_season": str(prev_cs) if prev_cs else None,
            "current_season": str(cs),
            "count": len(rows),
            # How the market->voto estimate was calibrated (r = fit quality on the
            # players having both signals), so the UI can be honest about it.
            "value_fit": ({"intercept": round(fit[0], 3), "slope": round(fit[1], 3),
                           "r": round(fit[2], 3), "n": fit[3]} if fit else None),
            "players": rows,
        })
